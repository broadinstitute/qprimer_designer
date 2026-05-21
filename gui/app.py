"""Streamlit GUI for the qPrimer Designer pipeline."""

import json
import os
import re
import select
import shutil
import signal
import subprocess
import sys
import time
from datetime import date, datetime
from pathlib import Path

import streamlit as st


def _sanitize_filename_component(name: str) -> str:
    """Return a safe single path component for output filenames."""
    sanitized = re.sub(r"[^\w\-.]", "_", (name or ""))
    sanitized = re.sub(r"_+", "_", sanitized).strip("._-")
    if sanitized in {"", ".", ".."}:
        return ""
    return sanitized


def _safe_path_under(base_dir: Path, filename: str) -> Path | None:
    """Build a safe file path under base_dir. Returns None if unsafe."""
    safe_name = _sanitize_filename_component(filename)
    if not safe_name:
        return None
    base_real = os.path.realpath(base_dir)
    base_prefix = base_real if base_real.endswith(os.sep) else base_real + os.sep
    candidate = os.path.realpath(os.path.join(base_real, safe_name))
    if not candidate.startswith(base_prefix):
        return None
    return Path(candidate)


def _build_fetch_command(
    gget_bin: str, taxid: str, out_dir: str, params: dict,
) -> list[str] | None:
    """Build and validate a gget virus command. Returns None if taxid is invalid."""
    safe_taxid = str(taxid).strip()
    if not re.fullmatch(r"\d+", safe_taxid):
        return None

    # Normalize out_dir to break taint chain for CodeQL
    real_out = os.path.realpath(out_dir)

    _ALLOWED_NUC = {"complete", "partial", ""}
    _DATE_RE = re.compile(r"\d{4}-\d{2}-\d{2}")
    _SAFE_STR = re.compile(r"[\w\s.,\-]+")

    cmd = [gget_bin, "virus", safe_taxid, "--out", real_out]

    nuc = params.get("nuc_completeness", "complete")
    if nuc and nuc in _ALLOWED_NUC:
        cmd.extend(["--nuc_completeness", nuc])
    seg = str(params.get("segment", "")).strip()
    if seg and _SAFE_STR.fullmatch(seg):
        cmd.extend(["--segment", seg])
    for key in ("min_seq_length", "max_seq_length"):
        val = params.get(key)
        if val:
            cmd.extend([f"--{key}", str(int(val))])
    for key in ("min_release_date", "max_release_date",
                "min_collection_date", "max_collection_date"):
        val = params.get(key)
        if val and _DATE_RE.fullmatch(str(val)):
            cmd.extend([f"--{key}", str(val)])
    for key, flag in (("geo_location", "geographic_location"), ("host", "host")):
        val = str(params.get(key, "")).strip()
        if val and _SAFE_STR.fullmatch(val):
            cmd.extend([f"--{flag}", val])
    if params.get("refseq_only"):
        cmd.append("--refseq_only")
    max_amb = params.get("max_ambiguous_chars")
    if max_amb is not None:
        cmd.extend(["--max_ambiguous_chars", str(int(max_amb))])

    return cmd

# ---------------------------------------------------------------------------
# Resolve project root (parent of gui/)
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
FASTA_DIR = PROJECT_ROOT / "target_seqs" / "original"
RUNS_DIR = PROJECT_ROOT / "runs"
SCHEMATIC_PATH = PROJECT_ROOT / "schematic.png"

# Ensure the upload directory exists
FASTA_DIR.mkdir(parents=True, exist_ok=True)

# Add src/ to path so we can import qprimer_designer utilities
sys.path.insert(0, str(PROJECT_ROOT / "src"))
sys.path.insert(0, str(PROJECT_ROOT))

from gui.snakefile_builder import build_params_txt, build_snakefile
from qprimer_designer.utils.params import parse_params
from qprimer_designer.adapt_cli import (
    _extract_spreadsheet_id,
    _download_spreadsheet_csv,
    _load_spreadsheet,
    _make_target_name,
    _build_gget_command,
    _extract_accessions,
    _filter_fasta_by_accessions,
    _deduplicate_fasta,
    _build_pset_fa,
    _send_email,
    _read_excel_summary,
    _get_new_seq_table,
    _generate_snakefile,
    _is_empty,
    _install_cron,
    _uninstall_cron,
)

MONITOR_DIR = PROJECT_ROOT / "monitor"
MONITOR_SCHEDULE_PATH = MONITOR_DIR / "schedule.json"
VIRUS_MAP_DATA_DIR = Path(__file__).parent / "virus_map_data"

# ---------------------------------------------------------------------------
# Page config
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="qPrimer Designer",
    layout="wide",
)

# ---------------------------------------------------------------------------
# Navigation helpers
# ---------------------------------------------------------------------------

def _navigate(page: str):
    """Set the current page in session state."""
    st.session_state.page = page
    st.session_state["_need_scroll_top"] = True


def _clear_pipeline_state():
    """Clear all pipeline run state for a fresh workflow."""
    for key in ("pipeline_return_code", "pipeline_log", "pipeline_running",
                "pipeline_completed_rules", "pipeline_should_start",
                "_run_page_fresh"):
        st.session_state.pop(key, None)


def _reset_workflow_state():
    """Reset all workflow selections (targets, config, pipeline state)."""
    for key in ("workflow", "targets", "cross", "host",
                "target_select", "cross_select", "host_select",
                "eval_method", "eval_for", "eval_rev", "eval_pro",
                "_eval_for_saved", "_eval_rev_saved", "_eval_pro_saved",
                "eval_pset_path", "eval_pset_upload",
                "mode", "design_mode", "probe_enabled", "_probe_enabled_saved",
                "run_id", "design_run_id",
                "monitor_spreadsheet_url", "monitor_spreadsheet_data",
                "monitor_query_ids", "monitor_email_recipients",
                "monitor_manual_primers", "monitor_pset_path",
                "monitor_pset_upload",
                "monitor_primer_name", "monitor_primer_fwd",
                "monitor_primer_rev", "monitor_primer_pro",
                "monitor_frequency", "monitor_skip_fetch",
                "monitor_resume_run"):
        st.session_state.pop(key, None)
    _clear_pipeline_state()


def _save_fasta_bytes(path: Path, data: bytes) -> None:
    """Write FASTA bytes ensuring the file ends with a newline.

    Some tools (e.g. bowtie2) silently drop the last base of the last
    sequence when a FASTA file does not end with a newline character.
    """
    if data and not data.endswith(b"\n"):
        data += b"\n"
    path.write_bytes(data)


def _pset_has_probe(fasta_path: str) -> bool:
    """Check if a FASTA file contains probe entries (*_pro)."""
    try:
        with open(fasta_path) as f:
            for line in f:
                if line.startswith(">") and line.strip().endswith("_pro"):
                    return True
    except (FileNotFoundError, OSError):
        pass
    return False


def _current_page() -> str:
    return st.session_state.get("page", "home")


# Workflow step definitions: (page_key, label)
_WORKFLOW_STEPS_DESIGN = [
    ("select_target", "Target"),
    ("select_offtarget", "Off-Target"),
    ("config", "Configuration"),
    ("run", "Run"),
    ("results", "Results"),
]

_WORKFLOW_STEPS_EVALUATE = [
    ("select_target", "Target"),
    ("select_offtarget", "Off-Target"),
    ("eval_primer", "Primer Set"),
    ("config", "Configuration"),
    ("run", "Run"),
    ("results", "Results"),
]

_WORKFLOW_STEPS_MONITOR = [
    ("monitor_target", "Target"),
    ("monitor_primer", "Primer Set"),
    ("config", "Configuration"),
    ("monitor_report", "Report"),
    ("run", "Run"),
    ("results", "Results"),
]


def _get_workflow_steps():
    workflow = st.session_state.get("workflow", "design")
    if workflow == "monitor":
        return _WORKFLOW_STEPS_MONITOR
    if workflow == "evaluate":
        return _WORKFLOW_STEPS_EVALUATE
    return _WORKFLOW_STEPS_DESIGN


def _render_workflow_progress(current_step_key: str):
    """Render a horizontal progress stepper at the top of workflow pages."""
    steps = _get_workflow_steps()
    current_idx = next(
        (i for i, (key, _) in enumerate(steps) if key == current_step_key),
        0,
    )

    steps_html = []
    for i, (key, label) in enumerate(steps):
        num = i + 1
        if i < current_idx:
            # Completed step
            circle = (
                f"<span style='display:inline-flex;align-items:center;justify-content:center;"
                f"width:28px;height:28px;border-radius:50%;background:#4CAF50;color:#fff;"
                f"font-size:14px;font-weight:600;'>&#10003;</span>"
            )
            lbl = f"<span style='color:#4CAF50;font-weight:500;'>{label}</span>"
        elif i == current_idx:
            # Current step
            circle = (
                f"<span style='display:inline-flex;align-items:center;justify-content:center;"
                f"width:28px;height:28px;border-radius:50%;background:#1976D2;color:#fff;"
                f"font-size:14px;font-weight:600;'>{num}</span>"
            )
            lbl = f"<span style='color:#1976D2;font-weight:600;'>{label}</span>"
        else:
            # Future step
            circle = (
                f"<span style='display:inline-flex;align-items:center;justify-content:center;"
                f"width:28px;height:28px;border-radius:50%;background:transparent;"
                f"border:2px solid #ccc;color:#ccc;"
                f"font-size:14px;font-weight:600;'>{num}</span>"
            )
            lbl = f"<span style='color:#ccc;'>{label}</span>"

        step = (
            f"<div style='display:flex;flex-direction:column;align-items:center;gap:4px;'>"
            f"{circle}{lbl}</div>"
        )
        steps_html.append(step)

    # Connector line between steps
    connector = (
        "<div style='flex:1;height:2px;background:#ddd;margin:0 4px;align-self:center;"
        "margin-bottom:20px;'></div>"
    )

    inner = connector.join(steps_html)
    st.markdown(
        f"<div style='display:flex;align-items:flex-start;justify-content:center;"
        f"margin:0.5em 2em 1em 2em;'>{inner}</div>",
        unsafe_allow_html=True,
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

# Common virus presets: (display name, TaxID)
VIRUS_PRESETS: list[tuple[str, int]] = [
    # --- Coronaviruses ---
    ("SARS-CoV-2", 2697049),
    ("SARS-CoV", 694009),
    ("MERS-CoV", 1335626),
    ("Human coronavirus 229E", 11137),
    ("Human coronavirus OC43", 31631),
    ("Human coronavirus NL63", 277944),
    ("Human coronavirus HKU1", 290028),
    # --- Influenza ---
    ("Influenza A", 11320),
    ("Influenza B", 11520),
    # --- Respiratory ---
    ("RSV-A (Respiratory Syncytial Virus A)", 208893),
    ("RSV-B (Respiratory Syncytial Virus B)", 208895),
    ("Human metapneumovirus", 162145),
    ("Human parainfluenza virus 1", 12730),
    ("Human parainfluenza virus 3", 11216),
    ("Human bocavirus 1", 329641),
    ("Human rhinovirus A", 147711),
    ("Adenovirus (Human mastadenovirus C)", 129951),
    ("Measles virus", 11234),
    ("Mumps virus", 2560602),
    ("Rubella virus", 11041),
    # --- Hemorrhagic fever ---
    ("Ebola virus (Zaire)", 186538),
    ("Marburg virus", 3052505),
    ("Lassa virus", 3052310),
    ("Crimean-Congo hemorrhagic fever virus", 3052518),
    ("Rift Valley Fever virus", 11588),
    ("Hantaan virus", 3052480),
    ("SFTS virus", 1003835),
    # --- Mosquito/tick-borne ---
    ("Dengue virus 1", 11053),
    ("Dengue virus 2", 11060),
    ("Dengue virus 3", 11069),
    ("Dengue virus 4", 11070),
    ("Zika virus", 64320),
    ("Chikungunya virus", 37124),
    ("Oropouche virus", 118655),
    ("West Nile virus", 11082),
    ("Yellow Fever virus", 11089),
    ("Japanese Encephalitis virus", 11072),
    ("Tick-borne encephalitis virus", 11084),
    # --- Blood-borne / hepatitis ---
    ("HIV-1", 11676),
    ("HIV-2", 11709),
    ("Hepatitis A virus", 12092),
    ("Hepatitis B virus", 10407),
    ("Hepatitis C virus", 3052230),
    ("Hepatitis E virus", 1678143),
    # --- Other ---
    ("MPOX (Monkeypox virus)", 10244),
    ("Nipah virus", 3052225),
    ("Rabies virus", 11292),
    ("Enterovirus A71", 39054),
    ("Enterovirus D68", 42789),
    ("Norovirus GII", 142786),
    ("Rotavirus A", 28875),
]


def _decompress_bz2_fastas(directory: Path) -> None:
    """Auto-decompress any .fa.bz2 files that don't have a corresponding .fa."""
    import bz2 as _bz2
    for p in directory.iterdir():
        if p.name.endswith(".fa.bz2") and p.is_file():
            fa_path = directory / p.name[:-4]  # strip .bz2
            if not fa_path.exists():
                with open(p, "rb") as fin, open(fa_path, "wb") as fout:
                    fout.write(_bz2.decompress(fin.read()))


def _available_fasta() -> list[str]:
    """Return sorted list of FASTA stem names in target_seqs/original/."""
    if not FASTA_DIR.exists():
        return []
    _decompress_bz2_fastas(FASTA_DIR)
    return sorted({
        p.stem for p in FASTA_DIR.iterdir()
        if p.suffix in (".fa", ".fasta", ".fna") and p.is_file()
    })


def _fasta_info(stem: str) -> dict:
    """Return sequence count and file size for a FASTA file."""
    path = FASTA_DIR / f"{stem}.fa"
    if not path.exists():
        # try other extensions
        for ext in (".fasta", ".fna"):
            alt = FASTA_DIR / f"{stem}{ext}"
            if alt.exists():
                path = alt
                break
    if not path.exists():
        return {"seqs": 0, "size": 0}
    with open(path) as f:
        n_seqs = sum(1 for line in f if line.startswith(">"))
    return {"seqs": n_seqs, "size": path.stat().st_size}


def _format_size(nbytes: int) -> str:
    for unit in ("B", "KB", "MB"):
        if nbytes < 1024:
            return f"{nbytes:.0f} {unit}"
        nbytes /= 1024
    return f"{nbytes:.1f} GB"


# Default parameter values (matches params.txt.template)
DEFAULT_PARAMS: dict = {
    "DESIGN_WINDOW": 500,
    "TM_MIN": 55,
    "TM_MAX": 60,
    "GC_MAX": 60,
    "DG_MIN": -6,
    "PRIMER_LEN_MIN": 19,
    "PRIMER_LEN_MAX": 21,
    "QUICK_N_POSITIONS": 100,
    "QUICK_N_VARIANTS": 10,
    "QUICK_COV_MIN": 1.0,
    "QUICK_ACT_MIN": 1.0,
    "PROBE_LEN_MIN": 24,
    "PROBE_LEN_MAX": 26,
    "PROBE_TM_MIN": 65,
    "PROBE_TM_MAX": 75,
    "PROBE_GC_MAX": 60,
    "PROBE_DG_MIN": -8,
    "PROBE_HOMOPOLYMER_MAX": 4,
    "PROBE_AVOID_5PRIME_G": False,
    "PROBE_CONSERVATION_THRESHOLD": 0.8,
    "PROBE_MAX_MISMATCHES": 2,
    "PROBE_MAX_INDELS": 0,
    "PROBE_AMPLICON_BUFFER": 15,
    "MIN_PROBES_PER_PAIR": 1,
    "MAX_PROBES_PER_PAIR": 5,
    "AMPLEN_MIN": 60,
    "AMPLEN_MAX": 200,
    "OFFLEN_MIN": 50,
    "OFFLEN_MAX": 5000,
    "NUM_TOP_SENSITIVITY": 200,
}


def _init_params():
    """Initialize session-state params dict from defaults if not set."""
    if "params" not in st.session_state:
        st.session_state.params = dict(DEFAULT_PARAMS)


def _find_tool(name: str) -> str | None:
    """Find a tool in PATH or the current Python environment's bin directory."""
    found = shutil.which(name)
    if found:
        return found
    candidate = Path(sys.executable).parent / name
    if candidate.exists():
        return str(candidate)
    return None


def _check_tool(name: str) -> bool:
    return _find_tool(name) is not None


# ---------------------------------------------------------------------------
# Sidebar
# ---------------------------------------------------------------------------

def _render_sidebar():
    with st.sidebar:
        st.title("qPrimer Designer")
        st.caption(f"Working directory: `{PROJECT_ROOT}`")

        st.divider()

        page = _current_page()

        # Reduce vertical spacing for sidebar nav buttons
        st.markdown(
            "<style>"
            "section[data-testid='stSidebar'] .stElementContainer:has(button[kind='tertiary']), "
            "section[data-testid='stSidebar'] .stMarkdown "
            "{ margin-bottom: -1rem; }"
            "</style>",
            unsafe_allow_html=True,
        )

        # Global nav buttons — always visible
        if st.button(":material/home: Home", key="sidebar_home", type="tertiary"):
            _reset_workflow_state()
            if page == "virus_map":
                for k in ("virus_map_selected_countries", "_virus_map_prev_clicks",
                           "virus_map_country_multiselect", "virus_map_chart",
                           "virus_map_select", "virus_map_date_range"):
                    st.session_state.pop(k, None)
            _navigate("home")
            st.rerun()
        if st.button(":material/public: Global Virus Distribution", key="sidebar_virus_map", type="tertiary"):
            _navigate("virus_map")
            st.rerun()
        if RUNS_DIR.exists() and any(RUNS_DIR.iterdir()):
            if st.button(":material/stacks: Past Results", key="sidebar_results", type="tertiary"):
                _navigate("results")
                st.rerun()
        if st.button(":material/mail: Contact", key="sidebar_contact", type="tertiary"):
            _navigate("contact")
            st.rerun()

        if page not in ("home", "virus_map", "contact", "results"):
            st.divider()

            workflow = st.session_state.get("workflow", "design")
            mode = st.session_state.get("mode", "Singleplex")

            if workflow == "monitor":
                st.markdown("**Mode:** Monitor")
                sheet_data = st.session_state.get("monitor_spreadsheet_data")
                if sheet_data:
                    n_rows = len(sheet_data["data_rows"])
                    st.markdown(f"**Spreadsheet:** {n_rows} row(s)")
            else:
                # Targets
                targets = st.session_state.get("targets", [])
                st.markdown(f"**Target:** {', '.join(targets) if targets else '—'}")

                # Off-targets
                cross = st.session_state.get("cross", [])
                host = st.session_state.get("host", [])
                off_targets = cross + host
                st.markdown(f"**Off-targets:** {', '.join(off_targets) if off_targets else '—'}")

                # Probe
                if "probe_enabled" in st.session_state or "_probe_enabled_saved" in st.session_state:
                    probe = st.session_state.get("probe_enabled",
                                                  st.session_state.get("_probe_enabled_saved", False))
                    st.markdown(f"**Probe:** {'On' if probe else 'Off'}")
                else:
                    st.markdown("**Probe:** —")

                # Mode
                if "design_mode" in st.session_state or "mode" in st.session_state:
                    st.markdown(f"**Mode:** {mode}")
                else:
                    st.markdown("**Mode:** —")

        # --- Scheduled monitor (always visible) ---
        active_cron = _get_active_cron_monitor()
        if active_cron:
            st.divider()
            st.markdown("**Scheduled monitor**")
            # Read saved schedule details
            if MONITOR_SCHEDULE_PATH.exists():
                import json
                try:
                    sched_info = json.loads(MONITOR_SCHEDULE_PATH.read_text())
                    targets = sched_info.get("targets", [])
                    target_summary = ", ".join(targets) if targets else "N/A"
                    with st.expander(target_summary):
                        lines = []
                        primers = sched_info.get("primer_sets", [])
                        if primers:
                            primer_list = "".join(
                                f"<br>&nbsp;&nbsp;&nbsp;{p}" for p in primers
                            )
                            lines.append(f"<b>Primer sets:</b>{primer_list}")
                        lines.append(f"<b>Frequency:</b> {sched_info.get('frequency', 'N/A')}")
                        recipients = sched_info.get("recipients", "")
                        if recipients:
                            lines.append(f"<b>Email:</b> {recipients}")
                        st.markdown(
                            "<div style='font-size:0.85em; line-height:1.4;'>"
                            + "<br>".join(lines)
                            + "</div>",
                            unsafe_allow_html=True,
                        )
                        if st.button("Stop monitoring", key="sidebar_stop_cron", use_container_width=True):
                            _uninstall_cron()
                            if MONITOR_SCHEDULE_PATH.exists():
                                MONITOR_SCHEDULE_PATH.unlink()
                            st.rerun()
                except (json.JSONDecodeError, KeyError):
                    pass
            else:
                if st.button("Stop monitoring", key="sidebar_stop_cron", use_container_width=True):
                    _uninstall_cron()
                    st.rerun()



# ---------------------------------------------------------------------------
# Home page
# ---------------------------------------------------------------------------

def _page_home():
    st.markdown(
        "<h1 style='text-align: center; margin-top: -2rem; margin-bottom: 0;'>qPrimer Designer</h1>"
        "<p style='text-align: center; font-size: 1.2em; color: gray; margin-bottom: 0;'>"
        "ML-guided PCR primer design and evaluation"
        "</p>"
        "<hr style='margin-top: 1rem; margin-bottom: 1rem;'>",
        unsafe_allow_html=True,
    )

    # Action button illustrations (inline SVG)
    _ICON_DESIGN = """
    <svg viewBox="0 0 112 80" width="112" height="80" xmlns="http://www.w3.org/2000/svg">
      <!-- Target sequence -->
      <line x1="5" y1="20" x2="107" y2="20" stroke="#888" stroke-width="3" stroke-linecap="round"/>
      <!-- Unselected pair left -->
      <rect x="4" y="44" width="10" height="3" rx="1" fill="#4CAF50" opacity="0.25"/>
      <polygon points="15,45.5 13,43.5 13,47.5" fill="#4CAF50" opacity="0.25"/>
      <polygon points="18,45.5 20,43.5 20,47.5" fill="#E8451E" opacity="0.25"/>
      <rect x="19" y="44" width="10" height="3" rx="1" fill="#E8451E" opacity="0.25"/>
      <!-- Selected primer pair -->
      <rect x="38" y="41" width="14" height="5" rx="2" fill="#4CAF50"/>
      <polygon points="54,43.5 50,39.5 50,47.5" fill="#4CAF50"/>
      <polygon points="58,43.5 62,39.5 62,47.5" fill="#E8451E"/>
      <rect x="60" y="41" width="14" height="5" rx="2" fill="#E8451E"/>
      <!-- Selection box -->
      <rect x="33" y="36" width="46" height="16" rx="3" fill="none" stroke="#2E86AB" stroke-width="2"/>
      <!-- Unselected pair right -->
      <rect x="84" y="44" width="10" height="3" rx="1" fill="#4CAF50" opacity="0.25"/>
      <polygon points="95,45.5 93,43.5 93,47.5" fill="#4CAF50" opacity="0.25"/>
      <polygon points="98,45.5 100,43.5 100,47.5" fill="#E8451E" opacity="0.25"/>
      <rect x="99" y="44" width="10" height="3" rx="1" fill="#E8451E" opacity="0.25"/>
      <!-- Unselected pairs below -->
      <rect x="22" y="62" width="10" height="3" rx="1" fill="#4CAF50" opacity="0.25"/>
      <polygon points="33,63.5 31,61.5 31,65.5" fill="#4CAF50" opacity="0.25"/>
      <polygon points="36,63.5 38,61.5 38,65.5" fill="#E8451E" opacity="0.25"/>
      <rect x="37" y="62" width="10" height="3" rx="1" fill="#E8451E" opacity="0.25"/>
      <rect x="60" y="62" width="10" height="3" rx="1" fill="#4CAF50" opacity="0.25"/>
      <polygon points="71,63.5 69,61.5 69,65.5" fill="#4CAF50" opacity="0.25"/>
      <polygon points="74,63.5 76,61.5 76,65.5" fill="#E8451E" opacity="0.25"/>
      <rect x="75" y="62" width="10" height="3" rx="1" fill="#E8451E" opacity="0.25"/>
    </svg>
    """

    _ICON_EVALUATE = """
    <svg viewBox="0 0 100 80" width="100" height="80" xmlns="http://www.w3.org/2000/svg">
      <!-- Target sequence line -->
      <line x1="6" y1="20" x2="94" y2="20" stroke="#888" stroke-width="3" stroke-linecap="round"/>
      <!-- Forward primer -->
      <rect x="10" y="28" width="18" height="4" rx="1.5" fill="#4CAF50"/>
      <polygon points="30,30 27,27 27,33" fill="#4CAF50"/>
      <!-- Match/mismatch ticks for forward -->
      <line x1="13" y1="23" x2="13" y2="27" stroke="#4CAF50" stroke-width="1.5"/>
      <line x1="17" y1="23" x2="17" y2="27" stroke="#4CAF50" stroke-width="1.5"/>
      <line x1="21" y1="23" x2="21" y2="27" stroke="#E8451E" stroke-width="1.5"/>
      <line x1="25" y1="23" x2="25" y2="27" stroke="#4CAF50" stroke-width="1.5"/>
      <line x1="29" y1="23" x2="29" y2="27" stroke="#4CAF50" stroke-width="1.5"/>
      <!-- Reverse primer -->
      <polygon points="58,30 61,27 61,33" fill="#E8451E"/>
      <rect x="60" y="28" width="18" height="4" rx="1.5" fill="#E8451E"/>
      <!-- Match/mismatch ticks for reverse -->
      <line x1="61" y1="23" x2="61" y2="27" stroke="#4CAF50" stroke-width="1.5"/>
      <line x1="65" y1="23" x2="65" y2="27" stroke="#4CAF50" stroke-width="1.5"/>
      <line x1="69" y1="23" x2="69" y2="27" stroke="#E8451E" stroke-width="1.5"/>
      <line x1="73" y1="23" x2="73" y2="27" stroke="#4CAF50" stroke-width="1.5"/>
      <line x1="77" y1="23" x2="77" y2="27" stroke="#4CAF50" stroke-width="1.5"/>
      <!-- Magnifying glass -->
      <circle cx="65" cy="57" r="13" fill="none" stroke="#555" stroke-width="2.5"/>
      <line x1="74" y1="66" x2="83" y2="75" stroke="#555" stroke-width="3" stroke-linecap="round"/>
      <!-- Checkmark inside -->
      <polyline points="59,57 63,62 72,52" fill="none" stroke="#4CAF50" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"/>
    </svg>
    """

    _ICON_MONITOR = """
    <svg viewBox="0 0 120 80" width="120" height="80" xmlns="http://www.w3.org/2000/svg">
      <!-- Phylogenetic tree -->
      <line x1="6" y1="40" x2="22" y2="40" stroke="#555" stroke-width="2.5"/>
      <line x1="22" y1="40" x2="22" y2="18" stroke="#555" stroke-width="2.5"/>
      <line x1="22" y1="40" x2="22" y2="62" stroke="#555" stroke-width="2.5"/>
      <!-- Upper branch -->
      <line x1="22" y1="18" x2="50" y2="18" stroke="#555" stroke-width="2.5"/>
      <line x1="50" y1="18" x2="50" y2="10" stroke="#555" stroke-width="2.5"/>
      <line x1="50" y1="18" x2="50" y2="26" stroke="#555" stroke-width="2.5"/>
      <line x1="50" y1="10" x2="78" y2="10" stroke="#555" stroke-width="2.5"/>
      <line x1="50" y1="26" x2="78" y2="26" stroke="#555" stroke-width="2.5"/>
      <!-- Lower branch -->
      <line x1="22" y1="62" x2="50" y2="62" stroke="#555" stroke-width="2.5"/>
      <line x1="50" y1="62" x2="50" y2="54" stroke="#555" stroke-width="2.5"/>
      <line x1="50" y1="62" x2="50" y2="70" stroke="#555" stroke-width="2.5"/>
      <line x1="50" y1="54" x2="78" y2="54" stroke="#555" stroke-width="2.5"/>
      <line x1="50" y1="70" x2="78" y2="70" stroke="#555" stroke-width="2.5"/>
      <!-- Leaf dots -->
      <circle cx="81" cy="10" r="3" fill="#2E86AB"/>
      <circle cx="81" cy="26" r="3" fill="#2E86AB"/>
      <circle cx="81" cy="54" r="3" fill="#2E86AB"/>
      <circle cx="81" cy="70" r="3" fill="#2E86AB"/>
      <!-- Time checkpoint lines (vertical dashed) -->
      <line x1="38" y1="4" x2="38" y2="76" stroke="#4CAF50" stroke-width="1.5" stroke-dasharray="3,3" opacity="0.6"/>
      <line x1="66" y1="4" x2="66" y2="76" stroke="#4CAF50" stroke-width="1.5" stroke-dasharray="3,3" opacity="0.6"/>
      <!-- Checkmarks in circles at checkpoint 1 -->
      <circle cx="38" cy="18" r="5" fill="white" stroke="#4CAF50" stroke-width="1.5"/>
      <polyline points="35,18 37,20 41,15" fill="none" stroke="#4CAF50" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/>
      <circle cx="38" cy="62" r="5" fill="white" stroke="#4CAF50" stroke-width="1.5"/>
      <polyline points="35,62 37,64 41,59" fill="none" stroke="#4CAF50" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/>
      <!-- Checkmarks in circles at checkpoint 2 -->
      <circle cx="66" cy="10" r="5" fill="white" stroke="#4CAF50" stroke-width="1.5"/>
      <polyline points="63,10 65,12 69,7" fill="none" stroke="#4CAF50" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/>
      <circle cx="66" cy="26" r="5" fill="white" stroke="#4CAF50" stroke-width="1.5"/>
      <polyline points="63,26 65,28 69,23" fill="none" stroke="#4CAF50" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/>
      <circle cx="66" cy="54" r="5" fill="white" stroke="#4CAF50" stroke-width="1.5"/>
      <polyline points="63,54 65,56 69,51" fill="none" stroke="#4CAF50" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/>
      <circle cx="66" cy="70" r="5" fill="white" stroke="#E8451E" stroke-width="1.5"/>
      <line x1="63" y1="67" x2="69" y2="73" stroke="#E8451E" stroke-width="1.8" stroke-linecap="round"/>
      <line x1="69" y1="67" x2="63" y2="73" stroke="#E8451E" stroke-width="1.8" stroke-linecap="round"/>
    </svg>
    """

    # Action buttons
    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown(
            f'<div style="text-align: center; padding: 0.3em 0;">{_ICON_DESIGN}</div>',
            unsafe_allow_html=True,
        )
        if st.button(
            "Design",
            use_container_width=True,
            type="primary",
        ):
            _reset_workflow_state()
            st.session_state.workflow = "design"
            _navigate("select_target")
            st.rerun()
        st.caption(
            "Generate optimized primer sets for your target pathogen sequences. "
            "Supports probe design."
        )

    with col2:
        st.markdown(
            f'<div style="text-align: center; padding: 0.3em 0;">{_ICON_EVALUATE}</div>',
            unsafe_allow_html=True,
        )
        if st.button(
            "Evaluate",
            use_container_width=True,
            type="primary",
        ):
            _reset_workflow_state()
            st.session_state.workflow = "evaluate"
            _navigate("select_target")
            st.rerun()
        st.caption(
            "Assess the performance of your existing primers against target sequences."
        )

    with col3:
        st.markdown(
            f'<div style="text-align: center; padding: 0.3em 0;">{_ICON_MONITOR}</div>',
            unsafe_allow_html=True,
        )
        if st.button(
            "Monitor",
            use_container_width=True,
            type="primary",
        ):
            _reset_workflow_state()
            st.session_state.workflow = "monitor"
            _navigate("monitor_target")
            st.rerun()
        st.caption(
            "Track the validity of existing primer designs against newly observed "
            "pathogen sequences."
        )

    st.markdown("<hr style='margin-top: 1rem; margin-bottom: 1rem;'>", unsafe_allow_html=True)

    # Description
    st.markdown(
        """
qPrimer Designer is an AI-powered tool for PCR diagnostics primer design and
evaluation. It enables a **scalable, data-driven approach for adaptive
diagnostic design** against rapidly evolving pathogens — accessible anywhere
in the world, without requiring deep specialist expertise.

The tool is built on **machine learning models** that predict PCR activity for
a given primer pair and target sequence. These models were trained on
experimental data from over **50,000 unique primer pair–target sequence
combinations**, and significantly outperform predictions based on mismatch
counts or free energy alone
([Baek, Hsu et al., NeurIPS 2025 Workshop on AI for Science](https://openreview.net/forum?id=GJaNhMEZGM#discussion)).

Its core capabilities include **Design** — generating new primer sets for
given pathogen sequences — and **Evaluate** — assessing the performance of
existing primer sets. Pathogen sequences can be uploaded directly as FASTA
files or fetched from NCBI Virus by specifying a taxonomy ID and metadata
parameters. In addition to the target pathogen, users can specify
cross-reactivity sequences from other pathogens or host genomes to evaluate
and minimize non-specific amplification. Finally, the **Monitor** feature
tracks the validity of existing designs against newly observed pathogen
sequences by periodically running fetch and evaluate, and delivering results
via email reports.
"""
    )



# ---------------------------------------------------------------------------
# Virus Map
# ---------------------------------------------------------------------------

# Map NCBI country names to ISO-3166 alpha-3 codes for plotly choropleth
_NCBI_TO_ISO3 = {
    "USA": "USA", "United States": "USA", "Puerto Rico": "PRI",
    "U.S. Virgin Islands": "VIR", "Guam": "GUM",
    "Viet Nam": "VNM", "Vietnam": "VNM",
    "Republic of Korea": "KOR", "Korea": "KOR", "South Korea": "KOR",
    "Democratic Republic of the Congo": "COD", "Zaire": "COD",
    "Republic of the Congo": "COG", "Congo": "COG",
    "Czech Republic": "CZE", "Czechia": "CZE",
    "Reunion": "REU", "French Guiana": "GUF",
    "Guadeloupe": "GLP", "Martinique": "MTQ",
    "Chinese Taipei": "TWN", "Taiwan": "TWN",
    "Myanmar": "MMR", "Burma": "MMR",
    "Cote d'Ivoire": "CIV", "Ivory Coast": "CIV",
    "Lao People's Democratic Republic": "LAO", "Laos": "LAO",
    "Russian Federation": "RUS", "Russia": "RUS",
    "Iran": "IRN", "Iran, Islamic Republic of": "IRN",
    "Syria": "SYR", "Syrian Arab Republic": "SYR",
    "Tanzania": "TZA", "United Republic of Tanzania": "TZA",
    "Venezuela": "VEN", "Bolivia": "BOL",
    "Micronesia": "FSM", "Micronesia, Federated States of": "FSM",
    "Palestine": "PSE", "State of Palestine": "PSE",
    "North Macedonia": "MKD", "Macedonia": "MKD",
    "Brunei": "BRN", "Brunei Darussalam": "BRN",
    "Cape Verde": "CPV", "Cabo Verde": "CPV",
    "East Timor": "TLS", "Timor-Leste": "TLS",
    "Micronesia": "FSM",
    "Turkey": "TUR", "Turkiye": "TUR",
    "Saint Barthelemy": "BLM", "Saint Martin": "MAF",
    "Virgin Islands": "VIR",
    "Borneo": "MYS",  # Borneo is split; default to Malaysia
}


def _country_to_iso3(name: str) -> str | None:
    """Convert a country name to ISO-3166 alpha-3 code."""
    if name in _NCBI_TO_ISO3:
        return _NCBI_TO_ISO3[name]
    try:
        import pycountry
        result = pycountry.countries.lookup(name)
        return result.alpha_3
    except (LookupError, ImportError):
        return None


def _download_virus_map_data():
    """Download virus map data from NCBI with progress display."""
    import importlib, sys
    # Ensure gui/ directory is importable regardless of working directory
    _gui_dir = str(Path(__file__).parent)
    if _gui_dir not in sys.path:
        sys.path.insert(0, _gui_dir)
    _mod = importlib.import_module("fetch_virus_map_data")
    DEFAULT_VIRUSES = _mod.DEFAULT_VIRUSES
    CONTINENTS = _mod.CONTINENTS
    fetch_continent = _mod.fetch_continent
    extract_record = _mod.extract_record

    st.subheader("Downloading virus geographic data...")
    st.caption("This is a one-time download from NCBI. It may take a few minutes.")

    VIRUS_MAP_DATA_DIR.mkdir(parents=True, exist_ok=True)

    progress = st.progress(0.0)
    status = st.empty()

    # Filter to viruses that still need downloading
    to_fetch = []
    for name, taxid, *_rest in DEFAULT_VIRUSES:
        out_path = VIRUS_MAP_DATA_DIR / f"{name}.json"
        if not out_path.exists():
            to_fetch.append((name, taxid))

    total = len(DEFAULT_VIRUSES)
    done = total - len(to_fetch)
    progress.progress(done / total if total else 1.0)

    if to_fetch:
        from concurrent.futures import ThreadPoolExecutor, as_completed

        def _fetch_one_virus(name, taxid):
            records = []
            for continent in CONTINENTS:
                raw = fetch_continent(taxid, continent)
                for rec in raw:
                    parsed = extract_record(rec)
                    if parsed:
                        parsed["continent"] = continent
                        records.append(parsed)
            countries = set(r["country"] for r in records)
            data = {
                "name": name,
                "taxid": taxid,
                "total_sequences": len(records),
                "num_countries": len(countries),
                "records": records,
            }
            out_path = VIRUS_MAP_DATA_DIR / f"{name}.json"
            with open(out_path, "w") as f:
                json.dump(data, f)
            return name

        max_workers = min(8, len(to_fetch))
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {
                executor.submit(_fetch_one_virus, name, taxid): name
                for name, taxid in to_fetch
            }
            for future in as_completed(futures):
                done += 1
                name = future.result()
                status.text(f"Fetched {name.replace('_', ' ')} ({done}/{total})")
                progress.progress(done / total)

    # Save index
    index = []
    for p in sorted(VIRUS_MAP_DATA_DIR.glob("*.json")):
        if p.name == "index.json":
            continue
        with open(p) as f:
            d = json.load(f)
        index.append({
            "name": d["name"], "taxid": d["taxid"],
            "total_sequences": d["total_sequences"],
            "num_countries": d["num_countries"],
        })
    with open(VIRUS_MAP_DATA_DIR / "index.json", "w") as f:
        json.dump(index, f, indent=2)

    status.text("Download complete!")
    st.rerun()


def _page_virus_map():
    """Virus Map page: world choropleth showing virus sequence geographic distribution."""
    import plotly.express as px
    import pandas as pd

    if not VIRUS_MAP_DATA_DIR.exists() or not any(
        p.name != "index.json" for p in VIRUS_MAP_DATA_DIR.glob("*.json")
    ):
        _download_virus_map_data()
        return

    # Load available viruses
    available = []
    for p in sorted(VIRUS_MAP_DATA_DIR.glob("*.json")):
        if p.name == "index.json":
            continue
        available.append(p.stem)

    if not available:
        st.warning("No virus data files found.")
        return

    # --- Left panel: controls | Right panel: map ---
    col_left, col_right = st.columns([1, 3])

    with col_left:
        selected = st.selectbox(
            "Virus",
            options=available,
            format_func=lambda x: x.replace("_", " "),
            key="virus_map_select",
        )
        if not selected:
            return

        # Load data
        data_path = VIRUS_MAP_DATA_DIR / f"{selected}.json"
        with open(data_path) as f:
            data = json.load(f)

        records = data.get("records", [])
        has_records = bool(records)

        if not has_records:
            st.warning("No geographic data available for this virus.")

        rec_df = pd.DataFrame(records) if has_records else pd.DataFrame(columns=["country", "collection_date", "release_date"])

        if has_records:
            # Parse dates
            rec_df["date_str"] = rec_df["collection_date"].where(
                rec_df["collection_date"] != "", rec_df["release_date"],
            )
            rec_df["date_parsed"] = pd.to_datetime(
                rec_df["date_str"].str[:10], errors="coerce", format="mixed",
            )

            valid_dates = rec_df["date_parsed"].dropna()
            date_min = date(2010, 1, 1)
            if len(valid_dates) > 0:
                date_max = valid_dates.max().date()
            else:
                date_max = date.today()

            date_today = date.today()
            if date_max < date_today:
                date_max = date_today

            filter_start, filter_end = st.slider(
                "Collection date",
                min_value=date_min,
                max_value=date_max,
                value=(date_min, date_max),
                key="virus_map_date_range",
            )

            # Apply date filter
            mask = pd.Series(True, index=rec_df.index)
            mask &= rec_df["date_parsed"].isna() | (rec_df["date_parsed"].dt.date >= filter_start)
            mask &= rec_df["date_parsed"].isna() | (rec_df["date_parsed"].dt.date <= filter_end)
            filtered = rec_df[mask]

            # Summary metrics (updated after df is built below)
            seq_count = len(filtered)

            # Build country counts
            country_counts = filtered["country"].value_counts()
            rows = []
            unmapped = []
            for country, count in country_counts.items():
                iso3 = _country_to_iso3(country)
                if iso3:
                    rows.append({"country": country, "iso3": iso3, "sequences": int(count)})
                else:
                    unmapped.append(country)
        else:
            filtered = rec_df
            seq_count = data.get("total_sequences", 0)
            rows = []
            unmapped = []

        if rows:
            df = pd.DataFrame(rows)
            df = df.groupby("iso3", as_index=False).agg(
                country=("country", "first"),
                sequences=("sequences", "sum"),
            )
            df = df.sort_values("sequences", ascending=False).reset_index(drop=True)
            display_df = df[["country", "sequences"]].copy()
        else:
            df = None
            display_df = pd.DataFrame(columns=["country", "sequences"])

        # Metrics placeholder — filled after selection is known
        metrics_seq = st.empty()
        metrics_countries = st.empty()

        # Country breakdown
        with st.expander(f"Country breakdown ({len(display_df)})", expanded=False):
            st.dataframe(
                display_df.rename(columns={"country": "Country", "sequences": "Sequences"}),
                hide_index=True,
                use_container_width=True,
            )

    # Initialize accumulated selections in session state
    if "virus_map_selected_countries" not in st.session_state:
        st.session_state["virus_map_selected_countries"] = []

    # Build continent -> country mapping from filtered data
    _continent_countries: dict[str, list[str]] = {}
    if has_records and df is not None:
        for _, row in filtered.iterrows():
            cont = row.get("continent", "")
            country = row.get("country", "")
            if cont and country and country in display_df["country"].values:
                _continent_countries.setdefault(cont, set()).add(country)
        _continent_countries = {k: sorted(v) for k, v in _continent_countries.items()}

    with col_right:
        if df is not None and not df.empty:
            # Highlight already-selected countries
            picked = set(st.session_state["virus_map_selected_countries"])

            fig = px.choropleth(
                df,
                locations="iso3",
                locationmode="ISO-3",
                color="sequences",
                hover_name="country",
                hover_data={"iso3": False, "sequences": True},
                color_continuous_scale="YlOrRd",
                title="",
                labels={"sequences": "Sequences"},
            )

            # Highlight selected countries with thicker borders on the same trace
            if picked:
                line_widths = [3 if c in picked else 0.5 for c in df["country"]]
                line_colors = ["#2E86AB" if c in picked else "#999" for c in df["country"]]
                fig.update_traces(
                    marker_line_width=line_widths,
                    marker_line_color=line_colors,
                )

            fig.update_layout(
                geo=dict(
                    showframe=False,
                    showcoastlines=True,
                    projection_type="natural earth",
                ),
                margin=dict(l=0, r=0, t=0, b=0),
                height=335,
                uirevision=selected,  # preserve zoom per virus, reset on virus change
            )
            event = st.plotly_chart(
                fig, use_container_width=True,
                on_select="rerun", selection_mode="points",
                key="virus_map_chart",
            )

            # Detect new clicks (toggle: add if not selected, remove if already selected)
            curr_click_set = set()
            if event and event.selection and event.selection.points:
                for pt in event.selection.points:
                    country_name = pt.get("hovertext", "")
                    if country_name:
                        curr_click_set.add(country_name)

            prev_click_set = set(st.session_state.get("_virus_map_prev_clicks", []))
            new_clicks = curr_click_set - prev_click_set
            st.session_state["_virus_map_prev_clicks"] = list(curr_click_set)

            if new_clicks:
                current = list(st.session_state["virus_map_selected_countries"])
                for c in new_clicks:
                    if c in current:
                        current.remove(c)
                    else:
                        current.append(c)
                st.session_state["virus_map_selected_countries"] = current
                st.session_state["_virus_map_sync"] = True
                # Clear plotly selection state to allow re-clicking same country
                st.session_state.pop("virus_map_chart", None)
                st.rerun()

            # Continent toggle buttons — directly below map
            if _continent_countries:
                # Pad with empty columns to compress buttons toward center
                n = len(_continent_countries)
                cont_cols_all = st.columns([0.5] + [1] * n + [0.5], gap="small")
                cont_cols = cont_cols_all[1:-1]
                for i, cont in enumerate(_continent_countries):
                    with cont_cols[i]:
                        cont_set = set(_continent_countries[cont])
                        all_selected = cont_set.issubset(set(st.session_state.get("virus_map_selected_countries", [])))
                        btn_type = "primary" if all_selected else "secondary"
                        short = cont
                        if st.button(short, key=f"cont_{cont}", use_container_width=True, type=btn_type):
                            current = list(st.session_state["virus_map_selected_countries"])
                            if all_selected:
                                current = [c for c in current if c not in cont_set]
                            else:
                                for c in cont_set:
                                    if c not in current:
                                        current.append(c)
                            st.session_state["virus_map_selected_countries"] = current
                            st.session_state["_virus_map_sync"] = True
                            st.session_state.pop("virus_map_chart", None)
                            st.rerun()
        else:
            # Show empty world map
            import plotly.graph_objects as go
            fig = go.Figure(go.Choropleth(locations=[], z=[]))
            fig.update_layout(
                geo=dict(showframe=False, showcoastlines=True, projection_type="natural earth"),
                margin=dict(l=0, r=0, t=0, b=0),
                height=450,
            )
            st.plotly_chart(fig, use_container_width=True)

    # --- Bottom section: selected countries + action buttons ---
    all_country_options = display_df["country"].tolist() if df is not None else []

    # Only overwrite multiselect from map clicks, not on every rerun
    if st.session_state.pop("_virus_map_sync", False):
        valid_accumulated = [c for c in st.session_state.get("virus_map_selected_countries", []) if c in all_country_options]
        st.session_state["virus_map_country_multiselect"] = valid_accumulated

    def _on_multiselect_change():
        st.session_state["virus_map_selected_countries"] = list(
            st.session_state.get("virus_map_country_multiselect", [])
        )

    col_sel, col_reset = st.columns([6, 1])
    with col_sel:
        selected_countries = st.multiselect(
            "Selected countries",
            options=all_country_options,
            default=[c for c in st.session_state.get("virus_map_selected_countries", []) if c in all_country_options]
                if "virus_map_country_multiselect" not in st.session_state else None,
            key="virus_map_country_multiselect",
            on_change=_on_multiselect_change,
        )
    with col_reset:
        st.markdown("<div style='height: 1.8rem'></div>", unsafe_allow_html=True)
        if st.button("Reset", key="virus_map_reset_sel"):
            st.session_state["virus_map_selected_countries"] = []
            st.session_state["_virus_map_prev_clicks"] = []
            st.session_state.pop("virus_map_country_multiselect", None)
            st.session_state.pop("virus_map_chart", None)
            st.session_state["_virus_map_sync"] = True
            st.rerun()
    # Sync multiselect back to accumulated state
    st.session_state["virus_map_selected_countries"] = list(selected_countries)

    # Update metrics based on selection
    if selected_countries and df is not None:
        sel_df = df[df["country"].isin(selected_countries)]
        metrics_seq.metric("Sequences", f"{int(sel_df['sequences'].sum()):,}")
        metrics_countries.metric("Countries", len(sel_df))
    else:
        metrics_seq.metric("Sequences", f"{seq_count:,}")
        metrics_countries.metric("Countries", len(display_df))

    def _prefill_virus(prefix: str):
        """Pre-fill virus, date, and country settings for the target workflow."""
        virus_display = next(
            (f"{name}  (TaxID: {tid})" for name, tid in VIRUS_PRESETS
             if str(tid) == str(data["taxid"])),
            None,
        )
        if virus_display:
            st.session_state[f"{prefix}_virus_select"] = virus_display
            # Set target name from virus name
            virus_name = virus_display.split("  (TaxID:")[0]
            sanitized = re.sub(r"[^\w\-]", "_", virus_name)
            sanitized = re.sub(r"_+", "_", sanitized).strip("_")
            # Pass subtype filter if present (e.g., H1N1 from Influenza A map data)
            subtype = data.get("subtype_filter", "")
            if subtype:
                st.session_state[f"{prefix}_subtype_filter"] = subtype
                subtype_clean = re.sub(r"[^\w\-]", "_", subtype)
                sanitized = f"{sanitized}_{subtype_clean}"
            st.session_state[f"{prefix}_target_name"] = sanitized
            st.session_state[f"_prev_auto_{prefix}"] = virus_name
        if selected_countries:
            # If all countries of a continent are selected, use the continent name
            # instead of individual country names (much faster for gget fetch)
            geo_parts = []
            remaining = set(selected_countries)
            for cont, cont_countries in _continent_countries.items():
                cont_set = set(cont_countries)
                if cont_set.issubset(remaining):
                    geo_parts.append(cont)
                    remaining -= cont_set
            # Add any individually selected countries
            geo_parts.extend(sorted(remaining))
            st.session_state[f"{prefix}_geo_location"] = ", ".join(geo_parts)
        # Pass date range as collection date filters
        st.session_state[f"{prefix}_min_collection_date"] = filter_start
        st.session_state[f"{prefix}_max_collection_date"] = filter_end
        # Clean up virus map state
        st.session_state["virus_map_selected_countries"] = []

    btn1, btn2, btn3 = st.columns(3)
    with btn1:
        if st.button("Start Design", type="primary", use_container_width=True):
            _reset_workflow_state()
            st.session_state.workflow = "design"
            _prefill_virus("fetch")
            _navigate("select_target")
            st.rerun()
    with btn2:
        if st.button("Start Evaluate", type="primary", use_container_width=True):
            _reset_workflow_state()
            st.session_state.workflow = "evaluate"
            _prefill_virus("fetch")
            _navigate("select_target")
            st.rerun()
    with btn3:
        if st.button("Start Monitor", type="primary", use_container_width=True):
            _reset_workflow_state()
            st.session_state.workflow = "monitor"
            _prefill_virus("monitor_fetch")
            _navigate("monitor_target")
            st.rerun()

    st.caption("SARS-CoV-2, Influenza A, and HIV-1 are excluded from the map due to excessive NCBI dataset size.")


# ---------------------------------------------------------------------------
# Tab 1: Files
# ---------------------------------------------------------------------------

def _tab_files():
    st.header("FASTA Files")
    st.write("Upload FASTA files to `target_seqs/original/` for use as targets, cross-reactivity, or host references.")

    # Upload
    uploaded = st.file_uploader(
        "Upload FASTA file(s)",
        type=["fa", "fasta", "fna", "bz2"],
        accept_multiple_files=True,
        key="fasta_upload",
    )
    if uploaded:
        import bz2 as _bz2
        already_saved = st.session_state.get("_fasta_saved_ids", set())
        new_saved = set()
        for f in uploaded:
            fid = (f.name, f.size)
            if fid in already_saved:
                continue
            p = Path(f.name)
            # Handle .fa.bz2 files
            if p.name.endswith(".fa.bz2"):
                stem = p.name[:-7]  # strip .fa.bz2
                dest = _safe_path_under(FASTA_DIR, stem + ".fa")
                if not dest:
                    st.error(f"Invalid filename: {f.name}")
                    continue
                _save_fasta_bytes(dest, _bz2.decompress(f.getvalue()))
            elif p.suffix.lower() not in (".fa", ".fasta", ".fna"):
                st.error(f"Unsupported extension: {p.suffix}")
                continue
            else:
                # Normalize to .fa extension
                dest = _safe_path_under(FASTA_DIR, p.stem + ".fa")
                if not dest:
                    st.error(f"Invalid filename: {f.name}")
                    continue
                _save_fasta_bytes(dest, f.getvalue())
            # Remove any leftover files with other FASTA extensions
            for ext in (".fasta", ".fna"):
                old = FASTA_DIR / (p.stem + ext)
                if old.exists():
                    old.unlink()
            st.success(f"Saved {dest.name}")
            new_saved.add(fid)
        if new_saved:
            st.session_state["_fasta_saved_ids"] = already_saved | new_saved

    st.divider()

    # File listing
    fastas = _available_fasta()
    if not fastas:
        st.info("No FASTA files found. Upload files above.")
        return

    st.subheader(f"Available files ({len(fastas)})")

    for stem in fastas:
        info = _fasta_info(stem)
        col1, col2, col3, col4 = st.columns([3, 2, 2, 1])
        col1.write(f"**{stem}**")
        col2.write(f"{info['seqs']} sequences")
        col3.write(_format_size(info["size"]))
        if col4.button("Delete", key=f"del_{stem}"):
            # Find and delete the file
            for ext in (".fa", ".fasta", ".fna"):
                p = FASTA_DIR / f"{stem}{ext}"
                if p.exists():
                    p.unlink()
                    break
            st.rerun()


# ---------------------------------------------------------------------------
# Tab 2: Configuration
# ---------------------------------------------------------------------------

def _tab_configuration():
    st.header("Pipeline Configuration")

    fastas = _available_fasta()

    mode = st.radio(
        "Design mode",
        ["Singleplex", "Multiplex", "Evaluate"],
        key="mode",
        horizontal=True,
    )

    probe = st.checkbox("Enable probe design", key="probe_enabled")

    st.divider()

    if mode == "Singleplex":
        st.subheader("Singleplex settings")
        st.multiselect(
            "TARGETS (on-target sequences to design primers for)",
            options=fastas,
            key="targets",
        )
        st.multiselect(
            "CROSS (cross-reactivity sequences to avoid)",
            options=fastas,
            key="cross",
        )
        st.multiselect(
            "HOST (host genome sequences to avoid)",
            options=fastas,
            key="host",
        )

    elif mode == "Multiplex":
        st.subheader("Multiplex settings")
        st.multiselect(
            "PANEL (targets for multiplex panel)",
            options=fastas,
            key="panel",
        )
        st.multiselect(
            "HOST (host genome sequences to avoid)",
            options=fastas,
            key="host_multiplex",
        )

    elif mode == "Evaluate":
        st.subheader("Evaluate existing primers")

        eval_method = st.radio(
            "Input method",
            ["Paste sequences", "Upload primer FASTA"],
            key="eval_method",
            horizontal=True,
        )

        if eval_method == "Paste sequences":
            st.text_input("Forward primer sequence", key="eval_for")
            st.text_input("Reverse primer sequence", key="eval_rev")
        else:
            pset_upload = st.file_uploader(
                "Upload primer set FASTA",
                type=["fa", "fasta"],
                key="eval_pset_upload",
            )
            if pset_upload:
                fid = (pset_upload.name, pset_upload.size)
                if st.session_state.get("_eval_pset_saved") != fid:
                    pset_dir = PROJECT_ROOT / "evaluate"
                    pset_dir.mkdir(parents=True, exist_ok=True)
                    pset_path = pset_dir / pset_upload.name
                    _save_fasta_bytes(pset_path, pset_upload.getvalue())
                    st.session_state["eval_pset_path"] = str(pset_path)
                    st.session_state["_eval_pset_saved"] = fid
                st.success(f"Saved {pset_upload.name}")

        st.selectbox(
            "Evaluation target (must match an uploaded FASTA)",
            options=fastas if fastas else ["(no files available)"],
            key="eval_target",
        )


# ---------------------------------------------------------------------------
# Tab 3: Parameters
# ---------------------------------------------------------------------------

def _tab_parameters(mode: str = "design"):
    _init_params()
    p = st.session_state.params

    if st.button("Reset to defaults"):
        st.session_state.params = dict(DEFAULT_PARAMS)
        st.rerun()

    if mode == "evaluate":
        # Evaluate only needs amplicon length parameters
        with st.expander("Amplicon lengths", expanded=True):
            c11, c12 = st.columns(2)
            p["AMPLEN_MIN"] = c11.number_input(
                "AMPLEN_MIN", value=int(p["AMPLEN_MIN"]),
                min_value=30, step=10, key="p_amp_min",
            )
            p["AMPLEN_MAX"] = c12.number_input(
                "AMPLEN_MAX", value=int(p["AMPLEN_MAX"]),
                min_value=30, step=10, key="p_amp_max",
            )
            if p["AMPLEN_MIN"] >= p["AMPLEN_MAX"]:
                st.warning("AMPLEN_MIN should be less than AMPLEN_MAX")

            c13, c14 = st.columns(2)
            p["OFFLEN_MIN"] = c13.number_input(
                "OFFLEN_MIN", value=int(p["OFFLEN_MIN"]),
                min_value=30, step=10, key="p_off_min",
            )
            p["OFFLEN_MAX"] = c14.number_input(
                "OFFLEN_MAX", value=int(p["OFFLEN_MAX"]),
                min_value=30, step=100, key="p_off_max",
            )

        # Show probe cutoffs when a probe sequence is provided
        eval_pro = st.session_state.get("_eval_pro_saved", "") or st.session_state.get("eval_pro", "")
        eval_pset = st.session_state.get("eval_pset_path", "")
        has_probe_seq = bool(eval_pro) or (bool(eval_pset) and _pset_has_probe(eval_pset))
        if has_probe_seq:
            with st.expander("Probe filtering", expanded=True):
                c15, c16 = st.columns(2)
                p["PROBE_MAX_MISMATCHES"] = c15.number_input(
                    "PROBE_MAX_MISMATCHES", value=int(p["PROBE_MAX_MISMATCHES"]),
                    min_value=0, max_value=10, key="p_eval_probe_mm",
                    help="Maximum mismatches for probe to count as matched",
                )
                p["PROBE_MAX_INDELS"] = c16.number_input(
                    "PROBE_MAX_INDELS", value=int(p["PROBE_MAX_INDELS"]),
                    min_value=0, max_value=5, key="p_eval_probe_indel",
                    help="Maximum indels for probe to count as matched",
                )
        return

    # Check if any selected target has >1 sequence (quick mode)
    selected_targets = st.session_state.get("targets", [])
    is_quick = any(
        _fasta_info(t).get("seqs", 0) > 1 for t in selected_targets
    ) if selected_targets else False

    # --- Primer generation ---
    with st.expander("Primer generation", expanded=True):
        if is_quick:
            st.caption("Multi-sequence target detected — using quick design mode.")
            c_q1, c_q2 = st.columns(2)
            p["QUICK_N_POSITIONS"] = c_q1.number_input(
                "MSA sites to explore", value=int(p["QUICK_N_POSITIONS"]),
                min_value=10, max_value=1000, step=10, key="p_quick_n_pos",
                help="Number of top conserved amplicon positions to evaluate across the MSA.",
            )
            p["QUICK_N_VARIANTS"] = c_q2.number_input(
                "Primer variants per site", value=int(p["QUICK_N_VARIANTS"]),
                min_value=1, max_value=50, step=1, key="p_quick_n_var",
                help="Number of wobble-optimized primer variants per position per direction.",
            )
        else:
            p["DESIGN_WINDOW"] = st.number_input(
                "Alignment window size (bp)", value=int(p["DESIGN_WINDOW"]),
                min_value=50, step=50, key="p_design_window",
                help="Window size for dividing the MSA when selecting representative sequences.",
            )

        c1, c2 = st.columns(2)
        p["TM_MIN"] = c1.number_input(
            "TM_MIN", value=float(p["TM_MIN"]), step=0.5, key="p_tm_min",
        )
        p["TM_MAX"] = c2.number_input(
            "TM_MAX", value=float(p["TM_MAX"]), step=0.5, key="p_tm_max",
        )
        if p["TM_MIN"] >= p["TM_MAX"]:
            st.warning("TM_MIN should be less than TM_MAX")

        p["GC_MAX"] = st.number_input(
            "GC_MAX (%)", value=float(p["GC_MAX"]), step=1.0, key="p_gc_max",
        )
        p["DG_MIN"] = st.number_input(
            "DG_MIN (kcal/mol)", value=float(p["DG_MIN"]), step=0.5, key="p_dg_min",
        )

        c3, c4 = st.columns(2)
        p["PRIMER_LEN_MIN"] = c3.number_input(
            "PRIMER_LEN_MIN", value=int(p["PRIMER_LEN_MIN"]),
            min_value=15, max_value=35, key="p_len_min",
        )
        p["PRIMER_LEN_MAX"] = c4.number_input(
            "PRIMER_LEN_MAX", value=int(p["PRIMER_LEN_MAX"]),
            min_value=15, max_value=35, key="p_len_max",
        )
        if p["PRIMER_LEN_MIN"] > p["PRIMER_LEN_MAX"]:
            st.warning("PRIMER_LEN_MIN should be <= PRIMER_LEN_MAX")


    # --- Probe generation ---
    probe_enabled = st.session_state.get("probe_enabled", False)
    if probe_enabled:
        with st.expander("Probe", expanded=True):
            c5, c6 = st.columns(2)
            p["PROBE_LEN_MIN"] = c5.number_input(
                "PROBE_LEN_MIN", value=int(p["PROBE_LEN_MIN"]),
                min_value=15, max_value=40, key="p_probe_len_min",
            )
            p["PROBE_LEN_MAX"] = c6.number_input(
                "PROBE_LEN_MAX", value=int(p["PROBE_LEN_MAX"]),
                min_value=15, max_value=40, key="p_probe_len_max",
            )
            if p["PROBE_LEN_MIN"] > p["PROBE_LEN_MAX"]:
                st.warning("PROBE_LEN_MIN should be <= PROBE_LEN_MAX")

            c7, c8 = st.columns(2)
            p["PROBE_TM_MIN"] = c7.number_input(
                "PROBE_TM_MIN", value=float(p["PROBE_TM_MIN"]),
                step=0.5, key="p_probe_tm_min",
            )
            p["PROBE_TM_MAX"] = c8.number_input(
                "PROBE_TM_MAX", value=float(p["PROBE_TM_MAX"]),
                step=0.5, key="p_probe_tm_max",
            )
            if p["PROBE_TM_MIN"] >= p["PROBE_TM_MAX"]:
                st.warning("PROBE_TM_MIN should be less than PROBE_TM_MAX")

            p["PROBE_GC_MAX"] = st.number_input(
                "PROBE_GC_MAX (%)", value=float(p["PROBE_GC_MAX"]),
                step=1.0, key="p_probe_gc_max",
            )
            p["PROBE_DG_MIN"] = st.number_input(
                "PROBE_DG_MIN (kcal/mol)", value=float(p["PROBE_DG_MIN"]),
                step=0.5, key="p_probe_dg_min",
            )

            p["PROBE_HOMOPOLYMER_MAX"] = st.number_input(
                "PROBE_HOMOPOLYMER_MAX", value=int(p["PROBE_HOMOPOLYMER_MAX"]),
                min_value=1, max_value=10, key="p_probe_hp",
            )
            p["PROBE_AVOID_5PRIME_G"] = st.checkbox(
                "PROBE_AVOID_5PRIME_G", value=bool(p["PROBE_AVOID_5PRIME_G"]),
                key="p_probe_5g",
            )

            c9, c10 = st.columns(2)
            p["PROBE_MAX_MISMATCHES"] = c9.number_input(
                "PROBE_MAX_MISMATCHES", value=int(p["PROBE_MAX_MISMATCHES"]),
                min_value=0, max_value=10, key="p_probe_mm",
            )
            p["PROBE_MAX_INDELS"] = c10.number_input(
                "PROBE_MAX_INDELS", value=int(p["PROBE_MAX_INDELS"]),
                min_value=0, max_value=10, key="p_probe_indels",
            )

    # --- Input preparation ---
    with st.expander("Input preparation (amplicon lengths)", expanded=False):
        c11, c12 = st.columns(2)
        p["AMPLEN_MIN"] = c11.number_input(
            "AMPLEN_MIN", value=int(p["AMPLEN_MIN"]),
            min_value=30, step=10, key="p_amp_min",
        )
        p["AMPLEN_MAX"] = c12.number_input(
            "AMPLEN_MAX", value=int(p["AMPLEN_MAX"]),
            min_value=30, step=10, key="p_amp_max",
        )
        if p["AMPLEN_MIN"] >= p["AMPLEN_MAX"]:
            st.warning("AMPLEN_MIN should be less than AMPLEN_MAX")

        c13, c14 = st.columns(2)
        p["OFFLEN_MIN"] = c13.number_input(
            "OFFLEN_MIN", value=int(p["OFFLEN_MIN"]),
            min_value=30, step=10, key="p_off_min",
        )
        p["OFFLEN_MAX"] = c14.number_input(
            "OFFLEN_MAX", value=int(p["OFFLEN_MAX"]),
            min_value=30, step=100, key="p_off_max",
        )

    # --- Evaluation ---
    with st.expander("Evaluation", expanded=False):
        p["NUM_TOP_SENSITIVITY"] = st.number_input(
            "NUM_TOP_SENSITIVITY", value=int(p["NUM_TOP_SENSITIVITY"]),
            min_value=10, step=10, key="p_num_top",
        )
        if is_quick:
            c_ev1, c_ev2 = st.columns(2)
            p["QUICK_COV_MIN"] = c_ev1.number_input(
                "Coverage goal", value=float(p.get("QUICK_COV_MIN", 1.0)),
                min_value=0.0, max_value=1.0, step=0.05, key="p_quick_cov",
                help="Minimum coverage fraction for early stopping (1.0 = all sequences).",
            )
            p["QUICK_ACT_MIN"] = c_ev2.number_input(
                "Activity goal", value=float(p.get("QUICK_ACT_MIN", 1.0)),
                min_value=0.0, max_value=1.0, step=0.05, key="p_quick_act",
                help="Minimum activity fraction for early stopping (1.0 = all sequences).",
            )


# ---------------------------------------------------------------------------
# Tab 4: Run
# ---------------------------------------------------------------------------

def _build_command() -> list[str]:
    """Build the snakemake command list from current session state."""
    mode = st.session_state.get("mode", "Singleplex")
    probe = st.session_state.get("probe_enabled",
                                  st.session_state.get("_probe_enabled_saved", False))
    cores = st.session_state.get("cores", 1)
    dry_run = st.session_state.get("dry_run", False)

    snakemake_bin = _find_tool("snakemake") or "snakemake"
    cmd = [snakemake_bin, "-s", "Snakefile", "--cores", str(cores)]

    config_args: list[str] = []
    if mode == "Multiplex":
        config_args.append("multiplex=1")
    if probe:
        config_args.append("probe=1")
    if mode == "Evaluate":
        config_args.append("evaluate=1")
        pset_path = st.session_state.get("eval_pset_path", "")
        if pset_path:
            config_args.append(f"pset={pset_path}")
        else:
            fwd = st.session_state.get("eval_for") or st.session_state.get("_eval_for_saved", "")
            rev = st.session_state.get("eval_rev") or st.session_state.get("_eval_rev_saved", "")
            if fwd:
                config_args.append(f"for={fwd}")
            if rev:
                config_args.append(f"rev={rev}")
            pro = st.session_state.get("eval_pro") or st.session_state.get("_eval_pro_saved", "")
            if pro:
                config_args.append(f"pro={pro}")

    if config_args:
        cmd += ["--config"] + config_args

    if dry_run:
        cmd.append("--dry-run")

    return cmd


def _preflight_checks() -> list[str]:
    """Return a list of preflight error messages. Empty = all OK."""
    errors: list[str] = []
    mode = st.session_state.get("mode", "Singleplex")

    # Check required targets
    if mode == "Singleplex":
        if not st.session_state.get("targets"):
            errors.append("No TARGETS selected in Configuration tab.")
    elif mode == "Multiplex":
        if not st.session_state.get("panel"):
            errors.append("No PANEL targets selected in Configuration tab.")
    elif mode == "Evaluate":
        has_pset = bool(st.session_state.get("eval_pset_path"))
        eval_fwd = st.session_state.get("eval_for") or st.session_state.get("_eval_for_saved", "")
        eval_rev = st.session_state.get("eval_rev") or st.session_state.get("_eval_rev_saved", "")
        has_seq = bool(eval_fwd) and bool(eval_rev)
        if not has_pset and not has_seq:
            errors.append("Provide primer sequences or upload a primer set FASTA.")
        if not st.session_state.get("targets"):
            errors.append("No target sequence selected.")

    # Check tools
    for tool in ("snakemake", "bowtie2", "mafft", "sam2pairwise"):
        if not _check_tool(tool):
            errors.append(f"Tool '{tool}' not found in PATH.")

    return errors


def _write_pipeline_files():
    """Write Snakefile and params.txt to project root."""
    mode = st.session_state.get("mode", "Singleplex")
    probe = st.session_state.get("probe_enabled",
                                  st.session_state.get("_probe_enabled_saved", False))

    targets = st.session_state.get("targets", [])
    cross = st.session_state.get("cross", [])
    panel = st.session_state.get("panel", [])

    if mode == "Multiplex":
        host = st.session_state.get("host_multiplex", [])
    else:
        host = st.session_state.get("host", [])

    run_id = st.session_state.get("design_run_id", "").strip()
    if not run_id:
        run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    st.session_state.run_id = run_id

    snakefile_content = build_snakefile(
        targets=targets,
        cross=cross,
        host=host,
        panel=panel,
        run_id=run_id,
    )
    (PROJECT_ROOT / "Snakefile").write_text(snakefile_content)

    _init_params()
    params_content = build_params_txt(st.session_state.params)
    (PROJECT_ROOT / "params.txt").write_text(params_content)


# Pipeline rule steps grouped for progress display
_PIPELINE_RULES_DESIGN_STANDARD = [
    ("make_MSA", "Building multiple sequence alignment"),
    ("pick_representative_seqs", "Selecting representative sequences"),
    ("generate_primers", "Generating primer candidates"),
    ("build_index", "Building bowtie2 index"),
    ("align", "Aligning primers to targets"),
    ("parse_map", "Parsing alignment results"),
    ("process_map", "Processing alignment maps"),
    ("prepare_input", "Preparing ML model input"),
    ("evaluate", "Evaluating with ML model"),
    ("rescue_evaluate", "Rescue evaluation"),
    ("filter_primer_list", "Filtering and ranking primers"),
    ("check_coverage", "Checking primer coverage"),
]

_PIPELINE_RULES_DESIGN_QUICK = [
    ("make_MSA", "Building multiple sequence alignment"),
    ("quick_design_on_target", "Designing with AI"),
    ("filter_primer_list", "Filtering and ranking primers"),
]

_PIPELINE_RULES_EVALUATE = [
    ("build_index", "Building bowtie2 index"),
    ("align", "Aligning primers to targets"),
    ("parse_map", "Parsing alignment results"),
    ("process_map", "Processing alignment maps"),
    ("prepare_input", "Preparing ML model input"),
    ("evaluate_pset", "Evaluating primer set"),
]


def _get_pipeline_rules() -> list[tuple[str, str]]:
    mode = st.session_state.get("mode", "Singleplex")
    if mode == "Evaluate":
        return _PIPELINE_RULES_EVALUATE
    # Check if any target has >1 sequence (quick mode)
    targets = st.session_state.get("targets", [])
    is_quick = any(_fasta_info(t).get("seqs", 0) > 1 for t in targets) if targets else False
    if is_quick:
        return _PIPELINE_RULES_DESIGN_QUICK
    return _PIPELINE_RULES_DESIGN_STANDARD


def _get_all_targets() -> list[str]:
    """Return all targets (on + off) for the current run."""
    targets = st.session_state.get("targets", [])
    cross = st.session_state.get("cross", [])
    host = st.session_state.get("host", [])
    return list(dict.fromkeys(targets + cross + host))  # preserve order, deduplicate


# Rules that run once per target (have a {target} wildcard)
_PER_TARGET_RULES = {
    "build_index", "align", "align_probes", "parse_map", "parse_probe_mapping",
    "process_map", "check_coverage", "prepare_input", "evaluate",
    "make_MSA", "pick_representative_seqs", "generate_primers",
    "resolve_target_seq", "generate_probes",
    "filter_primer_list", "build_final_output", "quick_design_on_target",
}


def _save_run_config():
    """Save all run settings to a JSON file alongside the final output."""
    run_id = st.session_state.get("run_id", "")
    if not run_id:
        return

    workflow = st.session_state.get("workflow", "design")
    mode = st.session_state.get("mode", "Singleplex")

    config = {
        "run_id": run_id,
        "timestamp": datetime.now().isoformat(),
        "workflow": workflow,
        "mode": mode,
        "targets": st.session_state.get("targets", []),
        "cross_reactivity": st.session_state.get("cross", []),
        "host": st.session_state.get("host", []),
        "probe_enabled": st.session_state.get("probe_enabled",
                                              st.session_state.get("_probe_enabled_saved", False)),
        "cores": st.session_state.get("cores", 1),
    }

    # Fetch settings (target)
    fetch_prefix = "fetch"
    fetch_config = {}
    virus_sel = st.session_state.get(f"{fetch_prefix}_virus_select")
    if virus_sel:
        fetch_config["virus"] = virus_sel
    for key_suffix in ("target_name", "nuc_completeness", "segment",
                       "min_seq_length", "max_seq_length",
                       "min_release_date", "max_release_date",
                       "geo_location", "host",
                       "limit", "min_collection_date", "max_collection_date",
                       "max_ambiguous_chars", "refseq_only"):
        val = st.session_state.get(f"{fetch_prefix}_{key_suffix}")
        if val is not None and val != "" and val is not False:
            fetch_config[key_suffix] = val
    if fetch_config:
        config["fetch_settings"] = fetch_config

    # Evaluate-specific
    if mode == "Evaluate":
        config["eval_method"] = st.session_state.get("eval_method", "")
        config["eval_target"] = st.session_state.get("eval_target", "")
        if st.session_state.get("eval_method") == "Paste sequences":
            config["eval_forward"] = st.session_state.get("eval_for") or st.session_state.get("_eval_for_saved", "")
            config["eval_reverse"] = st.session_state.get("eval_rev") or st.session_state.get("_eval_rev_saved", "")
        else:
            config["eval_pset_path"] = st.session_state.get("eval_pset_path", "")

    # Parameters
    config["parameters"] = dict(st.session_state.get("params", {}))

    # Determine output directory
    out_dir = RUNS_DIR / run_id
    out_dir.mkdir(parents=True, exist_ok=True)

    # Custom JSON serializer for dates
    def _json_default(obj):
        if isinstance(obj, (date, datetime)):
            return obj.isoformat()
        return str(obj)

    config_path = out_dir / "run_config.json"
    config_path.write_text(json.dumps(config, indent=2, default=_json_default, ensure_ascii=False))


def _tab_run():
    workflow = st.session_state.get("workflow", "design")
    st.header("Run Design" if workflow == "design" else "Run Evaluate")

    # Run ID
    default_run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    if "design_run_id" not in st.session_state:
        st.session_state["design_run_id"] = default_run_id
    st.text_input("Run ID", key="design_run_id",
                  help="Identifier for this run. Results will be saved under this name.")

    # Controls
    c1, c2 = st.columns(2)
    max_cpu = os.cpu_count() or 1
    c1.slider("CPU cores", min_value=1, max_value=max_cpu, value=min(4, max_cpu), key="cores")
    c2.checkbox("Dry run (plan only, no execution)", key="dry_run")

    # Preflight
    errors = _preflight_checks()
    if errors:
        for e in errors:
            st.error(e)

    rules = _get_pipeline_rules()

    # Clear previous result when entering Run page for first time
    if st.session_state.get("_run_page_fresh", True):
        st.session_state.pipeline_return_code = None
        st.session_state.pipeline_log = ""
        st.session_state.pipeline_completed_rules = set()
        st.session_state._run_page_fresh = False

    # Run / Stop buttons
    col_run, col_stop = st.columns(2)
    running = st.session_state.get("pipeline_running", False)

    with col_run:
        run_label = "Run Design" if workflow == "design" else "Run Evaluate"
        run_clicked = st.button(run_label, disabled=running or bool(errors), type="primary")
    with col_stop:
        stop_clicked = st.button("Stop", disabled=not running, type="secondary")

    # Handle Stop
    if stop_clicked:
        pid = st.session_state.get("pipeline_pid")
        if pid:
            try:
                os.killpg(os.getpgid(pid), signal.SIGTERM)
            except (ProcessLookupError, PermissionError):
                try:
                    os.kill(pid, signal.SIGTERM)
                except (ProcessLookupError, PermissionError):
                    pass
        st.session_state.pipeline_running = False
        st.session_state.pipeline_return_code = -1
        st.rerun()

    # Handle Run click — set state and rerun so buttons update first
    if run_clicked:
        st.session_state.pipeline_running = True
        st.session_state.pipeline_should_start = True
        st.session_state.pipeline_return_code = None
        st.session_state.pipeline_log = ""
        st.session_state.pipeline_completed_rules = set()
        st.rerun()

    # Execute pipeline (triggered after rerun from Run click)
    if st.session_state.get("pipeline_should_start"):
        st.session_state.pipeline_should_start = False

        _write_pipeline_files()

        # Unlock snakemake directory in case of leftover locks
        snakemake_bin = _find_tool("snakemake") or "snakemake"
        subprocess.run(
            [snakemake_bin, "-s", "Snakefile", "--unlock", "--cores", "1"],
            capture_output=True, cwd=str(PROJECT_ROOT),
        )

        env = os.environ.copy()
        env["PYTHONUNBUFFERED"] = "1"
        # Ensure the conda env bin dir is on PATH (for sam2pairwise, bowtie2, etc.)
        # Use snakemake's location to find the correct env bin dir, since
        # CONDA_PREFIX may point to the base env rather than qprimer-designer.
        snakemake_path = _find_tool("snakemake")
        env_bin = str(Path(sys.executable).parent)
        extra_dirs = [env_bin]
        if snakemake_path:
            extra_dirs.append(str(Path(snakemake_path).parent))
        for d in extra_dirs:
            if d and d not in env.get("PATH", ""):
                env["PATH"] = d + os.pathsep + env.get("PATH", "")

        cmd = _build_command()

        progress_area = st.empty()
        status_area = st.empty()

        all_targets = _get_all_targets()
        show_targets = len(all_targets) > 1

        completed_rules: set[str] = set()
        # Per-target tracking: rule_name -> set of completed targets
        rule_done_targets: dict[str, set[str]] = {}
        current_rule = rules[0][0] if rules else ""
        current_target = ""
        log = ""
        start_time = time.time()

        def _finish_current():
            """Mark the current rule+target as completed."""
            nonlocal current_target
            if current_rule and current_target:
                rule_done_targets.setdefault(current_rule, set()).add(current_target)
                current_target = ""

        def _render_rule_line(rule_name, label, is_current=False):
            done_targets = rule_done_targets.get(rule_name, set())
            is_per_target = rule_name in _PER_TARGET_RULES and show_targets

            if is_per_target:
                is_done = done_targets >= set(all_targets)
            else:
                is_done = rule_name in completed_rules

            if is_done:
                if is_per_target:
                    tstr = ", ".join(sorted(done_targets))
                    return f"✅ &nbsp; {label} — {tstr}"
                return f"✅ &nbsp; {label}"
            elif is_current:
                if is_per_target:
                    parts = []
                    for t in all_targets:
                        if t in done_targets:
                            parts.append(f"{t} ✓")
                        elif t == current_target:
                            parts.append(f"**{t}**")
                        else:
                            parts.append(f"<span style='color:#ccc'>{t}</span>")
                    tstr = ", ".join(parts)
                    return f"⏳ &nbsp; **{label}** — {tstr}"
                return f"⏳ &nbsp; **{label}** ..."
            elif is_per_target and done_targets:
                # Not current rule but has some targets done (partially complete)
                parts = []
                for t in all_targets:
                    if t in done_targets:
                        parts.append(f"{t} ✓")
                    else:
                        parts.append(f"<span style='color:#ccc'>{t}</span>")
                tstr = ", ".join(parts)
                return f"⏳ &nbsp; {label} — {tstr}"
            else:
                return f"⚪ &nbsp; {label}"

        def _update_progress():
            lines = []
            for rule_name, label in rules:
                is_current = rule_name == current_rule
                lines.append(_render_rule_line(rule_name, label, is_current))
            elapsed = int(time.time() - start_time)
            mins, secs = divmod(elapsed, 60)
            progress_area.markdown("<br>".join(lines), unsafe_allow_html=True)
            status_area.caption(f"Elapsed: {mins}m {secs}s")

        _update_progress()

        proc = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            cwd=str(PROJECT_ROOT),
            env=env,
            text=True,
        )
        st.session_state.pipeline_pid = proc.pid

        while proc.poll() is None:
            line = proc.stdout.readline()
            if line:
                log += line
                stripped = line.strip()
                if stripped.startswith("rule ") or stripped.startswith("localrule "):
                    rule_name = stripped.split("rule ")[1].rstrip(":")
                    # Finish previous rule+target
                    _finish_current()
                    if current_rule and current_rule != rule_name:
                        # Only mark non-per-target rules as completed on transition
                        if current_rule not in _PER_TARGET_RULES or not show_targets:
                            completed_rules.add(current_rule)
                    current_rule = rule_name
                    current_target = ""
                    _update_progress()
                elif stripped.startswith("wildcards:"):
                    # Parse target from wildcards line (target=, virus=, on=, or filename=)
                    wc_target = ""
                    for wc_name in ("target", "virus", "on"):
                        m = re.search(rf"{wc_name}=(\S+)", stripped)
                        if m:
                            wc_target = m.group(1).rstrip(",")
                            break
                    if not wc_target:
                        # Try filename= (format: "virus.target")
                        m = re.search(r"filename=(\S+)", stripped)
                        if m:
                            parts = m.group(1).rstrip(",").split(".")
                            if len(parts) >= 2:
                                wc_target = parts[-1]
                    if wc_target:
                        current_target = wc_target
                        _update_progress()
                elif "Finished job" in stripped:
                    _finish_current()
                    _update_progress()

        # Read remaining output
        remaining = proc.stdout.read()
        if remaining:
            log += remaining

        # Mark last rule as complete
        _finish_current()
        if current_rule:
            if current_rule not in _PER_TARGET_RULES or not show_targets:
                completed_rules.add(current_rule)
        # Mark per-target rules with any completed targets as done
        # (pipeline finished, so all instances have run)
        for rn, _ in rules:
            if rn in _PER_TARGET_RULES and rule_done_targets.get(rn):
                completed_rules.add(rn)

        rc = proc.returncode
        st.session_state.pipeline_running = False
        st.session_state.pipeline_return_code = rc
        st.session_state.pipeline_log = log
        st.session_state.pipeline_completed_rules = completed_rules
        st.session_state.pipeline_rule_done_targets = rule_done_targets

        elapsed = int(time.time() - start_time)
        mins, secs = divmod(elapsed, 60)

        # Final progress
        lines = []
        for rule_name, label in rules:
            if rule_name in completed_rules:
                done_targets = rule_done_targets.get(rule_name, set())
                if rule_name in _PER_TARGET_RULES and show_targets and done_targets:
                    tstr = ", ".join(sorted(done_targets))
                    lines.append(f"✅ &nbsp; {label} — {tstr}")
                else:
                    lines.append(f"✅ &nbsp; {label}")
            else:
                lines.append(f"⚪ &nbsp; {label}")
        progress_area.markdown("<br>".join(lines), unsafe_allow_html=True)
        status_area.caption(f"Completed in {mins}m {secs}s")

        if rc == 0:
            _save_run_config()
            st.success("Pipeline finished successfully!")
        else:
            st.error(f"Pipeline failed with exit code {rc}.")
            with st.expander("Show error log"):
                last_lines = "\n".join(log.splitlines()[-50:])
                st.code(last_lines, language="text")

    # Show previous run result if not currently running
    elif not running and st.session_state.get("pipeline_return_code") is not None:
        rc = st.session_state.pipeline_return_code
        completed_rules = st.session_state.get("pipeline_completed_rules", set())
        rule_done_targets_saved = st.session_state.get("pipeline_rule_done_targets", {})
        all_targets_saved = _get_all_targets()
        show_targets_saved = len(all_targets_saved) > 1

        lines = []
        for rule_name, label in rules:
            if rule_name in completed_rules:
                done_targets = rule_done_targets_saved.get(rule_name, set())
                if rule_name in _PER_TARGET_RULES and show_targets_saved and done_targets:
                    tstr = ", ".join(sorted(done_targets))
                    lines.append(f"✅ &nbsp; {label} — {tstr}")
                else:
                    lines.append(f"✅ &nbsp; {label}")
            else:
                lines.append(f"⚪ &nbsp; {label}")
        st.markdown("<br>".join(lines), unsafe_allow_html=True)

        if rc == 0:
            st.success("Pipeline finished successfully!")
        else:
            st.error(f"Pipeline failed with exit code {rc}.")
            log = st.session_state.get("pipeline_log", "")
            if log:
                with st.expander("Show error log"):
                    last_lines = "\n".join(log.splitlines()[-50:])
                    st.code(last_lines, language="text")


# ---------------------------------------------------------------------------
# Delphy phylogenetic tree helpers
# ---------------------------------------------------------------------------

DELPHY_CACHE_DIR = PROJECT_ROOT / ".delphy_cache"
DELPHY_VERSION = "1.3.4"

import platform as _platform
_DELPHY_PLATFORM = (
    "macos-arm64" if _platform.system() == "Darwin" and _platform.machine() == "arm64"
    else "linux-x86_64" if _platform.system() == "Linux" and _platform.machine() == "x86_64"
    else "linux-arm64" if _platform.system() == "Linux" and _platform.machine() == "aarch64"
    else None
)


def _get_delphy_binary() -> Path | None:
    """Return path to Delphy binary, downloading if needed. Returns None if unsupported platform."""
    if _DELPHY_PLATFORM is None:
        return None

    DELPHY_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    binary = DELPHY_CACHE_DIR / "delphy"

    if binary.exists():
        return binary

    tarball_name = f"delphy-{_DELPHY_PLATFORM}.tar.gz"
    url = f"https://github.com/broadinstitute/delphy/releases/download/{DELPHY_VERSION}/{tarball_name}"
    tarball = DELPHY_CACHE_DIR / tarball_name

    import urllib.request
    urllib.request.urlretrieve(url, tarball)

    import tarfile
    with tarfile.open(tarball, "r:gz") as tf:
        # Find the delphy binary inside the tarball
        for member in tf.getmembers():
            if member.name.endswith("/delphy") or member.name == "delphy":
                member.name = "delphy"
                tf.extract(member, DELPHY_CACHE_DIR)
                break

    tarball.unlink(missing_ok=True)
    binary.chmod(binary.stat().st_mode | 0o755)
    return binary


def _extract_and_format_date(date_string: str) -> str | None:
    """Return YYYY-MM-DD date or None if not a complete date."""
    if not date_string or date_string != date_string:  # NaN check
        return None
    date_string = str(date_string).strip()
    date_string = re.sub(r"[/.]", "-", date_string)
    # Strict YYYY-MM-DD
    if re.match(r"^\d{4}-\d{1,2}-\d{1,2}$", date_string):
        return date_string
    # Try flexible formats
    for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d %b %Y", "%b %d %Y", "%Y/%m/%d"):
        try:
            dt = datetime.strptime(date_string, fmt)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            continue
    return None


def _country_to_alpha2(country_name: str) -> str:
    """Convert country name to 2-letter ISO code. Returns '' on failure."""
    try:
        import pycountry
        results = pycountry.countries.search_fuzzy(country_name)
        if results:
            return results[0].alpha_2
    except (LookupError, Exception):
        pass
    return ""


def _msa_has_delphy_headers(msa_path: Path) -> bool:
    """Check if MSA headers end with '|YYYY-MM-DD' (e.g. acc|date or acc|geo|date)."""
    date_re = re.compile(r"^[^|]+\|(?:[^|]+\|)*\d{4}-\d{2}-\d{2}$")
    n_checked = 0
    n_valid = 0
    with open(msa_path) as f:
        for line in f:
            if line.startswith(">"):
                header = line[1:].strip()
                n_checked += 1
                if date_re.match(header):
                    n_valid += 1
    return n_checked > 0 and n_valid == n_checked


def _reformat_msa_for_delphy(msa_path: Path, metadata_csv: Path, output_path: Path) -> dict:
    """Rewrite MSA headers to 'accession|YYYY-MM-DD' (with optional country code) using metadata CSV.

    If sequences come from 2+ countries, headers become 'accession|CC|YYYY-MM-DD'.
    Returns {"included": N, "excluded_no_date": N}.
    """
    import pandas as pd
    from Bio import SeqIO

    meta = pd.read_csv(metadata_csv, dtype=str)
    # Find relevant columns
    acc_col = None
    for col in meta.columns:
        if re.search(r"^accession$", col, flags=re.I):
            acc_col = col
            break
    date_col = None
    for col in meta.columns:
        if re.search(r"collection.?date", col, flags=re.I):
            date_col = col
            break
    release_col = None
    for col in meta.columns:
        if re.search(r"release.?date", col, flags=re.I):
            release_col = col
            break
    geo_col = None
    for col in meta.columns:
        if re.search(r"geographic.?location", col, flags=re.I):
            geo_col = col
            break

    if not acc_col or (not date_col and not release_col):
        return {"included": 0, "excluded_no_date": 0, "error": "Missing accession/date columns"}

    # Build accession -> date and accession -> country_code mappings
    acc_to_date = {}
    acc_to_cc = {}
    for _, row in meta.iterrows():
        acc = str(row[acc_col]).strip()
        # Date: prefer Collection Date, fallback to Release date
        formatted = None
        if date_col:
            formatted = _extract_and_format_date(row.get(date_col, ""))
        if not formatted and release_col:
            formatted = _extract_and_format_date(row.get(release_col, ""))
        if formatted:
            acc_to_date[acc] = formatted
        # Country code
        if geo_col:
            loc = str(row.get(geo_col, "")).strip()
            country = loc.split(":")[0].strip() if loc else ""
            if country and country != "nan":
                cc = _country_to_alpha2(country)
                if cc:
                    acc_to_cc[acc] = cc

    # Only include country codes if 2+ distinct countries
    unique_countries = set(acc_to_cc.values())
    include_cc = len(unique_countries) >= 2

    included = 0
    excluded = 0
    records_out = []

    for record in SeqIO.parse(str(msa_path), "fasta"):
        acc = record.id.split()[0]
        # Try with and without version suffix
        d = acc_to_date.get(acc) or acc_to_date.get(acc.split(".")[0])
        if d:
            if include_cc:
                cc = acc_to_cc.get(acc) or acc_to_cc.get(acc.split(".")[0], "")
                header = f"{acc}|{cc}|{d}" if cc else f"{acc}|{d}"
            else:
                header = f"{acc}|{d}"
            record.id = header
            record.description = ""
            records_out.append(record)
            included += 1
        else:
            excluded += 1

    with open(output_path, "w") as fh:
        SeqIO.write(records_out, fh, "fasta")

    return {"included": included, "excluded_no_date": excluded}


def _run_delphy(delphy_bin: Path, delphy_fasta: Path, output_dir: Path,
                num_seqs: int, progress_placeholder=None) -> dict | None:
    """Run Delphy inference. Returns dict of output file paths or None on failure."""
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = delphy_fasta.stem.replace("_delphy", "")

    log_file = output_dir / f"{stem}_delphy.log"
    trees_file = output_dir / f"{stem}_delphy.trees"
    dphy_file = output_dir / f"{stem}_delphy.dphy"

    steps = 500_000 * max(1, num_seqs)
    samples = 200
    log_every = max(1, steps // samples)
    tree_every = log_every

    cmd = [
        str(delphy_bin),
        "--v0-in-fasta", str(delphy_fasta),
        "--v0-steps", str(steps),
        "--v0-out-log-file", str(log_file),
        "--v0-out-trees-file", str(trees_file),
        "--v0-out-delphy-file", str(dphy_file),
        "--v0-log-every", str(log_every),
        "--v0-tree-every", str(tree_every),
    ]

    proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                            text=True, bufsize=1)
    output_lines = []
    progress_bar = None
    if progress_placeholder:
        progress_bar = progress_placeholder.progress(0.0, text="Starting Delphy...")
    for line in proc.stdout:
        output_lines.append(line.rstrip())
        if progress_bar:
            # Parse step number from Delphy output (e.g., "... Step 165300000, ...")
            m = re.search(r"Step\s+(\d+)", line)
            if m:
                current_step = int(m.group(1))
                pct = min(current_step / steps, 1.0)
                progress_bar.progress(pct, text=f"Step {current_step:,} / {steps:,} ({pct:.0%}) — t_MRCA: {tm}"
                    if (tm := re.search(r"t_MRCA\s*=\s*([\d-]+)", line)) and (tm := tm.group(1))
                    else f"Step {current_step:,} / {steps:,} ({pct:.0%})")

    ret = proc.wait()
    if ret != 0:
        return None

    return {
        "log": log_file,
        "trees": trees_file,
        "dphy": dphy_file,
    }


def _parse_beast_trees(trees_path: Path):
    """Parse BEAST2 .trees (Nexus) file and return the last tree.

    Delphy's Nexus output contains [&mutations=...] annotations that
    Bio.Phylo's Nexus parser can't handle. We extract the last Newick
    string, strip annotations, build a taxon label map from the
    'translate' block, and parse as plain Newick.
    """
    from Bio import Phylo
    import io

    text = trees_path.read_text()

    # Build translate table: number -> taxon label
    translate_map = {}
    in_translate = False
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.lower().startswith("translate"):
            in_translate = True
            continue
        if in_translate:
            if stripped == ";":
                in_translate = False
                continue
            # e.g. "1 PV549935.1|2025-04-27,"
            parts = stripped.rstrip(",;").split(None, 1)
            if len(parts) == 2:
                translate_map[parts[0]] = parts[1]

    # Find the last "tree STATE_xxx = ..." line
    last_tree_line = None
    for line in text.splitlines():
        if line.strip().startswith("tree "):
            last_tree_line = line.strip()

    if not last_tree_line:
        return None

    # Extract Newick part after "= "
    eq_idx = last_tree_line.index("= ")
    newick_str = last_tree_line[eq_idx + 2:]

    # Strip [&...] annotations
    newick_clean = re.sub(r"\[&[^\]]*\]", "", newick_str)

    # Replace numeric taxon IDs with actual labels
    if translate_map:
        def _replace_taxon(m):
            num = m.group(1)
            return translate_map.get(num, num)
        # Match numbers that appear as leaf labels (preceded by '(' or ',' and followed by ':')
        newick_clean = re.sub(r"(?<=[(,])(\d+)(?=:)", _replace_taxon, newick_clean)

    tree = Phylo.read(io.StringIO(newick_clean), "newick")
    return tree


def _plot_phylo_tree(tree, covered_ids: set[str] | None = None):
    """Convert Bio.Phylo tree to a Plotly figure with optional coverage coloring."""
    import plotly.graph_objects as go

    # Compute x (distance from root) and y (leaf ordering) coordinates
    # Using a simple recursive layout
    tip_count = tree.count_terminals()
    y_counter = [0]  # mutable counter

    node_coords = {}  # clade -> (x, y)

    def _assign_coords(clade, x_start=0):
        branch_len = clade.branch_length or 0
        x = x_start + branch_len

        if clade.is_terminal():
            y = y_counter[0]
            y_counter[0] += 1
            node_coords[id(clade)] = (x, y)
        else:
            child_ys = []
            for child in clade.clades:
                _assign_coords(child, x)
                child_ys.append(node_coords[id(child)][1])
            y = sum(child_ys) / len(child_ys)
            node_coords[id(clade)] = (x, y)

    _assign_coords(tree.root)

    # Build line segments (horizontal + vertical branches)
    line_x = []
    line_y = []

    def _draw_branches(clade):
        cx, cy = node_coords[id(clade)]
        for child in clade.clades:
            ccx, ccy = node_coords[id(child)]
            # Horizontal line from parent x to child x at child y
            line_x.extend([cx, ccx, None])
            line_y.extend([ccy, ccy, None])
            # Vertical line at parent x connecting children
            line_x.extend([cx, cx, None])
            line_y.extend([cy, ccy, None])
            _draw_branches(child)

    _draw_branches(tree.root)

    # Tip markers and labels
    tip_x, tip_y, tip_text, tip_color = [], [], [], []
    tip_labels = []
    # Pre-build version-stripped lookup for covered_ids
    if covered_ids is not None:
        _covered_stripped = {c.split(".")[0] for c in covered_ids} | covered_ids
    else:
        _covered_stripped = set()
    for tip in tree.get_terminals():
        x, y = node_coords[id(tip)]
        tip_x.append(x)
        tip_y.append(y)
        name = tip.name or ""
        # Parse accession, optional country code, and date from "accession|CC|date" or "accession|date"
        parts = name.split("|")
        acc = parts[0] if parts else name
        if len(parts) == 3:
            cc = parts[1]
            date_str = parts[2]
        elif len(parts) == 2:
            date_str = parts[1]
            cc = ""
        else:
            date_str = ""
            cc = ""
        label = acc
        if date_str:
            label += f"  {date_str}"
        if cc:
            label += f"  {cc}"
        tip_labels.append(f"  {label}")

        if covered_ids is not None:
            is_covered = acc in _covered_stripped or acc.split(".")[0] in _covered_stripped
            tip_color.append("#2ecc71" if is_covered else "#e74c3c")
            status = "Covered" if is_covered else "Not covered"
            tip_text.append(f"{acc}<br>{date_str}<br>{status}")
        else:
            tip_color.append("#3498db")
            tip_text.append(f"{acc}<br>{date_str}" if date_str else acc)

    # Compute x range for label offset
    all_tip_x = [v for v in tip_x if v is not None]
    x_range = max(all_tip_x) - min(all_tip_x) if all_tip_x else 1
    label_offset = x_range * 0.02

    fig = go.Figure()

    # Branch lines — rectangular cladogram style
    fig.add_trace(go.Scatter(
        x=line_x, y=line_y,
        mode="lines",
        line=dict(color="#555", width=1.5),
        hoverinfo="skip",
        showlegend=False,
    ))

    # Tip markers
    fig.add_trace(go.Scatter(
        x=tip_x, y=tip_y,
        mode="markers",
        marker=dict(
            size=7,
            color=tip_color,
            line=dict(width=0.5, color="white"),
        ),
        text=tip_text,
        hoverinfo="text",
        showlegend=False,
    ))

    # Tip labels (right of markers)
    fig.add_trace(go.Scatter(
        x=[x + label_offset for x in tip_x],
        y=tip_y,
        mode="text",
        text=tip_labels,
        textposition="middle right",
        textfont=dict(size=10, color="#333"),
        hoverinfo="skip",
        showlegend=False,
        cliponaxis=False,
    ))

    # Add legend entries for coverage
    if covered_ids is not None:
        n_covered = sum(1 for c in tip_color if c == "#2ecc71")
        n_uncovered = sum(1 for c in tip_color if c == "#e74c3c")
        fig.add_trace(go.Scatter(
            x=[None], y=[None], mode="markers",
            marker=dict(size=10, color="#2ecc71"),
            name=f"Covered ({n_covered})",
        ))
        fig.add_trace(go.Scatter(
            x=[None], y=[None], mode="markers",
            marker=dict(size=10, color="#e74c3c"),
            name=f"Not covered ({n_uncovered})",
        ))

    height = max(400, tip_count * 18)
    # Right margin for tip labels that overflow the plot area
    max_label_len = max((len(l) for l in tip_labels), default=0)
    right_margin = max(20, max_label_len * 5)

    fig.update_layout(
        title=dict(text="Phylogenetic Tree (Delphy)", font=dict(size=16)),
        xaxis_title="Evolutionary distance",
        yaxis=dict(showticklabels=False, showgrid=False, zeroline=False),
        xaxis=dict(showgrid=False, zeroline=False),
        height=height,
        margin=dict(l=20, r=right_margin, t=60, b=40),
        plot_bgcolor="white",
        legend=dict(orientation="h", yanchor="bottom", y=0.1, font=dict(size=13)),
    )

    return fig


def _get_activity_scores_from_eval(eval_re_path: Path, pname_f: str,
                                   pname_r: str) -> dict[str, float]:
    """Extract per-sequence activity scores from .eval.re file for a primer pair.

    The .eval.re file is a matrix: rows are primer pairs (pname_f, pname_r),
    columns are accession IDs, values are regressor activity scores.

    Returns {accession: score}.
    """
    import pandas as pd

    df = pd.read_csv(eval_re_path)
    pair_row = df[(df["pname_f"] == pname_f) & (df["pname_r"] == pname_r)]
    if pair_row.empty:
        return {}

    row = pair_row.iloc[0]
    scores = {}
    for acc in df.columns[2:]:
        try:
            scores[acc] = float(row[acc])
        except (ValueError, KeyError):
            pass
    return scores


# ---------------------------------------------------------------------------
# Tab 5: Results
# ---------------------------------------------------------------------------

def _tab_results():
    import io
    import zipfile

    st.header("Results")

    st.markdown(
        "For help interpreting output columns, see the "
        "[Output Interpretation Guide]"
        "(https://github.com/broadinstitute/qprimer_designer/blob/main/docs/output_interpretation_guide.md)."
    )

    run_id = st.session_state.get("run_id", "")
    workflow = st.session_state.get("workflow", "design")

    # Allow browsing past runs
    past_runs = []
    if workflow == "design" and RUNS_DIR.exists():
        past_runs = sorted(
            [d.name for d in RUNS_DIR.iterdir() if d.is_dir()],
            reverse=True,
        )
    elif workflow in ("evaluate", "monitor"):
        browse_dir = MONITOR_DIR if workflow == "monitor" else RUNS_DIR
        if browse_dir.exists():
            past_runs = sorted(
                [d.name for d in browse_dir.iterdir() if d.is_dir()],
                reverse=True,
            )

    if past_runs:
        # Default to current run_id if it exists in the list
        default_idx = 0
        if run_id in past_runs:
            default_idx = past_runs.index(run_id)
        selected_run = st.selectbox(
            "Select run",
            options=past_runs,
            index=default_idx,
            key="results_run_selector",
        )
        if selected_run:
            run_id = selected_run
            st.session_state["run_id"] = run_id

    # --- Final CSV files (design workflow) ---
    if workflow == "design":
        st.subheader("Final output files")

        run_dir = RUNS_DIR / run_id if run_id else None
        if run_dir and run_dir.exists():
            csvs = sorted(run_dir.glob("*_final.csv"), reverse=True)
        else:
            csvs = []

        if csvs:
            csv_labels = [c.name for c in csvs]

            selected_csv_label = st.selectbox(
                "Select CSV to view",
                options=csv_labels,
                key="result_csv",
            )
            if selected_csv_label:
                import pandas as pd

                csv_path = run_dir / selected_csv_label
                try:
                    df = pd.read_csv(csv_path)
                    st.write(f"**{len(df)} primer pairs**")

                    mc1, mc2, mc3 = st.columns(3)
                    mc1.metric("Pairs", len(df))
                    if "coverage" in df.columns:
                        mc2.metric("Avg coverage", f"{df['coverage'].mean():.2f}")
                    if "score" in df.columns:
                        mc3.metric("Best score", f"{df['score'].max():.4f}")

                    st.dataframe(df, use_container_width=True)

                    st.download_button(
                        "Download CSV",
                        data=csv_path.read_bytes(),
                        file_name=csv_path.name,
                        mime="text/csv",
                    )
                except Exception as exc:
                    st.error(f"Error reading {selected_csv_label}: {exc}")
        else:
            st.info("No CSV results for this run yet.")

        # --- Phylogenetic Tree (Delphy) ---
        st.divider()
        st.subheader("Phylogenetic Tree")

        # Check prerequisites: MSA file and metadata
        MSA_DIR = PROJECT_ROOT / "target_seqs" / "msa"
        target_names = st.session_state.get("targets", [])
        # Fallback: discover targets from MSA directory
        if not target_names and MSA_DIR.exists():
            target_names = sorted(
                p.stem for p in MSA_DIR.glob("*.aln")
                if not p.stem.endswith("_delphy")
            )

        if not target_names:
            st.info("No target sequences available for tree generation.")
        else:
            tree_target = st.selectbox(
                "Select target for phylogenetic tree",
                options=target_names,
                key="delphy_target",
            )

            if tree_target:
                msa_path = MSA_DIR / f"{tree_target}.aln"
                meta_path = FASTA_DIR / f"{tree_target}_metadata.csv"
                tree_dir = run_dir / "trees" if run_dir else None

                if not msa_path.exists():
                    st.warning(f"MSA file not found: `{msa_path.name}`. Run the pipeline first.")
                else:
                    has_meta = meta_path.exists()
                    headers_ready = _msa_has_delphy_headers(msa_path)

                    if not has_meta and not headers_ready:
                        st.warning(
                            f"Metadata file not found: `{meta_path.name}`. "
                            "Re-fetch sequences to generate metadata for Delphy, "
                            "or provide an MSA with headers in `accession|YYYY-MM-DD` format."
                        )
                    else:
                        # Check if tree already exists
                        dphy_file = tree_dir / f"{tree_target}_delphy.dphy" if tree_dir else None
                        trees_file = tree_dir / f"{tree_target}_delphy.trees" if tree_dir else None
                        tree_exists = trees_file and trees_file.exists()

                        if not tree_exists:
                            st.markdown(
                                "Generate a phylogenetic tree using [Delphy](https://github.com/broadinstitute/delphy) "
                                "to visualize primer coverage across the evolutionary tree."
                            )
                            if st.button("Generate Phylogenetic Tree", type="primary"):
                                delphy_bin = _get_delphy_binary()
                                if delphy_bin is None:
                                    st.error(
                                        f"Delphy is not available for this platform "
                                        f"({_platform.system()} {_platform.machine()})."
                                    )
                                else:
                                    delphy_fasta = MSA_DIR / f"{tree_target}_delphy.afa"

                                    if headers_ready:
                                        # MSA headers already in Delphy format — use directly
                                        import shutil as _shutil
                                        _shutil.copy2(msa_path, delphy_fasta)
                                        with open(msa_path) as _f:
                                            num_seqs = sum(1 for _l in _f if _l.startswith(">"))
                                        stats = {"included": num_seqs, "excluded_no_date": 0}
                                        st.info(f"MSA headers already in Delphy format ({num_seqs} sequences).")
                                    else:
                                        with st.spinner("Reformatting MSA headers for Delphy..."):
                                            stats = _reformat_msa_for_delphy(msa_path, meta_path, delphy_fasta)

                                    if stats.get("error"):
                                        st.error(f"Header reformat error: {stats['error']}")
                                    elif stats["included"] == 0:
                                        st.error(
                                            "No sequences with valid collection dates (YYYY-MM-DD). "
                                            "Delphy requires complete dates."
                                        )
                                    else:
                                        if not headers_ready:
                                            st.info(
                                                f"Reformatted: {stats['included']} sequences included, "
                                                f"{stats['excluded_no_date']} excluded (missing dates)."
                                            )
                                        progress_area = st.empty()
                                        with st.spinner("Running Delphy inference... This may take a while."):
                                            result = _run_delphy(
                                                delphy_bin, delphy_fasta, tree_dir,
                                                stats["included"], progress_area,
                                            )

                                        if result is None:
                                            st.error("Delphy inference failed. Check logs.")
                                        else:
                                            st.success("Delphy inference complete!")
                                            st.rerun()

                        if tree_exists:
                            # Parse and display tree
                            try:
                                tree = _parse_beast_trees(trees_file)
                                if tree is None:
                                    st.warning("Could not parse tree from Delphy output.")
                                else:
                                    # Two-column layout: controls (left) | tree (right)
                                    ctrl_col, tree_col = st.columns([1, 3])

                                    covered_ids = None
                                    with ctrl_col:
                                        # Find eval.re files for this target
                                        eval_re_files = sorted(run_dir.glob(
                                            f"**/*{tree_target}*.eval.re"
                                        )) if run_dir and run_dir.exists() else []

                                        if eval_re_files and csvs:
                                            import pandas as pd
                                            csv_path = run_dir / csv_labels[0]
                                            final_df = pd.read_csv(csv_path)
                                            if len(final_df) > 0:
                                                pair_options = [
                                                    f"{r['pname_f']} / {r['pname_r']}"
                                                    for _, r in final_df.iterrows()
                                                    if "pname_f" in r and "pname_r" in r
                                                ]
                                                if pair_options:
                                                    selected_pair = st.selectbox(
                                                        "Primer pair",
                                                        options=["(none)"] + pair_options,
                                                        key="delphy_primer_pair",
                                                    )
                                                    if selected_pair != "(none)":
                                                        parts = selected_pair.split(" / ")
                                                        pf, pr = parts[0], parts[1]
                                                        scores = _get_activity_scores_from_eval(
                                                            eval_re_files[0], pf, pr,
                                                        )
                                                        if scores:
                                                            cutoff = st.slider(
                                                                "Activity score cutoff",
                                                                min_value=0.0,
                                                                max_value=1.1,
                                                                value=0.5,
                                                                step=0.05,
                                                                key="delphy_activity_cutoff",
                                                                help="Sequences with activity ≥ cutoff are considered covered.",
                                                            )
                                                            covered_ids = {
                                                                acc for acc, s in scores.items()
                                                                if s >= cutoff
                                                            }
                                                            n_total = len(scores)
                                                            st.metric(
                                                                "Coverage",
                                                                f"{len(covered_ids)} / {n_total}",
                                                            )

                                        # Download .dphy for Delphy web viewer
                                        if dphy_file and dphy_file.exists():
                                            st.download_button(
                                                "Download .dphy (for delphy.fathom.info)",
                                                data=dphy_file.read_bytes(),
                                                file_name=dphy_file.name,
                                                mime="application/octet-stream",
                                            )

                                    with tree_col:
                                        fig = _plot_phylo_tree(tree, covered_ids)
                                        st.plotly_chart(fig, use_container_width=True)
                            except Exception as exc:
                                st.error(f"Error displaying tree: {exc}")


    # --- Evaluate reports (evaluate/monitor workflows only) ---
    if workflow in ("evaluate", "monitor"):
        st.subheader("Evaluation reports")

        if workflow == "monitor" and run_id and (MONITOR_DIR / run_id).exists():
            xlsx_files = sorted((MONITOR_DIR / run_id).glob("*.xlsx"), reverse=True)
        elif run_id and (RUNS_DIR / run_id).exists():
            xlsx_files = sorted((RUNS_DIR / run_id).rglob("*.xlsx"), reverse=True)
        else:
            xlsx_files = []

    if workflow in ("evaluate", "monitor") and xlsx_files:
        from openpyxl import load_workbook

        # Prepare zip for "download all" (built once, used below)
        zip_buf = None
        if len(xlsx_files) > 1:
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for xf in xlsx_files:
                    zf.write(xf, xf.name)

        # Individual report viewer
        xlsx_labels = [xf.stem for xf in xlsx_files]

        selected_label = st.selectbox(
            "Select primer set",
            options=xlsx_labels,
            key="result_xlsx",
        )
        if selected_label:
            xf = next(f for f in xlsx_files if f.stem == selected_label)
            try:
                wb = load_workbook(xf, read_only=True)
                ws = wb["summary"]
                rows = list(ws.iter_rows(values_only=True))
                wb.close()

                # Parse summary sheet sections
                sections: dict[str, list] = {}
                current_section = "_header"
                sections[current_section] = []
                for row in rows:
                    first = row[0]
                    if first in ("Dimerization", "Sensitivity", "Specificity"):
                        current_section = first
                        sections[current_section] = []
                    else:
                        sections[current_section].append(row)

                # --- Sequences ---
                header_rows = sections.get("_header", [])
                if len(header_rows) >= 2:
                    seq_labels = [v for v in header_rows[0] if v]
                    seq_values = list(header_rows[1])
                    seq_parts = []
                    for lbl, val in zip(seq_labels, seq_values):
                        if val:
                            seq_parts.append(f"**{lbl}:** `{val}`")
                    st.markdown("  \n".join(seq_parts))

                # --- Sensitivity & Specificity ---
                metric_lines = []

                sens_rows = sections.get("Sensitivity", [])
                sens_data = [r for r in sens_rows if r[0] is not None]
                if sens_data:
                    for r in sens_data:
                        target, coverage, act_mean = r[0], r[1], r[2]
                        line = f"**Sensitivity** ({target}): **{coverage}** sequences covered"
                        if act_mean is not None:
                            line += f" (mean activity {act_mean})"
                        metric_lines.append(line)
                else:
                    metric_lines.append("**Sensitivity:** no data")

                spec_rows = sections.get("Specificity", [])
                spec_data = [r for r in spec_rows if r[0] is not None]
                if spec_data:
                    for r in spec_data:
                        target, coverage, act_mean = r[0], r[1], r[2]
                        line = f"**Specificity** ({target}): **{coverage}** off-target amplifications predicted"
                        if act_mean is not None:
                            line += f" (mean activity {act_mean})"
                        metric_lines.append(line)
                else:
                    metric_lines.append("**Specificity:** no off-target evaluated")

                st.markdown("  \n".join(metric_lines))

            except Exception as exc:
                st.error(f"Error reading {selected_label}: {exc}")

            # Download buttons side by side
            col_dl1, col_dl2 = st.columns(2)
            with col_dl1:
                st.download_button(
                    f"Download {xf.name}",
                    data=xf.read_bytes(),
                    file_name=xf.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_{selected_label}",
                    use_container_width=True,
                )
            with col_dl2:
                if zip_buf is not None:
                    st.download_button(
                        f"Download all ({len(xlsx_files)} files)",
                        data=zip_buf.getvalue(),
                        file_name=f"evaluate_{run_id}.zip",
                        mime="application/zip",
                        key="dl_all_xlsx",
                        use_container_width=True,
                    )
    elif workflow in ("evaluate", "monitor"):
        st.info("No evaluation reports for this run.")

    st.divider()

    # --- Cleanup ---
    st.subheader("Cleanup")
    if st.button("Delete all pipeline outputs", type="secondary"):
        result = subprocess.run(
            ["snakemake", "-s", "Snakefile", "--delete-all-output", "--cores", "1"],
            capture_output=True,
            text=True,
            cwd=str(PROJECT_ROOT),
        )
        if result.returncode == 0:
            st.success("Pipeline outputs deleted.")
        else:
            st.error(f"Cleanup failed:\n{result.stderr}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def _page_select_target():
    """Common page for selecting pathogen target sequences."""
    _render_workflow_progress("select_target")

    workflow = st.session_state.get("workflow", "design")
    workflow_label = {"design": "Design", "evaluate": "Evaluate", "monitor": "Monitor"}[workflow]

    st.header(f"{workflow_label} — Select Target Sequences")
    st.markdown(
        "Which pathogen do you want to detect? "
        "Select target sequences from the list below, or add new ones."
    )

    st.divider()

    # --- Select from existing FASTA files ---
    fastas = _available_fasta()

    if fastas:
        # Build display labels with sequence count and file size
        labels = []
        for stem in fastas:
            info = _fasta_info(stem)
            labels.append(f"{stem}  ({info['seqs']} seqs, {_format_size(info['size'])})")

        # Restore selectbox value from stored targets when returning to this page
        stored_targets = st.session_state.get("targets", [])
        if stored_targets and "target_select" not in st.session_state:
            for label in labels:
                if label.startswith(stored_targets[0] + "  ("):
                    st.session_state.target_select = label
                    break

        selected_label = st.selectbox(
            "Select a target sequence",
            options=[""] + labels,
            format_func=lambda x: "Choose a pathogen..." if x == "" else x,
            key="target_select",
        )

        if selected_label:
            # Extract stem name (everything before the first double-space)
            selected_stem = selected_label.split("  (")[0]
            st.session_state.targets = [selected_stem]
        else:
            st.session_state.targets = []
    else:
        st.info("No FASTA files found. Upload or fetch sequences below.")

    st.divider()

    # --- Add new sequences ---
    st.subheader("Don't see your pathogen?")

    add_col1, add_col2 = st.columns(2)

    with add_col1:
        st.markdown("**Upload FASTA file**")
        uploaded = st.file_uploader(
            "Upload FASTA file(s)",
            type=["fa", "fasta", "fna"],
            accept_multiple_files=True,
            key="fasta_upload",
            label_visibility="collapsed",
        )
        if uploaded:
            already_saved = st.session_state.get("_fasta_saved_ids", set())
            new_saved = set()
            for f in uploaded:
                fid = (f.name, f.size)
                if fid in already_saved:
                    continue
                p = Path(f.name)
                if p.suffix.lower() not in (".fa", ".fasta", ".fna"):
                    st.error(f"Unsupported extension: {p.suffix}")
                    continue
                dest = _safe_path_under(FASTA_DIR, p.stem + ".fa")
                if not dest:
                    st.error(f"Invalid filename: {f.name}")
                    continue
                _save_fasta_bytes(dest, f.getvalue())
                for ext in (".fasta", ".fna"):
                    old = FASTA_DIR / (p.stem + ext)
                    if old.exists():
                        old.unlink()
                st.success(f"Saved {dest.name}")
                new_saved.add(fid)
            if new_saved:
                st.session_state["_fasta_saved_ids"] = already_saved | new_saved
                st.rerun()

    with add_col2:
        st.markdown("**Fetch from NCBI Virus (using gget)**")
        _render_fetch_ui("fetch")

    st.divider()

    # --- Continue button ---
    selected = st.session_state.get("targets", [])
    if st.button(
        "Continue →",
        disabled=not selected,
        type="primary",
        use_container_width=True,
    ):
        _navigate("select_offtarget")
        st.rerun()

    if not selected:
        st.caption("Select a target sequence to continue.")


def _render_fetch_ui(prefix: str, monitor: bool = False):
    """Render fetch-from-NCBI UI. prefix is used to namespace session state keys."""
    _OTHER_OPTION = "Other (enter TaxID manually)"
    virus_options = sorted([f"{name}  (TaxID: {tid})" for name, tid in VIRUS_PRESETS])
    virus_options.append(_OTHER_OPTION)

    selected_virus = st.selectbox(
        "Virus",
        options=virus_options,
        index=None,
        placeholder="Type a virus name to search...",
        key=f"{prefix}_virus_select",
    )

    taxid = ""
    auto_name = ""
    if selected_virus == _OTHER_OPTION:
        taxid = st.text_input(
            "Taxonomy ID",
            placeholder="e.g., 2697049",
            key=f"{prefix}_taxid_manual",
        )
    elif selected_virus:
        taxid = selected_virus.split("TaxID: ")[1].rstrip(")")
        auto_name = selected_virus.split("  (TaxID:")[0]

    # Subtype input (e.g., H5N1 for Influenza A)
    st.text_input(
        "Subtype",
        placeholder="e.g., H5N1, H5N*, H3N2 (wildcards supported)",
        key=f"{prefix}_subtype_filter",
        help="Filter fetched sequences by subtype in the FASTA header. "
             "Useful for Influenza A where all subtypes (H1N1, H3N2, H5N1, ...) are fetched together. "
             "Supports wildcards: * matches anything, ? matches one character.",
    )

    # Build auto target name from virus + subtype
    subtype_val = st.session_state.get(f"{prefix}_subtype_filter", "").strip()
    auto_target = auto_name
    if not auto_name and taxid:
        for name, tid in VIRUS_PRESETS:
            if str(tid) == taxid.strip():
                auto_target = name
                break

    # Append subtype to target name (e.g., "Influenza_A" → "Influenza_A_H5N1")
    if subtype_val:
        subtype_clean = re.sub(r"[^\w\-]", "_", subtype_val)
        subtype_clean = re.sub(r"_+", "_", subtype_clean).strip("_")
        if subtype_clean:
            auto_target = f"{auto_target}_{subtype_clean}" if auto_target else subtype_clean

    # Track (virus + subtype) together for auto-name updates
    prev_auto = st.session_state.get(f"_prev_auto_{prefix}", "")
    if auto_target != prev_auto:
        st.session_state[f"_prev_auto_{prefix}"] = auto_target
        sanitized = re.sub(r"[^\w\-]", "_", auto_target)
        sanitized = re.sub(r"_+", "_", sanitized).strip("_")
        st.session_state[f"{prefix}_target_name"] = sanitized

    target_name = st.text_input(
        "Target name",
        placeholder="e.g., SARS-CoV-2",
        key=f"{prefix}_target_name",
        help="Name for the output FASTA file.",
    )

    nuc_completeness = st.selectbox(
        "Nucleotide completeness",
        options=["", "complete", "partial"],
        format_func=lambda x: "Any" if x == "" else x,
        key=f"{prefix}_nuc_completeness",
    )

    segment = st.text_input(
        "Gene segment",
        placeholder="e.g., 1, 2, 3 (leave blank for non-segmented)",
        key=f"{prefix}_segment",
    )

    col_len1, col_len2 = st.columns(2)
    with col_len1:
        st.number_input(
            "Min sequence length (bp)", min_value=0, value=None, step=100,
            key=f"{prefix}_min_seq_length",
        )
    with col_len2:
        st.number_input(
            "Max sequence length (bp)", min_value=0, value=None, step=100,
            key=f"{prefix}_max_seq_length",
        )

    col_d1, col_d2 = st.columns(2)
    _date_min = date(1900, 1, 1)
    with col_d1:
        st.date_input("Min release date", value=None, min_value=_date_min,
                       key=f"{prefix}_min_release_date")
    with col_d2:
        if monitor:
            st.date_input("Max release date", value=date.today(), min_value=_date_min,
                          disabled=True, key=f"{prefix}_max_release_date")
        else:
            st.date_input("Max release date", value=None, min_value=_date_min,
                           key=f"{prefix}_max_release_date")

    st.text_input("Geographic location", placeholder="e.g., USA, Africa, Nigeria",
                   key=f"{prefix}_geo_location")
    st.text_input("Host", placeholder="e.g., human, Homo sapiens", key=f"{prefix}_host")

    with st.expander("Additional parameters"):
        st.number_input("Max sequences", min_value=1, max_value=100000, value=None,
                        step=100, key=f"{prefix}_limit")
        col_c1, col_c2 = st.columns(2)
        with col_c1:
            st.date_input("Min collection date", value=None, min_value=_date_min,
                           key=f"{prefix}_min_collection_date")
        with col_c2:
            st.date_input("Max collection date", value=None, min_value=_date_min,
                           key=f"{prefix}_max_collection_date")
        st.number_input("Max ambiguous characters", min_value=0, value=10, step=1,
                        key=f"{prefix}_max_ambiguous_chars")
        st.checkbox("RefSeq only", key=f"{prefix}_refseq_only")

    if st.button("Fetch sequences", disabled=not taxid or not target_name, key=f"{prefix}_fetch_btn"):
        gget_bin = shutil.which("gget")
        if not gget_bin:
            env_bin = Path(sys.executable).parent
            gget_candidate = env_bin / "gget"
            if gget_candidate.exists():
                gget_bin = str(gget_candidate)
        if not gget_bin:
            st.error("'gget' is not installed. Run: `pip install gget`")
        else:
            gget_tmp = _safe_path_under(FASTA_DIR, f".gget_tmp_{target_name}")
            if not gget_tmp:
                st.error("Invalid target name.")
                return
            gget_tmp.mkdir(parents=True, exist_ok=True)

            fetch_params = {
                "nuc_completeness": st.session_state.get(f"{prefix}_nuc_completeness", "complete"),
                "segment": st.session_state.get(f"{prefix}_segment", ""),
                "min_seq_length": st.session_state.get(f"{prefix}_min_seq_length"),
                "max_seq_length": st.session_state.get(f"{prefix}_max_seq_length"),
                "min_release_date": st.session_state.get(f"{prefix}_min_release_date"),
                "max_release_date": st.session_state.get(f"{prefix}_max_release_date"),
                "geo_location": st.session_state.get(f"{prefix}_geo_location", ""),
                "host": st.session_state.get(f"{prefix}_host", ""),
                "refseq_only": st.session_state.get(f"{prefix}_refseq_only"),
                "min_collection_date": st.session_state.get(f"{prefix}_min_collection_date"),
                "max_collection_date": st.session_state.get(f"{prefix}_max_collection_date"),
                "max_ambiguous_chars": st.session_state.get(f"{prefix}_max_ambiguous_chars"),
            }

            # Split multi-country geo_location into individual locations
            raw_geo = fetch_params.get("geo_location", "").strip()
            geo_locations = [g.strip() for g in raw_geo.split(",") if g.strip()] if raw_geo else [""]

            _FETCH_STEPS = [
                ("STEP 1", "Validating input"),
                ("STEP 2", "Checking optimized pathways"),
                ("STEP 3", "Fetching metadata from NCBI"),
                ("STEP 4", "Applying metadata filters"),
                ("STEP 5", "Downloading sequences"),
                ("STEP 6", "Applying sequence filters"),
                ("STEP 7", "Saving output files"),
                ("STEP 8", "Fetching GenBank metadata"),
            ]

            progress_area = st.empty()
            detail_area = st.empty()
            status_area = st.empty()

            def _render_progress(current_step, detail="", loc_label=""):
                lines = []
                if loc_label:
                    lines.append(f"**{loc_label}**")
                for i, (_, label) in enumerate(_FETCH_STEPS):
                    if i < current_step:
                        lines.append(f"✅  {label}")
                    elif i == current_step:
                        lines.append(f"🔵  {label}...")
                    else:
                        lines.append(f"⚪  {label}")
                progress_area.markdown("\n".join(lines))
                if detail:
                    detail_area.caption(detail)

            def _run_single_fetch(cmd, loc_label=""):
                """Run a single fetch command, updating progress. Returns True on success."""
                if not isinstance(cmd, list) or not all(isinstance(arg, str) for arg in cmd):
                    st.error("Internal error: invalid command structure")
                    return False

                _render_progress(-1, loc_label=loc_label)
                env = os.environ.copy()
                env["PYTHONUNBUFFERED"] = "1"
                proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                        text=True, env=env)
                current_step = -1
                detail = ""
                start_time = time.time()
                timeout = 1800

                while proc.poll() is None:
                    if time.time() - start_time > timeout:
                        proc.kill()
                        st.error("Fetch timed out after 30 minutes.")
                        return False
                    line = proc.stdout.readline()
                    if line:
                        for i, (key, _) in enumerate(_FETCH_STEPS):
                            if key in line:
                                current_step = i
                                detail = ""
                                break
                        if "retrieved" in line and "records" in line:
                            detail = line.split("INFO - ")[-1].strip() if "INFO - " in line else ""
                        elif "Downloading sequences for" in line:
                            detail = line.split("INFO - ")[-1].strip() if "INFO - " in line else ""
                        elif "Received" in line and "sequences" in line:
                            detail = line.split("INFO - ")[-1].strip() if "INFO - " in line else ""
                        elif "PROCESS COMPLETED" in line:
                            current_step = len(_FETCH_STEPS)
                        _render_progress(current_step, detail, loc_label=loc_label)
                    elapsed = int(time.time() - start_time)
                    mins, secs = divmod(elapsed, 60)
                    status_area.caption(f"Elapsed: {mins}m {secs}s")

                remaining = proc.stdout.read()
                if remaining and "PROCESS COMPLETED" in remaining:
                    _render_progress(len(_FETCH_STEPS), loc_label=loc_label)

                return proc.returncode == 0

            try:
                all_fasta_parts = []
                all_meta_parts = []
                multi = len(geo_locations) > 1

                for loc_idx, geo_loc in enumerate(geo_locations):
                    loc_label = f"Fetching {loc_idx + 1}/{len(geo_locations)}: {geo_loc}" if multi else ""
                    # Use a sub-tmp dir per location for multi-country
                    if multi:
                        sub_tmp = gget_tmp / f"loc_{loc_idx}"
                        sub_tmp.mkdir(parents=True, exist_ok=True)
                        validated_out_dir = os.path.realpath(str(sub_tmp))
                    else:
                        validated_out_dir = os.path.realpath(str(gget_tmp))

                    loc_params = dict(fetch_params, geo_location=geo_loc)
                    cmd = _build_fetch_command(gget_bin, taxid, validated_out_dir, loc_params)
                    if not cmd:
                        st.error("TaxID must be numeric.")
                        return

                    ok = _run_single_fetch(cmd, loc_label=loc_label)
                    if not ok:
                        if not multi:
                            st.error("gget fetch failed.")
                            return
                        st.warning(f"Fetch failed for {geo_loc}, skipping.")
                        continue

                    out_dir_path = Path(validated_out_dir)
                    fastas = list(out_dir_path.glob("*.fa")) + list(out_dir_path.glob("*.fasta"))
                    all_fasta_parts.extend(fastas)
                    # Collect metadata CSVs for Delphy integration
                    meta_csvs = list(out_dir_path.glob("*_metadata.csv"))
                    all_meta_parts.extend(meta_csvs)

                if not all_fasta_parts:
                    st.warning("No FASTA output found. The query may have returned no results.")
                else:
                    dest = _safe_path_under(FASTA_DIR, f"{target_name}.fa")
                    if not dest:
                        st.error("Invalid target name. Please use letters, numbers, '.', '_' or '-'.")
                        detail_area.empty()
                        return
                    # Merge all FASTA parts into one file
                    with open(dest, "w") as fout:
                        for fa in all_fasta_parts:
                            with open(fa) as fin:
                                content = fin.read()
                                fout.write(content)
                                if not content.endswith("\n"):
                                    fout.write("\n")
                    # Merge metadata CSVs for Delphy integration
                    if all_meta_parts:
                        import pandas as pd
                        meta_dest = FASTA_DIR / f"{target_name}_metadata.csv"
                        meta_dfs = [pd.read_csv(m) for m in all_meta_parts]
                        pd.concat(meta_dfs, ignore_index=True).to_csv(meta_dest, index=False)

                    # Filter by subtype if specified
                    from qprimer_designer.adapt_cli import _filter_fasta_by_subtype, _deduplicate_fasta
                    subtype = st.session_state.get(f"{prefix}_subtype_filter", "").strip()
                    n_subtype_removed = 0
                    if subtype:
                        n_subtype_removed = _filter_fasta_by_subtype(dest, subtype)
                    n_deduped = _deduplicate_fasta(dest)
                    with open(dest) as f:
                        n_seqs = sum(1 for line in f if line.startswith(">"))
                    msg = f"Downloaded {n_seqs} unique sequences → `{dest.name}`"
                    if multi:
                        msg += f" (from {len(geo_locations)} locations)"
                    if n_subtype_removed > 0:
                        msg += f" ({n_subtype_removed} filtered by subtype '{subtype}')"
                    if n_deduped > 0:
                        msg += f" ({n_deduped} duplicates removed)"
                    st.success(msg)
                    detail_area.empty()
                    st.rerun()
            except Exception as exc:
                st.error(f"Fetch error: {exc}")
            finally:
                shutil.rmtree(gget_tmp, ignore_errors=True)


def _page_select_offtarget():
    """Page for selecting off-target (cross-reactivity) and host sequences."""
    _render_workflow_progress("select_offtarget")

    workflow = st.session_state.get("workflow", "design")
    workflow_label = {"design": "Design", "evaluate": "Evaluate", "monitor": "Monitor"}[workflow]

    st.header(f"{workflow_label} — Off-Target & Host Sequences")
    st.markdown(
        "Optionally specify sequences to check for cross-reactivity or host background. "
        "You can skip this step if not needed."
    )

    st.divider()

    fastas = _available_fasta()
    # Exclude already-selected targets from the options
    selected_targets = st.session_state.get("targets", [])
    offtarget_options = [s for s in fastas if s not in selected_targets]

    # --- Cross-reactivity sequences ---
    st.subheader("Cross-reactivity sequences")
    st.caption("Other pathogens you want to avoid detecting (off-target amplification)")

    if offtarget_options:
        cross_labels = []
        for stem in offtarget_options:
            info = _fasta_info(stem)
            cross_labels.append(f"{stem}  ({info['seqs']} seqs, {_format_size(info['size'])})")

        # Restore multiselect value from stored cross when returning to this page
        stored_cross = st.session_state.get("cross", [])
        if stored_cross and "cross_select" not in st.session_state:
            restored = [l for l in cross_labels if l.split("  (")[0] in stored_cross]
            if restored:
                st.session_state.cross_select = restored

        selected_cross = st.multiselect(
            "Select cross-reactivity sequences",
            options=cross_labels,
            key="cross_select",
            placeholder="Choose pathogens to avoid...",
        )
        st.session_state.cross = [label.split("  (")[0] for label in selected_cross]
    else:
        st.info("No additional FASTA files available.")
        st.session_state.cross = []

    # Add new cross-reactivity sequences
    with st.expander("Add new cross-reactivity sequences"):
        add_cross_col1, add_cross_col2 = st.columns(2)

        with add_cross_col1:
            st.markdown("**Upload FASTA file**")
            uploaded_cross = st.file_uploader(
                "Upload FASTA",
                type=["fa", "fasta", "fna"],
                accept_multiple_files=True,
                key="cross_fasta_upload",
                label_visibility="collapsed",
            )
            if uploaded_cross:
                already_saved = st.session_state.get("_fasta_saved_ids", set())
                new_saved = set()
                for f in uploaded_cross:
                    fid = (f.name, f.size)
                    if fid in already_saved:
                        continue
                    p = Path(f.name)
                    if p.suffix.lower() not in (".fa", ".fasta", ".fna"):
                        st.error(f"Unsupported extension: {p.suffix}")
                        continue
                    dest = _safe_path_under(FASTA_DIR, p.stem + ".fa")
                    if not dest:
                        st.error(f"Invalid filename: {f.name}")
                        continue
                    _save_fasta_bytes(dest, f.getvalue())
                    for ext in (".fasta", ".fna"):
                        old = FASTA_DIR / (p.stem + ext)
                        if old.exists():
                            old.unlink()
                    st.success(f"Saved {dest.name}")
                    new_saved.add(fid)
                if new_saved:
                    st.session_state["_fasta_saved_ids"] = already_saved | new_saved
                    st.rerun()

        with add_cross_col2:
            st.markdown("**Fetch from NCBI Virus (using gget)**")
            _render_fetch_ui("cross")

    st.divider()

    # --- Host sequences ---
    st.subheader("Host sequences")
    st.caption("Host genome to check for non-specific amplification (e.g., human genome)")

    if offtarget_options:
        host_labels = []
        for stem in offtarget_options:
            info = _fasta_info(stem)
            host_labels.append(f"{stem}  ({info['seqs']} seqs, {_format_size(info['size'])})")

        # Restore multiselect value from stored host when returning to this page
        stored_host = st.session_state.get("host", [])
        if stored_host and "host_select" not in st.session_state:
            restored = [l for l in host_labels if l.split("  (")[0] in stored_host]
            if restored:
                st.session_state.host_select = restored

        selected_host = st.multiselect(
            "Select host sequences",
            options=host_labels,
            key="host_select",
            placeholder="Choose host genome...",
        )
        st.session_state.host = [label.split("  (")[0] for label in selected_host]
    else:
        st.info("No FASTA files available.")
        st.session_state.host = []

    # Upload host sequences
    with st.expander("Upload host sequences"):
        uploaded_host = st.file_uploader(
            "Upload FASTA",
            type=["fa", "fasta", "fna"],
            accept_multiple_files=True,
            key="host_fasta_upload",
            label_visibility="collapsed",
        )
        if uploaded_host:
            already_saved = st.session_state.get("_fasta_saved_ids", set())
            new_saved = set()
            for f in uploaded_host:
                fid = (f.name, f.size)
                if fid in already_saved:
                    continue
                p = Path(f.name)
                if p.suffix.lower() not in (".fa", ".fasta", ".fna"):
                    st.error(f"Unsupported extension: {p.suffix}")
                    continue
                dest = _safe_path_under(FASTA_DIR, p.stem + ".fa")
                if not dest:
                    st.error(f"Invalid filename: {f.name}")
                    continue
                _save_fasta_bytes(dest, f.getvalue())
                for ext in (".fasta", ".fna"):
                    old = FASTA_DIR / (p.stem + ext)
                    if old.exists():
                        old.unlink()
                st.success(f"Saved {dest.name}")
                new_saved.add(fid)
            if new_saved:
                st.session_state["_fasta_saved_ids"] = already_saved | new_saved
                st.rerun()

    st.divider()

    # --- Back / Continue ---
    col_back, col_cont = st.columns(2)
    with col_back:
        if st.button("← Back", use_container_width=True):
            _navigate("select_target")
            st.rerun()
    with col_cont:
        if st.button("Continue →", type="primary", use_container_width=True):
            if workflow == "evaluate":
                _navigate("eval_primer")
            else:
                _navigate(f"{workflow}_config")
            st.rerun()

    st.caption("You can skip this step — cross-reactivity and host checks are optional.")


def _config_design():
    """Design-specific configuration."""
    st.header("Design — Configuration")
    st.markdown(
        "Configure design mode and parameters for primer generation."
    )

    # Design mode
    st.subheader("Design mode")
    col_single, col_multi = st.columns(2)
    with col_single:
        singleplex = st.button(
            "Singleplex",
            use_container_width=True,
            type="primary" if st.session_state.get("design_mode", "Singleplex") == "Singleplex" else "secondary",
            key="btn_singleplex",
        )
        if singleplex:
            st.session_state.design_mode = "Singleplex"
            st.rerun()
        st.caption("Design primers for a single target pathogen.")
    with col_multi:
        st.button(
            "Multiplex (coming soon)",
            use_container_width=True,
            disabled=True,
            key="btn_multiplex",
        )
        st.caption("Design a multiplexed panel for multiple targets simultaneously.")

    # Set internal mode for pipeline
    if st.session_state.get("design_mode", "Singleplex") == "Singleplex":
        st.session_state.mode = "Singleplex"

    st.divider()

    # Probe design — restore saved value if widget key was cleared by navigation
    if "probe_enabled" not in st.session_state and "_probe_enabled_saved" in st.session_state:
        st.session_state["probe_enabled"] = st.session_state["_probe_enabled_saved"]
    st.checkbox("Enable probe design", key="probe_enabled")

    # Parameters
    _tab_parameters()

    st.divider()

    col_back, col_cont = st.columns(2)
    with col_back:
        if st.button("← Back", use_container_width=True, key="design_back"):
            _navigate("select_offtarget")
            st.rerun()
    with col_cont:
        if st.button("Continue →", type="primary", use_container_width=True, key="design_cont"):
            # Persist widget value before navigating — Streamlit clears
            # widget-bound keys when the widget is no longer rendered.
            st.session_state["_probe_enabled_saved"] = st.session_state.get("probe_enabled", False)
            st.session_state._run_page_fresh = True
            _navigate("run")
            st.rerun()


def _page_eval_primer():
    """Evaluate step 3: Primer set input."""
    _render_workflow_progress("eval_primer")
    st.header("Evaluate — Primer Set")
    st.markdown(
        "Add primer sets by entering sequences or uploading a FASTA file."
    )

    st.divider()

    # Set internal mode
    st.session_state.mode = "Evaluate"

    eval_method = st.radio(
        "Input method",
        ["Paste sequences", "Upload primer FASTA"],
        key="eval_method",
        horizontal=True,
        label_visibility="collapsed",
    )

    if eval_method == "Paste sequences":
        st.text_input("Forward primer sequence (5'→3')", key="eval_for",
                       placeholder="e.g., ATGCGATCGATCGATCG")
        st.text_input("Reverse primer sequence (5'→3')", key="eval_rev",
                       placeholder="e.g., TAGCTAGCTAGCTAGCT")
        st.text_input("Probe sequence (5'→3', optional)", key="eval_pro",
                       placeholder="e.g., AACCGGTTAACCGGTTAACC")
    else:
        pset_upload = st.file_uploader(
            "Upload primer set FASTA",
            type=["fa", "fasta"],
            key="eval_pset_upload",
        )
        if pset_upload:
            pset_dir = PROJECT_ROOT / "evaluate"
            pset_dir.mkdir(parents=True, exist_ok=True)
            pset_path = pset_dir / pset_upload.name
            _save_fasta_bytes(pset_path, pset_upload.getvalue())
            st.session_state["eval_pset_path"] = str(pset_path)
            st.success(f"Saved {pset_upload.name}")

    st.divider()

    col_back, col_cont = st.columns(2)
    with col_back:
        if st.button("← Back", use_container_width=True, key="eval_primer_back"):
            _navigate("select_offtarget")
            st.rerun()
    with col_cont:
        if st.button("Continue →", type="primary", use_container_width=True, key="eval_primer_cont"):
            # Persist widget values before navigating away (widgets clear when
            # not rendered, so copy to non-widget keys)
            for k in ("eval_for", "eval_rev", "eval_pro"):
                st.session_state[f"_{k}_saved"] = st.session_state.get(k, "")
            _navigate("evaluate_config")
            st.rerun()


def _config_evaluate():
    """Evaluate-specific configuration (parameters only)."""
    st.header("Evaluate — Configuration")
    st.markdown(
        "Adjust amplicon length parameters for the evaluation pipeline."
    )

    # Summary of selections from previous steps
    targets = st.session_state.get("targets", [])
    cross = st.session_state.get("cross", [])
    host = st.session_state.get("host", [])
    eval_fwd = st.session_state.get("_eval_for_saved", "")
    eval_rev_seq = st.session_state.get("_eval_rev_saved", "")
    pset_path = st.session_state.get("eval_pset_path", "")

    st.markdown(f"**Target:** {', '.join(targets) if targets else 'none'}")
    if cross:
        st.markdown(f"**Cross-reactivity:** {', '.join(cross)}")
    if host:
        st.markdown(f"**Host:** {', '.join(host)}")
    if pset_path:
        st.markdown(f"**Primer set:** {Path(pset_path).name}")
    elif eval_fwd and eval_rev_seq:
        st.markdown(f"**Forward:** {eval_fwd}")
        st.markdown(f"**Reverse:** {eval_rev_seq}")

    st.divider()

    # Set internal mode
    st.session_state.mode = "Evaluate"

    # Parameters
    _tab_parameters(mode="evaluate")

    st.divider()

    col_back, col_cont = st.columns(2)
    with col_back:
        if st.button("← Back", use_container_width=True, key="eval_back"):
            _navigate("eval_primer")
            st.rerun()
    with col_cont:
        if st.button("Continue →", type="primary", use_container_width=True, key="eval_cont"):
            st.session_state._run_page_fresh = True
            _navigate("run")
            st.rerun()


def _page_monitor_target():
    """Monitor step 1: Select pathogens to monitor via spreadsheet or manual fetch."""
    _render_workflow_progress("monitor_target")

    st.header("Monitor — Select Pathogens")

    st.session_state.mode = "Monitor"

    # --- Resume existing run ---
    existing_runs = []
    if MONITOR_DIR.exists():
        for d in sorted(MONITOR_DIR.iterdir(), reverse=True):
            if d.is_dir() and not d.name.startswith("."):
                # Find date subdirectories
                date_dirs = sorted(
                    [dd for dd in d.iterdir() if dd.is_dir() and dd.name.isdigit()],
                    reverse=True,
                )
                if date_dirs:
                    existing_runs.append((d.name, date_dirs))

    if existing_runs:
        st.subheader("Resume existing run")
        st.caption("Skip fetching and re-evaluate with sequences from a previous run.")
        run_options = [r[0] for r in existing_runs]
        selected_run = st.selectbox(
            "Run ID", options=run_options, index=None,
            placeholder="Select a previous run...",
            key="monitor_resume_run",
        )

        if selected_run:
            # Find the run and show its details
            for run_id, date_dirs in existing_runs:
                if run_id == selected_run:
                    # Show available dates
                    date_labels = [dd.name for dd in date_dirs]
                    st.caption(f"Available dates: {', '.join(date_labels)}")

                    # Load mastersheet from latest date dir
                    latest = date_dirs[0]
                    mastersheet = latest / "mastersheet.csv"
                    if mastersheet.exists():
                        csv_text = mastersheet.read_text()
                        try:
                            _, data_rows = _load_spreadsheet(csv_text)
                            import pandas as pd
                            seen = {}
                            for row in data_rows:
                                t = _make_target_name(row)
                                if t not in seen:
                                    seen[t] = {
                                        "Target": t,
                                        "TaxID": str(row.get("TaxID", "")).strip(),
                                        "Host": str(row.get("host", "")).strip(),
                                        "Geographic location": str(row.get("geographic_location", "")).strip(),
                                    }
                            st.dataframe(pd.DataFrame(list(seen.values())),
                                         use_container_width=True, hide_index=True)

                            # Show primer info
                            primer_rows = []
                            for row in data_rows:
                                primer_rows.append({
                                    "Target": _make_target_name(row),
                                    "Primer name": row.get("Primer name", ""),
                                    "Forward": row.get("Forward", "")[:25],
                                    "Reverse": row.get("Reverse", "")[:25],
                                    "Probe": "yes" if not _is_empty(row.get("Probe")) else "",
                                })
                            with st.expander("Primer sets"):
                                st.dataframe(pd.DataFrame(primer_rows),
                                             use_container_width=True, hide_index=True)

                            # Show sequence counts
                            acc_files = list(latest.glob("*_accessions.txt"))
                            if acc_files:
                                counts = []
                                for af in acc_files:
                                    target = af.stem.replace("_accessions", "")
                                    n = len([l for l in af.read_text().strip().split("\n") if l.strip()])
                                    counts.append(f"{target}: {n} sequences")
                                st.caption("Sequences: " + ", ".join(counts))

                        except Exception:
                            st.caption("Could not parse mastersheet.")
                    break

            if st.button("Continue with skip fetch →", type="primary",
                         use_container_width=True, key="monitor_resume_cont"):
                st.session_state["run_id"] = selected_run
                st.session_state["monitor_skip_fetch"] = True
                # Load mastersheet data as spreadsheet data
                latest_dir = None
                for run_id, date_dirs in existing_runs:
                    if run_id == selected_run:
                        latest_dir = date_dirs[0]
                        break
                if latest_dir:
                    ms = latest_dir / "mastersheet.csv"
                    if ms.exists():
                        csv_text = ms.read_text()
                        try:
                            headers, data_rows = _load_spreadsheet(csv_text)
                            st.session_state["monitor_spreadsheet_data"] = {
                                "csv_text": csv_text,
                                "headers": headers,
                                "data_rows": data_rows,
                                "spreadsheet_id": selected_run,
                            }
                        except Exception:
                            pass
                _navigate("monitor_primer")
                st.rerun()

        st.divider()

    # --- Option 1: Load from Google Sheets ---
    st.subheader("Load from Google Sheets ([template](https://docs.google.com/spreadsheets/d/127GY5lBqUKuri4MgM4XTlKLKto45_Yn9NvYJhF2WIvo/copy))")
    st.info(
        '**After copying, set sharing to "Anyone with the link" → Viewer so the app can read it.**'
    )
    spreadsheet_url = st.text_input(
        "Spreadsheet URL",
        key="monitor_spreadsheet_url",
        placeholder="https://docs.google.com/spreadsheets/d/...",
    )

    if spreadsheet_url:
        if st.button("Load spreadsheet", key="monitor_load_sheet"):
            try:
                sid = _extract_spreadsheet_id(spreadsheet_url)
                csv_text = _download_spreadsheet_csv(sid)
                headers, data_rows = _load_spreadsheet(csv_text)
                st.session_state["monitor_spreadsheet_data"] = {
                    "csv_text": csv_text,
                    "headers": headers,
                    "data_rows": data_rows,
                    "spreadsheet_id": sid,
                }
                st.success(f"Loaded {len(data_rows)} row(s) from spreadsheet.")
            except Exception as e:
                st.error(f"Failed to load spreadsheet: {e}")

    # Show loaded data
    sheet_data = st.session_state.get("monitor_spreadsheet_data")
    if sheet_data:
        data_rows = sheet_data["data_rows"]

        import pandas as pd
        from datetime import date
        today_str = date.today().strftime("%Y-%m-%d")
        # Group by target — show target-level info, not per-primer
        seen_targets = {}
        for row in data_rows:
            target_name = _make_target_name(row)
            if target_name not in seen_targets:
                seen_targets[target_name] = {
                    "Target": target_name,
                    "TaxID": str(row.get("TaxID", "")).strip(),
                    "Min release date": str(row.get("min_release_date", "")).strip(),
                    "Max release date": today_str,
                    "Geographic location": str(row.get("geographic_location", "")).strip(),
                    "Host": str(row.get("host", "")).strip(),
                }
        st.dataframe(pd.DataFrame(list(seen_targets.values())), use_container_width=True, hide_index=True)

        target_names = list(seen_targets.keys())
        st.markdown(f"**Targets to monitor:** {', '.join(target_names)}")

    st.divider()

    # --- Option 2: Manual setup ---
    st.subheader("Manual setup")
    st.caption("Set up a single target manually by fetching sequences from NCBI Virus.")
    with st.expander("Configure manual fetch"):
        _render_fetch_ui("monitor_fetch", monitor=True)

    st.divider()

    # --- Continue ---
    has_sheet = bool(st.session_state.get("monitor_spreadsheet_data"))
    if st.button(
        "Continue →",
        disabled=not has_sheet,
        type="primary",
        use_container_width=True,
        key="monitor_target_cont",
    ):
        _navigate("monitor_primer")
        st.rerun()

    if not has_sheet:
        st.caption("Load a spreadsheet to continue.")


def _page_monitor_primer():
    """Monitor step 2: Primer set configuration."""
    _render_workflow_progress("monitor_primer")

    st.header("Monitor — Primer Sets")
    st.markdown(
        "Review primer sets from the spreadsheet, or add additional ones by entering sequences or uploading a FASTA file."
    )

    st.divider()

    sheet_data = st.session_state.get("monitor_spreadsheet_data")

    # --- From spreadsheet ---
    if sheet_data:
        st.subheader("Primers from spreadsheet")
        import pandas as pd
        display_rows = []
        for row in sheet_data["data_rows"]:
            fwd = row.get("Forward", "")
            rev = row.get("Reverse", "")
            pro = row.get("Probe", "")
            display_rows.append({
                "Target": _make_target_name(row),
                "Primer name": row.get("Primer name", ""),
                "Forward": fwd[:25] + "..." if len(fwd) > 25 else fwd,
                "Reverse": rev[:25] + "..." if len(rev) > 25 else rev,
                "Probe": pro[:25] + "..." if len(pro) > 25 else pro if not _is_empty(pro) else "",
            })
        st.dataframe(pd.DataFrame(display_rows), use_container_width=True, hide_index=True)

    st.divider()

    # --- Manual input ---
    st.subheader("Add primer sets manually")

    # Initialize manual primer sets list
    if "monitor_manual_primers" not in st.session_state:
        st.session_state["monitor_manual_primers"] = []

    with st.expander("Enter primer sequences"):
        name = st.text_input("Primer set name", key="monitor_primer_name",
                             placeholder="e.g. My_primer_1")
        fwd = st.text_input("Forward primer (5'→3')", key="monitor_primer_fwd",
                            placeholder="ATCGATCG...")
        rev = st.text_input("Reverse primer (5'→3')", key="monitor_primer_rev",
                            placeholder="ATCGATCG...")
        pro = st.text_input("Probe (5'→3', optional)", key="monitor_primer_pro",
                            placeholder="ATCGATCG...")

        if st.button("Add primer set", key="monitor_add_primer"):
            if name and fwd and rev:
                st.session_state["monitor_manual_primers"].append({
                    "name": name, "Forward": fwd, "Reverse": rev, "Probe": pro,
                })
                st.session_state["monitor_primer_name"] = ""
                st.session_state["monitor_primer_fwd"] = ""
                st.session_state["monitor_primer_rev"] = ""
                st.session_state["monitor_primer_pro"] = ""
                st.rerun()
            else:
                st.warning("Name, Forward, and Reverse are required.")

    # Show manually added primers
    manual = st.session_state.get("monitor_manual_primers", [])
    if manual:
        import pandas as pd
        st.markdown(f"**{len(manual)} manual primer set(s) added:**")
        manual_display = [{
            "Primer name": p["name"],
            "Forward": p["Forward"],
            "Reverse": p["Reverse"],
            "Probe": p.get("Probe", ""),
        } for p in manual]
        st.dataframe(pd.DataFrame(manual_display), use_container_width=True, hide_index=True)
        if st.button("Clear all manual primers", key="monitor_clear_primers"):
            st.session_state["monitor_manual_primers"] = []
            st.rerun()

    # FASTA upload
    with st.expander("Upload primer set FASTA"):
        uploaded = st.file_uploader(
            "Upload a FASTA file with primer sequences",
            type=["fa", "fasta", "fna"],
            key="monitor_pset_upload",
        )
        if uploaded:
            _id = (uploaded.name, uploaded.size)
            saved = st.session_state.setdefault("_fasta_saved_ids", set())
            if _id not in saved:
                pset_path = PROJECT_ROOT / "evaluate" / uploaded.name
                pset_path.parent.mkdir(parents=True, exist_ok=True)
                _save_fasta_bytes(pset_path, uploaded.getvalue())
                st.session_state["monitor_pset_path"] = str(pset_path)
                saved.add(_id)
            st.success(f"Uploaded: {uploaded.name}")

    st.divider()

    # --- Back / Continue ---
    col_back, col_cont = st.columns(2)
    with col_back:
        if st.button("← Back", use_container_width=True, key="monitor_primer_back"):
            _navigate("monitor_target")
            st.rerun()
    with col_cont:
        if st.button("Continue →", type="primary", use_container_width=True, key="monitor_primer_cont"):
            _navigate("monitor_config")
            st.rerun()


def _config_monitor():
    """Monitor step 3: Amplicon length configuration."""
    st.header("Monitor — Configuration")
    st.markdown(
        "Adjust amplicon length parameters for the monitoring pipeline."
    )

    st.divider()

    # --- Amplicon length ---
    st.subheader("Amplicon length")
    p = st.session_state.setdefault("params", dict(DEFAULT_PARAMS))
    c1, c2 = st.columns(2)
    p["AMPLEN_MIN"] = c1.number_input(
        "AMPLEN_MIN", value=int(p["AMPLEN_MIN"]),
        min_value=1, step=10, key="monitor_amplen_min",
    )
    p["AMPLEN_MAX"] = c2.number_input(
        "AMPLEN_MAX", value=int(p["AMPLEN_MAX"]),
        min_value=1, step=10, key="monitor_amplen_max",
    )
    if p["AMPLEN_MIN"] >= p["AMPLEN_MAX"]:
        st.warning("AMPLEN_MIN should be less than AMPLEN_MAX")

    st.divider()

    # --- Back / Continue ---
    col_back, col_cont = st.columns(2)
    with col_back:
        if st.button("← Back", use_container_width=True, key="monitor_back"):
            _navigate("monitor_primer")
            st.rerun()
    with col_cont:
        if st.button("Continue →", type="primary", use_container_width=True, key="monitor_cont"):
            _navigate("monitor_report")
            st.rerun()


def _page_design():
    """Design workflow — config page."""
    _render_workflow_progress("config")
    _config_design()


def _page_evaluate():
    """Evaluate workflow — config page."""
    _render_workflow_progress("config")
    _config_evaluate()


def _page_monitor():
    """Monitor workflow — config page."""
    _render_workflow_progress("config")
    _config_monitor()


def _monitor_preflight_checks() -> list[str]:
    """Preflight checks specific to monitor workflow."""
    errors = []
    if not st.session_state.get("monitor_spreadsheet_data"):
        errors.append("Load a spreadsheet first.")
    if not shutil.which("gget"):
        errors.append("Tool 'gget' not found in PATH. Run: pip install gget")
    for tool in ("snakemake", "bowtie2", "sam2pairwise"):
        if not _check_tool(tool):
            errors.append(f"Tool '{tool}' not found in PATH.")
    return errors


def _run_monitor():
    """Execute the full monitor pipeline: fetch → evaluate → email."""
    import io as _io

    sheet_data = st.session_state["monitor_spreadsheet_data"]
    csv_text = sheet_data["csv_text"]
    data_rows = sheet_data["data_rows"]
    spreadsheet_id = sheet_data["spreadsheet_id"]

    selected_qids = st.session_state.get("monitor_query_ids", [])
    if selected_qids:
        data_rows = [r for r in data_rows if str(r.get("query_id", "")).strip() in selected_qids]

    # Group rows by target
    target_groups: dict[str, list[dict]] = {}
    for row in data_rows:
        name = _make_target_name(row)
        target_groups.setdefault(name, []).append(row)

    target_names = list(target_groups.keys())

    # Setup directories
    skip_fetch = st.session_state.get("monitor_skip_fetch", False)
    existing_run_id = st.session_state.get("run_id", "")

    if skip_fetch and existing_run_id and "/" in existing_run_id:
        # Reuse existing run directory
        runid, date_str = existing_run_id.split("/", 1)
        work_dir = MONITOR_DIR / runid
        date_dir = work_dir / date_str
    else:
        date_str = datetime.now().strftime("%Y%m%d")
        runid = st.session_state.get("monitor_run_id", "").strip()
        if not runid:
            runid = spreadsheet_id[:8]
        work_dir = MONITOR_DIR / runid
        date_dir = work_dir / date_str

    date_dir.mkdir(parents=True, exist_ok=True)

    # Save spreadsheet snapshot
    (date_dir / "mastersheet.csv").write_text(csv_text, encoding="utf-8")

    st.session_state.run_id = f"{runid}/{date_str}"

    # Read params and email config
    _init_params()
    params_content = build_params_txt(st.session_state.params)
    params_path = date_dir / "params.txt"
    params_path.write_text(params_content)

    # Read email config from project params.txt (password deferred to send time)
    params_source = PROJECT_ROOT / "params.txt"
    email_sender = ""
    default_recipients = ""
    if params_source.exists():
        p = parse_params(params_source)
        email_sender = str(p.get("EMAIL_SENDER", "")).strip()
        default_recipients = str(p.get("EMAIL_RECIPIENTS", "")).strip()

    recipients_str = st.session_state.get("monitor_email_recipients", "").strip()
    if not recipients_str:
        recipients_str = default_recipients
    email_recipients = [
        e.strip() for e in recipients_str.split(",")
        if e.strip()
    ]

    # Append non-sensitive email config to run params.txt
    with open(params_path, "a") as f:
        f.write(f"\n## Email configuration\n")
        f.write(f"EMAIL_SENDER = {email_sender}\n")
        f.write(f"EMAIL_RECIPIENTS = {recipients_str}\n")
    # Note: EMAIL_PASSWORD is read from project params.txt at send time, never copied

    cores = st.session_state.get("cores", 1)
    progress_area = st.empty()
    status_area = st.empty()
    start_time = time.time()

    def _elapsed():
        s = int(time.time() - start_time)
        m, s = divmod(s, 60)
        return f"{m}m {s}s"

    def _update(lines):
        progress_area.markdown("<br>".join(lines), unsafe_allow_html=True)
        status_area.caption(f"Elapsed: {_elapsed()}")

    progress_lines = []
    fetch_results: dict[str, dict] = {}

    if skip_fetch:
        # ===================== SKIP FETCH — use existing sequences =====================
        progress_lines.append("⏳ &nbsp; **Scanning existing sequences (skip fetch)...**")
        _update(progress_lines)

        for target_name in target_names:
            new_fasta = date_dir / f"{target_name}_new.fa"
            acc_file = date_dir / f"{target_name}_accessions.txt"
            meta_file = date_dir / f"{target_name}_metadata.csv"

            if new_fasta.exists():
                new_acc = _extract_accessions(new_fasta)
                total_acc = set()
                if acc_file.exists():
                    total_acc = set(l.strip() for l in acc_file.read_text().strip().split("\n") if l.strip())
                fetch_results[target_name] = {
                    "status": "success",
                    "new_fasta_path": new_fasta,
                    "metadata_path": meta_file if meta_file.exists() else None,
                    "new_accessions": new_acc,
                    "total_count": len(total_acc) if total_acc else len(new_acc),
                    "new_count": len(new_acc),
                    "is_first_fetch": False,
                }
            else:
                fetch_results[target_name] = {"status": "no_output", "new_count": 0}

        total_new = sum(r.get("new_count", 0) for r in fetch_results.values() if r.get("status") == "success")
        progress_lines[-1] = f"✅ &nbsp; Found {total_new} existing sequence(s) — fetch skipped"
        _update(progress_lines)
    else:
        # ===================== STEP 1: FETCH =====================
        progress_lines.append("⏳ &nbsp; **Fetching sequences...**")
        _update(progress_lines)

        _OVERLAP_DAYS = 7

        for i, (target_name, rows) in enumerate(target_groups.items()):
            row = {k: v for k, v in rows[0].items()
                   if k not in {"max_collection_date", "max_release_date"}}

            # Target status
            tparts = []
            for j, tn in enumerate(target_names):
                if j < i:
                    tparts.append(f"{tn} ✓")
                elif j == i:
                    tparts.append(f"**{tn}**")
                else:
                    tparts.append(f"<span style='color:#ccc'>{tn}</span>")
            progress_lines[-1] = f"⏳ &nbsp; **Fetching sequences** — {', '.join(tparts)}"
            _update(progress_lines)

            dated_fasta = date_dir / f"{target_name}.fa"
            dated_meta = date_dir / f"{target_name}_metadata.csv"

            # Find previous accessions for diff
            prev_acc_files = sorted(
                [d / f"{target_name}_accessions.txt"
                 for d in work_dir.iterdir()
                 if d.is_dir() and d.name != date_str
                 and (d / f"{target_name}_accessions.txt").exists()],
                key=lambda p: p.parent.name,
                reverse=True,
            )
            is_first = len(prev_acc_files) == 0

            # Narrow date window for subsequent fetches
            if not is_first:
                from datetime import timedelta
                prev_date_str = prev_acc_files[0].parent.name
                try:
                    prev_date = datetime.strptime(prev_date_str, "%Y%m%d")
                    cutoff = (prev_date - timedelta(days=_OVERLAP_DAYS)).strftime("%Y-%m-%d")
                    for date_col in ("min_release_date", "min_collection_date"):
                        orig = str(row.get(date_col, "")).strip()
                        if not orig or orig < cutoff:
                            row[date_col] = cutoff
                except ValueError:
                    pass

            # Run gget
            gget_tmp = date_dir / f".gget_tmp_{target_name}"
            gget_tmp.mkdir(parents=True, exist_ok=True)
            cmd = _build_gget_command(row, str(gget_tmp))

            if not cmd:
                fetch_results[target_name] = {"status": "error"}
                shutil.rmtree(gget_tmp, ignore_errors=True)
                continue

            try:
                subprocess.run(cmd, check=True, capture_output=True)
            except subprocess.CalledProcessError:
                fetch_results[target_name] = {"status": "error"}
                shutil.rmtree(gget_tmp, ignore_errors=True)
                continue

            # Move FASTA output
            fasta_files = list(gget_tmp.glob("*.fa")) + list(gget_tmp.glob("*.fasta"))
            meta_files = list(gget_tmp.glob("*_metadata.csv"))

            if not fasta_files:
                fetch_results[target_name] = {"status": "no_output"}
                shutil.rmtree(gget_tmp, ignore_errors=True)
                continue

            shutil.move(str(fasta_files[0]), str(dated_fasta))
            if meta_files:
                shutil.move(str(meta_files[0]), str(dated_meta))
            shutil.rmtree(gget_tmp, ignore_errors=True)

            _deduplicate_fasta(dated_fasta)

            # Diff accessions
            current_acc = _extract_accessions(dated_fasta)
            total_count = len(current_acc)

            if not is_first:
                prev_acc = set(prev_acc_files[0].read_text().strip().split("\n"))
                new_acc = current_acc - prev_acc
            else:
                new_acc = current_acc

            # Write new-only FASTA
            new_fasta = None
            if new_acc:
                new_fasta = date_dir / f"{target_name}_new.fa"
                _filter_fasta_by_accessions(dated_fasta, new_acc, new_fasta)

            # Save accessions, delete full FASTA
            acc_file = date_dir / f"{target_name}_accessions.txt"
            acc_file.write_text("\n".join(sorted(current_acc)) + "\n")
            dated_fasta.unlink(missing_ok=True)

            fetch_results[target_name] = {
                "status": "success",
                "new_fasta_path": new_fasta,
                "metadata_path": dated_meta if dated_meta.exists() else None,
                "new_accessions": new_acc,
                "total_count": total_count,
                "new_count": len(new_acc),
                "is_first_fetch": is_first,
            }

        # Summarize fetch
        total_new = sum(r.get("new_count", 0) for r in fetch_results.values() if r.get("status") == "success")
        success_targets = [t for t, r in fetch_results.items() if r.get("status") == "success"]
        progress_lines[-1] = f"✅ &nbsp; Fetch complete — {total_new} new sequence(s) across {len(success_targets)} target(s)"
        _update(progress_lines)

    # ===================== STEP 2: EVALUATE =====================
    targets_with_new = [
        t for t in target_names
        if fetch_results.get(t, {}).get("new_count", 0) > 0
    ]

    all_xlsx: list[Path] = []
    email_body_parts = [f"ADAPT Monitor Report — {date_str}\n{'=' * 50}\n"]

    if targets_with_new:
        progress_lines.append("⏳ &nbsp; **Evaluating primers...**")
        _update(progress_lines)

        for i, target_name in enumerate(targets_with_new):
            result = fetch_results[target_name]
            rows = target_groups[target_name]

            # Update progress
            tparts = []
            for j, tn in enumerate(targets_with_new):
                if j < i:
                    tparts.append(f"{tn} ✓")
                elif j == i:
                    tparts.append(f"**{tn}**")
                else:
                    tparts.append(f"<span style='color:#ccc'>{tn}</span>")
            progress_lines[-1] = f"⏳ &nbsp; **Evaluating primers** — {', '.join(tparts)}"
            _update(progress_lines)

            email_body_parts.append(f"\n== {target_name} ==")
            email_body_parts.append(
                f"{result['new_count']} new sequence(s) ({result['total_count']} total)\n"
            )
            seq_table = _get_new_seq_table(result["new_accessions"], result.get("metadata_path"))
            email_body_parts.append(seq_table)

            target_fasta = result["new_fasta_path"]
            if not target_fasta:
                continue

            # Build pset.fa from spreadsheet rows
            pset_fa = date_dir / f"{target_name}_pset.fa"
            if not pset_fa.exists():
                primer_rows = [
                    r for r in rows
                    if not _is_empty(r.get("Forward")) and not _is_empty(r.get("Reverse"))
                ]
                if not primer_rows:
                    email_body_parts.append("No primer sequences — skipping evaluate.\n")
                    continue
                _build_pset_fa(primer_rows, pset_fa)

            # Setup eval directory
            eval_dir = date_dir / target_name
            if eval_dir.exists():
                shutil.rmtree(eval_dir)
            eval_dir.mkdir(parents=True, exist_ok=True)

            shutil.copy2(params_path, eval_dir / "params.txt")
            shutil.copy2(pset_fa, eval_dir / pset_fa.name)

            target_seq_dir = eval_dir / "target_seqs" / "original"
            target_seq_dir.mkdir(parents=True, exist_ok=True)
            shutil.copy2(target_fasta, target_seq_dir / f"{target_name}.fa")

            # Ensure TARGETS is set in params
            eval_params = eval_dir / "params.txt"
            ptxt = eval_params.read_text()
            if "TARGETS" not in ptxt:
                with open(eval_params, "a") as f:
                    f.write(f"\nTARGETS = {target_name}\n")
            else:
                ptxt = re.sub(
                    r"^TARGETS\s*=.*$",
                    f"TARGETS = {target_name}",
                    ptxt,
                    flags=re.MULTILINE,
                )
                eval_params.write_text(ptxt)

            # Generate Snakefile
            _generate_snakefile(eval_dir, [target_name], [], [], [])

            # Run snakemake evaluate
            snakemake_bin = _find_tool("snakemake") or "snakemake"
            eval_cmd = [
                snakemake_bin, "-s", "Snakefile",
                "--cores", str(cores),
                "--config", "evaluate=1", f"pset={pset_fa.name}",
            ]

            env = os.environ.copy()
            env["PYTHONUNBUFFERED"] = "1"
            snakemake_path = _find_tool("snakemake")
            env_bin = str(Path(sys.executable).parent)
            extra_dirs = [env_bin]
            if snakemake_path:
                extra_dirs.append(str(Path(snakemake_path).parent))
            for d in extra_dirs:
                if d and d not in env.get("PATH", ""):
                    env["PATH"] = d + os.pathsep + env.get("PATH", "")

            proc = subprocess.run(
                eval_cmd,
                cwd=str(eval_dir),
                env=env,
                capture_output=True,
                text=True,
            )

            if proc.returncode != 0:
                email_body_parts.append("Evaluate failed.\n")
                continue

            # Collect Excel files
            xlsx_files = sorted(eval_dir.rglob("*.xlsx"))
            for xlsx in xlsx_files:
                dest = date_dir / f"{target_name}_{xlsx.name}"
                shutil.move(str(xlsx), str(dest))
                all_xlsx.append(dest)

            email_body_parts.append("Evaluation results:")
            for xlsx in all_xlsx[-len(xlsx_files):]:
                email_body_parts.append(f"\n  --- {xlsx.stem} ---")
                summary_text = _read_excel_summary(xlsx)
                email_body_parts.append(summary_text)

        progress_lines[-1] = f"✅ &nbsp; Evaluation complete — {len(all_xlsx)} report(s)"
        _update(progress_lines)
    else:
        progress_lines.append("✅ &nbsp; No new sequences — evaluation skipped")
        _update(progress_lines)

    # ===================== STEP 3: EMAIL =====================
    email_body = "\n".join(email_body_parts)

    # Read password at send time only (avoid storing in long-lived variable)
    _email_password = ""
    if params_source.exists():
        _email_password = str(parse_params(params_source).get("EMAIL_PASSWORD", "")).strip()

    if email_sender and _email_password and email_recipients and all_xlsx:
        progress_lines.append("⏳ &nbsp; **Sending email report...**")
        _update(progress_lines)

        subject = f"[ADAPT Monitor] {date_str} — {len(target_names)} target(s)"
        ok = _send_email(
            sender=email_sender,
            password=_email_password,
            recipients=email_recipients,
            subject=subject,
            body=email_body,
            attachments=all_xlsx,
        )
        _email_password = ""  # clear after use
        if ok:
            progress_lines[-1] = f"✅ &nbsp; Email sent to {', '.join(email_recipients)}"
        else:
            progress_lines[-1] = "⚠️ &nbsp; Email sending failed"
        _update(progress_lines)
    elif not email_sender or not _email_password:
        progress_lines.append("⚪ &nbsp; Email skipped (not configured)")
        _update(progress_lines)
    elif not all_xlsx:
        progress_lines.append("⚪ &nbsp; Email skipped (no reports to send)")
        _update(progress_lines)

    # Done
    elapsed = int(time.time() - start_time)
    mins, secs = divmod(elapsed, 60)
    status_area.caption(f"Completed in {mins}m {secs}s")

    st.session_state.pipeline_running = False
    st.session_state.pipeline_return_code = 0
    st.session_state.monitor_results_dir = str(date_dir)

    if all_xlsx:
        st.success(f"Monitor complete — {len(all_xlsx)} report(s) generated.")
    elif total_new == 0:
        st.info("No new sequences found. Nothing to evaluate.")
    else:
        st.warning("Fetch succeeded but evaluation produced no reports.")


def _tab_run_monitor():
    """Run page for monitor workflow."""
    st.header("Run Monitor")

    # Run ID for identifying this run later
    sheet_data = st.session_state.get("monitor_spreadsheet_data")
    default_run_id = ""
    if sheet_data:
        default_run_id = sheet_data.get("spreadsheet_id", "")[:8]
    if "monitor_run_id" not in st.session_state:
        st.session_state["monitor_run_id"] = default_run_id
    st.text_input(
        "Run ID",
        key="monitor_run_id",
        help="A short identifier for this run. Used to organize output directories.",
    )

    c1, _ = st.columns(2)
    max_cpu = os.cpu_count() or 1
    c1.slider("CPU cores", min_value=1, max_value=max_cpu, value=min(4, max_cpu), key="cores")

    # Preflight
    errors = _monitor_preflight_checks()
    if errors:
        for e in errors:
            st.error(e)

    # Clear previous result when entering Run page for first time
    if st.session_state.get("_run_page_fresh", True):
        st.session_state.pipeline_return_code = None
        st.session_state._run_page_fresh = False

    running = st.session_state.get("pipeline_running", False)

    run_clicked = st.button("Run Monitor", disabled=running or bool(errors), type="primary")

    if run_clicked:
        st.session_state.pipeline_running = True
        st.session_state.pipeline_should_start = True
        st.session_state.pipeline_return_code = None
        st.rerun()

    if st.session_state.get("pipeline_should_start"):
        st.session_state.pipeline_should_start = False
        _run_monitor()

    elif not running and st.session_state.get("pipeline_return_code") is not None:
        rc = st.session_state.pipeline_return_code
        if rc == 0:
            st.success("Monitor completed successfully!")
        else:
            st.error("Monitor failed.")


def _page_run():
    """Run pipeline page."""
    _render_workflow_progress("run")

    workflow = st.session_state.get("workflow", "design")

    # Back button (only when pipeline is not running)
    if not st.session_state.get("pipeline_running"):
        if workflow == "monitor":
            back_label = "← Back to Report"
            back_target = "monitor_report"
        else:
            back_label = "← Back to Configuration"
            back_target = f"{workflow}_config"
        if st.button(back_label, key="run_back"):
            _navigate(back_target)
            st.rerun()

    if workflow == "monitor":
        _tab_run_monitor()
    else:
        _tab_run()

    # Show results button after pipeline completion
    rc = st.session_state.get("pipeline_return_code")
    if rc is not None and not st.session_state.get("pipeline_running"):
        st.divider()
        if st.button("View Results →", type="primary", use_container_width=True):
            _navigate("results")
            st.rerun()


def _page_results():
    """Results page."""
    _render_workflow_progress("results")
    if st.button("← Back to Run", key="results_back"):
        _navigate("run")
        st.rerun()
    _tab_results()


def _get_active_cron_monitor() -> str | None:
    """Check if an adapt-monitor cron job is currently active. Returns the cron line or None."""
    try:
        result = subprocess.run(["crontab", "-l"], capture_output=True, text=True)
        if result.returncode != 0:
            return None
        for line in result.stdout.strip().split("\n"):
            if "adapt-monitor" in line:
                return line
    except FileNotFoundError:
        pass
    return None


def _page_monitor_report():
    """Monitor: Email & scheduling page."""
    _render_workflow_progress("monitor_report")

    st.header("Monitor — Report")

    # --- Email ---
    st.subheader("Email")

    params_path = PROJECT_ROOT / "params.txt"
    email_configured = False
    if params_path.exists():
        params = parse_params(params_path)
        email_sender = str(params.get("EMAIL_SENDER", "")).strip()
        email_password = str(params.get("EMAIL_PASSWORD", "")).strip()
        email_configured = bool(email_sender and email_password)

    if email_configured:
        st.markdown(f"Sender: `{email_sender}`")
        default_recipients = str(params.get("EMAIL_RECIPIENTS", "")).strip()
    else:
        st.warning(
            "Admin email not configured. Set `EMAIL_SENDER` and `EMAIL_PASSWORD` "
            "in `params.txt` to enable email reports."
        )
        default_recipients = ""

    if "monitor_email_recipients" not in st.session_state and default_recipients:
        st.session_state["monitor_email_recipients"] = default_recipients
    st.text_input(
        "Recipient email(s), comma-separated",
        key="monitor_email_recipients",
        disabled=not email_configured,
        placeholder="user1@example.com, user2@example.com",
    )

    st.divider()

    # --- Schedule ---
    st.subheader("Schedule")

    frequency_options = {
        "Biweekly (1st & 15th)": "biweekly",
        "Monthly (1st)": "monthly",
        "Quarterly (Jan, Apr, Jul, Oct)": "quarterly",
    }
    selected_label = st.radio(
        "Report frequency",
        options=list(frequency_options.keys()),
        index=1,
        key="monitor_frequency",
    )
    frequency = frequency_options[selected_label]

    st.divider()

    # --- Back / Run ---
    col_back, col_cont = st.columns(2)
    with col_back:
        if st.button("← Back", use_container_width=True, key="report_back"):
            _navigate("monitor_config")
            st.rerun()
    with col_cont:
        if st.button("Continue →", type="primary", use_container_width=True,
                      key="report_cont", disabled=not email_configured):
            # Install cron job
            recipients = st.session_state.get("monitor_email_recipients", "").strip()
            try:
                _install_cron(
                    work_dir=PROJECT_ROOT / "monitor",
                    params_file=PROJECT_ROOT / "params.txt",
                    schedule_email=recipients,
                    frequency=frequency,
                )
                # Save schedule details for display
                import json
                sheet_data = st.session_state.get("monitor_spreadsheet_data")
                targets = []
                if sheet_data:
                    seen = set()
                    for row in sheet_data["data_rows"]:
                        t = _make_target_name(row)
                        if t not in seen:
                            targets.append(t)
                            seen.add(t)
                primer_names = []
                if sheet_data:
                    for row in sheet_data["data_rows"]:
                        pn = row.get("Primer name", "").strip()
                        if pn and pn not in primer_names:
                            primer_names.append(pn)
                MONITOR_DIR.mkdir(parents=True, exist_ok=True)
                MONITOR_SCHEDULE_PATH.write_text(json.dumps({
                    "targets": targets,
                    "primer_sets": primer_names,
                    "frequency": selected_label,
                    "recipients": recipients,
                }, indent=2))
                st.session_state["_monitor_cron_installed"] = True
            except SystemExit:
                st.error("Failed to install cron job. Is 'adapt' in PATH?")
            st.session_state._run_page_fresh = True
            _navigate("run")
            st.rerun()


_SCROLL_JS = """
<script>
setTimeout(function() {
    const doc = window.parent.document;
    const containers = doc.querySelectorAll('[data-testid="stMain"], [data-testid="ScrollToBottomContainer"], section.main');
    containers.forEach(el => { el.scrollTop = 0; });
    let node = doc.querySelector('[data-testid="stAppViewBlockContainer"]');
    while (node) {
        node.scrollTop = 0;
        node = node.parentElement;
    }
    doc.documentElement.scrollTop = 0;
    doc.body.scrollTop = 0;
}, 100);
</script>
"""


def _scroll_to_top():
    """Inject JS to scroll to the top of the page if navigation just happened."""
    if st.session_state.pop("_need_scroll_top", False):
        st.components.v1.html(_SCROLL_JS, height=0)


def main():
    _render_sidebar()

    page = _current_page()

    if page == "home":
        _page_home()
    elif page == "virus_map":
        _page_virus_map()
    elif page == "contact":
        st.subheader("Contact")
        st.markdown(
            "**S Chan Baek** — [baekseun@broadinstitute.org](mailto:baekseun@broadinstitute.org)  \n"
            "**Kenneth B Hsu** — [khsu@broadinstitute.org](mailto:khsu@broadinstitute.org)"
        )
    elif page == "select_target":
        _page_select_target()
    elif page == "select_offtarget":
        _page_select_offtarget()
    elif page in ("design", "design_config"):
        _page_design()
    elif page == "eval_primer":
        _page_eval_primer()
    elif page in ("evaluate", "evaluate_config"):
        _page_evaluate()
    elif page == "monitor_target":
        _page_monitor_target()
    elif page == "monitor_primer":
        _page_monitor_primer()
    elif page in ("monitor", "monitor_config"):
        _page_monitor()
    elif page == "run":
        _page_run()
    elif page == "results":
        _page_results()
    elif page == "monitor_report":
        _page_monitor_report()
    else:
        _page_home()

    _scroll_to_top()


if __name__ == "__main__":
    main()
