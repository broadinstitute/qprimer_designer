"""Rescue evaluate: re-evaluate decision=0 targets with a subset reference."""

import argparse
import ast
import shlex
import subprocess
import sys
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
from Bio import SeqIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sklearn.preprocessing import MultiLabelBinarizer


def register(subparsers):
    """Register the rescue-evaluate subcommand."""
    parser = subparsers.add_parser(
        "rescue-evaluate",
        help="Re-evaluate failed targets with subset reference",
        description=(
            "After initial evaluation, re-run bowtie2 + ML on targets with "
            "decision=0 using a smaller reference. This avoids bowtie2 -k "
            "saturation on large references. On-target only."
        ),
    )
    parser.add_argument("--eval", dest="eval_path", required=True,
                        help="Path to .eval file (derives .full, .cl, .re)")
    parser.add_argument("--report-dir", dest="report_dir", required=True,
                        help="Directory containing per-primer-set .xlsx reports")
    parser.add_argument("--fasta", required=True, help="Primer FASTA file")
    parser.add_argument("--ref", required=True, help="Reference FASTA file")
    parser.add_argument("--features", required=True, help="Primer features file")
    parser.add_argument("--params", required=True, help="Parameters file")
    parser.add_argument("--reftype", required=True, choices=["on", "off"])
    parser.add_argument("--threads", type=int, default=2)
    parser.set_defaults(func=run)


def _find_failed_targets(report_dir):
    """Find seq_ids with decision=0 across all xlsx reports in report_dir."""
    report_dir = Path(report_dir)
    failed_ids = set()

    for xlsx in report_dir.glob("*.xlsx"):
        if xlsx.name.startswith("~$"):
            continue
        df = pd.read_excel(xlsx, sheet_name="detail")
        failed = df[df["decision"] == 0]
        failed_ids.update(failed["seq_id"].tolist())

    return failed_ids


def _extract_sequences(ref_path, target_ids, output_path):
    """Extract specific sequences from a FASTA file."""
    records = [r for r in SeqIO.parse(ref_path, "fasta") if r.id in target_ids]
    with open(output_path, "w") as f:
        SeqIO.write(records, f, "fasta")
    return len(records)


def _run_rescue_pipeline(primer_fa, subset_ref, features, params_file,
                         workdir, threads):
    """Run bowtie2 → sam2pairwise → process → prepare-input → evaluate.

    Returns path to rescue .eval, or None if no results.
    workdir is a persistent directory (not tmpdir).
    """
    workdir = Path(workdir)
    workdir.mkdir(parents=True, exist_ok=True)
    index_prefix = str(workdir / "subset")

    subprocess.run(
        ["bowtie2-build", "--quiet", str(subset_ref), index_prefix],
        check=True,
    )

    n_seqs = sum(1 for line in open(subset_ref) if line.startswith(">"))
    k_value = min(n_seqs * 5, 50000)
    sam_file = workdir / "rescue.sam"

    subprocess.run([
        "bowtie2", "-x", index_prefix,
        "-U", str(primer_fa), "-f",
        "-p", str(threads),
        "-k", str(k_value),
        "--mp", "2,2", "--np", "2",
        "--rdg", "4,4", "--rfg", "4,4",
        "-L", "8", "-N", "1",
        "--score-min", "L,-0.6,-0.6",
        "--no-hd", "--no-unal",
        "-S", str(sam_file),
    ], check=True)

    if not sam_file.exists() or sam_file.stat().st_size == 0:
        return None

    parsed_file = workdir / "rescue.parsed"
    with open(sam_file) as fin, open(parsed_file, "w") as fout:
        subprocess.run(["sam2pairwise"], stdin=fin, stdout=fout, check=True)

    # process_map: extract pairwise alignment columns
    mapped_file = workdir / "rescue.mapped"
    pseqs = str(parsed_file) + ".pseq"
    tseqs = str(parsed_file) + ".tseq"
    matches = str(parsed_file) + ".match"

    pf = shlex.quote(str(parsed_file))
    sf = shlex.quote(str(sam_file))
    mf = shlex.quote(str(mapped_file))

    subprocess.run(f'awk "NR % 4 == 2" {pf} > {shlex.quote(pseqs)}',
                   shell=True, check=True)
    subprocess.run(f'awk "NR % 4 == 3" {pf} > {shlex.quote(matches)}',
                   shell=True, check=True)
    subprocess.run(f'awk "NR % 4 == 0" {pf} > {shlex.quote(tseqs)}',
                   shell=True, check=True)
    subprocess.run(
        f"cut -f1-4 {sf} | "
        f"paste - {shlex.quote(pseqs)} | "
        f"paste - {shlex.quote(tseqs)} | "
        f"paste - {shlex.quote(matches)} > {mf}",
        shell=True, check=True,
    )

    if not mapped_file.exists() or mapped_file.stat().st_size == 0:
        return None

    # prepare-input (no coverage check for rescue)
    input_file = workdir / "rescue.input"
    subprocess.run([
        "qprimer", "prepare-input",
        "--in", str(mapped_file),
        "--out", str(input_file),
        "--ref", str(subset_ref),
        "--reftype", "on",
        "--features", str(features),
        "--params", str(params_file),
    ], check=True)

    if not input_file.exists() or input_file.stat().st_size == 0:
        return None

    # ML evaluate
    eval_out = workdir / "rescue.eval"
    subprocess.run([
        "qprimer", "evaluate",
        "--in", str(input_file),
        "--out", str(eval_out),
        "--ref", str(subset_ref),
        "--reftype", "on",
        "--threads", str(threads),
    ], check=True)

    rescue_full = Path(f"{eval_out}.full")
    if not rescue_full.exists() or rescue_full.stat().st_size == 0:
        return None

    return eval_out


def _export_rescue_report(rescue_eval_path, ref_path, primer_fa, report_dir,
                          probe_mapping=None, probe_seqs=None):
    """Generate xlsx reports from rescue evaluation results."""
    from Bio import SeqIO
    names = []
    with open(primer_fa) as f:
        for line in f:
            if line.startswith(">"):
                name = line[1:].strip()
                if name.endswith("_for"):
                    pair_id = name[:-4]
                    names.append(pair_id)

    if not names:
        return

    cmd = [
        "qprimer", "export-report",
        "--on", str(rescue_eval_path),
        "--out", str(report_dir),
        "--names", *names,
        "--ref", str(ref_path),
    ]
    if probe_mapping and Path(probe_mapping).exists():
        cmd.extend(["--probe-mapping-on", str(probe_mapping)])
    if probe_seqs and Path(probe_seqs).exists():
        cmd.extend(["--probe-seqs", str(probe_seqs)])

    subprocess.run(cmd, check=True)


def _update_summary_sensitivity(summary_df, detail_df):
    """Update Sensitivity section in summary based on updated detail."""
    total = len(detail_df)
    positive = detail_df[detail_df["decision"] == 1]
    n_positive = len(positive)
    regressors = pd.to_numeric(positive["regressor"], errors="coerce").dropna()

    sens_row = None
    for i, row in summary_df.iterrows():
        if str(row.iloc[0]).strip() == "Sensitivity":
            sens_row = i
            break

    if sens_row is None:
        return

    # Data rows start 2 rows after "Sensitivity" header (header + column names)
    data_start = sens_row + 2

    # Find where Specificity starts (or end of data)
    spec_row = None
    for i in range(data_start, len(summary_df)):
        val = str(summary_df.iloc[i, 0]).strip()
        if val == "Specificity" or val == "":
            spec_row = i
            break

    if spec_row is None:
        spec_row = len(summary_df)

    for i in range(data_start, spec_row):
        target_name = summary_df.iloc[i, 0]
        if pd.isna(target_name) or str(target_name).strip() == "":
            continue
        summary_df.iloc[i, 1] = f"{n_positive} / {total}"
        if len(regressors) > 0:
            summary_df.iloc[i, 2] = round(regressors.mean(), 3)
            summary_df.iloc[i, 3] = round(regressors.median(), 3)
            summary_df.iloc[i, 4] = round(regressors.min(), 3)
            summary_df.iloc[i, 5] = round(regressors.max(), 3)


def _apply_detail_formatting(xlsx_path):
    """Apply Courier New font, column width 40, and wrap text to align columns."""
    wb = load_workbook(xlsx_path)
    ws = wb["detail"]
    courier = Font(name="Courier New")
    wrap = Alignment(wrap_text=True)

    align_cols = [
        idx for idx, cell in enumerate(ws[1], start=1)
        if "align" in str(cell.value)
    ]
    for col_idx in align_cols:
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 40
        for row in ws.iter_rows(min_row=2):
            cell = row[col_idx - 1]
            if cell.value:
                cell.font = courier
                cell.alignment = wrap

    wb.save(xlsx_path)


def _merge_final_xlsx(original_dir, rescue_dir):
    """Merge rescue results into original xlsx, saving as _final.xlsx.

    For each primer set: if a seq_id had decision=0 in the original but
    decision=1 in the rescue, replace that row with the rescue row.
    """
    original_dir = Path(original_dir)
    rescue_dir = Path(rescue_dir)
    updated_count = 0

    for orig_xlsx in original_dir.glob("*.xlsx"):
        rescue_xlsx = rescue_dir / orig_xlsx.name
        if not rescue_xlsx.exists():
            continue

        orig_detail = pd.read_excel(orig_xlsx, sheet_name="detail")
        rescue_detail = pd.read_excel(rescue_xlsx, sheet_name="detail")

        rescue_improved = rescue_detail[rescue_detail["decision"] == 1]
        if rescue_improved.empty:
            continue

        improved_by_id = rescue_improved.set_index("seq_id")
        keep_original = {"seq_id", "target", "probe", "mm_p", "indel_p"}
        update_cols = [c for c in orig_detail.columns
                       if c in improved_by_id.columns and c not in keep_original]
        n_replaced = 0

        for idx, row in orig_detail.iterrows():
            if row["decision"] == 0 and row["seq_id"] in improved_by_id.index:
                for col in update_cols:
                    orig_detail.at[idx, col] = improved_by_id.at[row["seq_id"], col]
                n_replaced += 1

        if n_replaced == 0:
            continue

        orig_summary = pd.read_excel(orig_xlsx, sheet_name="summary",
                                      header=None)
        _update_summary_sensitivity(orig_summary, orig_detail)

        final_path = original_dir / orig_xlsx.name.replace(".xlsx", "_final.xlsx")
        with pd.ExcelWriter(final_path, engine="openpyxl") as writer:
            orig_summary.to_excel(writer, sheet_name="summary",
                                  index=False, header=False)
            orig_detail.to_excel(writer, sheet_name="detail", index=False)

        _apply_detail_formatting(final_path)

        updated_count += n_replaced
        print(f"  {orig_xlsx.name}: {n_replaced} row(s) rescued")

    # Rename _final.xlsx over original, remove rescue folder
    for final_xlsx in original_dir.glob("*_final.xlsx"):
        orig_name = final_xlsx.name.replace("_final.xlsx", ".xlsx")
        orig_path = original_dir / orig_name
        orig_path.unlink(missing_ok=True)
        final_xlsx.rename(orig_path)

    import shutil
    if rescue_dir.exists():
        shutil.rmtree(rescue_dir)

    return updated_count


def run(args):
    """Run the rescue-evaluate command."""
    if args.reftype != "on":
        print("Rescue only applies to on-target. Skipping.")
        return

    eval_path = Path(args.eval_path)
    full_path = Path(f"{eval_path}.full")

    if not full_path.exists() or full_path.stat().st_size == 0:
        print("No .full file found. Skipping rescue.")
        return

    failed = _find_failed_targets(args.report_dir)

    if not failed:
        print("No decision=0 targets found. No rescue needed.")
        return

    ref_total = sum(1 for _ in SeqIO.parse(args.ref, "fasta"))
    print(f"Rescue: {len(failed)}/{ref_total} target(s) need re-evaluation.")

    rescue_dir = eval_path.parent / "rescue"
    rescue_dir.mkdir(parents=True, exist_ok=True)
    subset_ref = rescue_dir / "subset.fa"

    n_extracted = _extract_sequences(args.ref, failed, subset_ref)
    print(f"  Extracted {n_extracted} sequence(s) to subset reference.")

    rescue_eval = _run_rescue_pipeline(
        primer_fa=args.fasta,
        subset_ref=subset_ref,
        features=args.features,
        params_file=args.params,
        workdir=rescue_dir,
        threads=args.threads,
    )

    if rescue_eval is None:
        print("  Rescue produced no new alignments.")
        return

    print(f"  Rescue eval saved to {rescue_eval}")

    rescue_report_dir = Path(args.report_dir) / "rescue"
    _export_rescue_report(
        rescue_eval_path=str(rescue_eval),
        ref_path=str(subset_ref),
        primer_fa=args.fasta,
        report_dir=str(rescue_report_dir),
    )
    print(f"  Rescue xlsx reports saved to {rescue_report_dir}")

    n_updated = _merge_final_xlsx(args.report_dir, str(rescue_report_dir))
    if n_updated:
        print(f"  Total {n_updated} row(s) updated in _final.xlsx files.")
    else:
        print("  No rows improved by rescue.")
