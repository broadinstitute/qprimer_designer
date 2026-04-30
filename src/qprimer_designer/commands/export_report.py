"""Export evaluation results to Excel reports with alignment details."""

import argparse
import ast
from pathlib import Path

import numpy as np
import pandas as pd
from Bio import SeqIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from qprimer_designer.utils import complement_dna
from qprimer_designer.external import compute_dimer_dg


def register(subparsers):
    """Register the export-report subcommand."""
    parser = subparsers.add_parser(
        "export-report",
        help="Export evaluation results to Excel reports",
        description="""
Export evaluation results to Excel format with detailed alignment information.
Creates one Excel file per primer set with:
  - Summary sheet: dimerization, sensitivity, and specificity metrics
  - Detail sheet: per-target alignments and scores
""",
    )
    parser.add_argument("--on", required=True, help="On-target evaluation file (without .full extension)")
    parser.add_argument("--off", nargs="*", default=[], help="Off-target evaluation files")
    parser.add_argument("--out", required=True, help="Output directory for Excel reports")
    parser.add_argument("--names", nargs="+", required=True, help="Primer set IDs to process")
    parser.add_argument("--off-ref", nargs="*", default=[], help="Off-target reference FASTA files (same order as --off)")
    parser.add_argument("--probe-mapping-on", default=None, help="Probe mapping CSV for on-target")
    parser.add_argument("--probe-mapping-off", nargs="*", default=[], help="Probe mapping CSVs for off-targets")
    parser.add_argument("--probe-seqs", default=None, help="Probe FASTA file")
    parser.add_argument("--ref", default=None, help="On-target reference FASTA (for counting total sequences)")
    parser.add_argument("--probe-max-mismatches", type=int, default=2,
                        help="Maximum mismatches for probe to count as matched (default: 2)")
    parser.set_defaults(func=run)


def filter_by_primer_id(df, pid, eval_type="on"):
    """Filter dataframe for a specific primer pair."""
    pid_for = f"{pid}_for"
    pid_rev = f"{pid}_rev"

    if eval_type == "on":
        mask = (df["pname_f"] == pid_for) & (df["pname_r"] == pid_rev)
    elif eval_type == "off":
        valid = {pid_for, pid_rev}
        mask = df["pname_f"].isin(valid) & df["pname_r"].isin(valid)
    else:
        raise ValueError(f"Unknown eval_type: {eval_type}")

    return df.loc[mask].copy()


def build_alignment_string(df, side):
    """
    Build formatted alignment string for display.

    Args:
        df: DataFrame with primer alignment data
        side: 'f' (forward) or 'r' (reverse)
    """
    if side == "f":
        return (
            "1     " + df["pseq_f"] + df["len_f"].astype(str).str.rjust(6) + "\n" +
            "      " + df["match_f"] + "      \n" +
            df["starts"].astype(str).str.ljust(6) + df["tseq_f"].apply(complement_dna) +
            (df["starts"] + df["tseq_f"].str.replace("-", "").str.len() - 1).astype(str).str.rjust(6)
        )
    else:
        # Target coordinates are shown high→low (3'→5') since primer is 5'→3'
        r_end = df["starts"] + df["prod_len"] - 1
        r_start = df["starts"] + df["prod_len"] - df["tseq_r"].str.replace("-", "").str.len()
        return (
            "1     " + df["pseq_r"] + df["len_r"].astype(str).str.rjust(6) + "\n" +
            "      " + df["match_r"].str[::-1] + "      \n" +
            r_end.astype(str).str.ljust(6) +
            df["tseq_r"].apply(complement_dna) +
            r_start.astype(str).str.rjust(6)
        )


def eval_to_target_df(eval_path, pid, eval_type="on"):
    """
    Load and process evaluation results for a specific primer set.

    Args:
        eval_path: Path to evaluation file (without .full extension)
        pid: Primer set ID
        eval_type: 'on' or 'off' target

    Returns:
        DataFrame with per-target results
    """
    df = pd.read_csv(f"{eval_path}.full")

    df = filter_by_primer_id(df, pid, eval_type)

    if df.empty:
        raise ValueError(f"No rows found for primer set {pid} in {eval_path}")

    # Explode target lists to per-target rows
    df["targets"] = df["targets"].apply(ast.literal_eval)
    df["starts"] = df["starts"].apply(ast.literal_eval)
    df = df.explode(["targets", "starts"])

    df = df.set_index("targets")
    df.index.name = "seq_id"

    # For off-target, filter to likely amplicons
    if eval_type == "off":
        df = df[df["classifier"] > 0.5].copy()
        if df.empty:
            return pd.DataFrame()

    # Build alignment strings
    df["align_f"] = build_alignment_string(df, "f")
    df["align_r"] = build_alignment_string(df, "r")

    # Binarize classifier for coverage counting
    df["classifier"] = (df["classifier"] > 0.5).astype(int)

    # Drop raw alignment columns
    df = df.drop(columns=["match_f", "tseq_f", "match_r", "tseq_r"])

    # Keep only the best alignment per sequence (highest regressor score)
    if eval_type == "on":
        df = df.reset_index()
        df["_reg_numeric"] = pd.to_numeric(df["regressor"], errors="coerce")
        df = df.sort_values("_reg_numeric", ascending=False).drop_duplicates(
            subset="seq_id", keep="first"
        )
        df = df.drop(columns=["_reg_numeric"]).set_index("seq_id")

    return df


def annotate_probe(df, probe_mapping_path, pid, max_mismatches=2):
    """
    Add probe, mm_p, and indel_p columns to detail dataframe.

    probe = 1 if probe aligns within amplicon AND mismatches <= max_mismatches AND indels == 0.
    mm_p = mismatch count (substitutions only, recorded whenever probe is within amplicon).
    indel_p = indel count (recorded whenever probe is within amplicon).
    """
    df["probe"] = 0
    df["mm_p"] = np.nan
    df["indel_p"] = np.nan

    if probe_mapping_path is None:
        return df

    probe_df = pd.read_csv(probe_mapping_path)
    if probe_df.empty:
        return df

    # Filter to probes matching this primer set
    probe_name = f"{pid}_pro"
    probe_df = probe_df[probe_df["probe_name"] == probe_name]
    if probe_df.empty:
        return df

    # Ensure indels column exists (backward compatibility)
    if "indels" not in probe_df.columns:
        probe_df["indels"] = 0

    # Reset index to avoid duplicate index issues, then restore
    df = df.reset_index()

    for i, row in df.iterrows():
        seq_id = row["seq_id"]
        amp_start = row["starts"]
        amp_end = amp_start + row["prod_len"]

        # Find probe alignments for this target sequence
        hits = probe_df[probe_df["target_id"] == seq_id]
        for _, hit in hits.iterrows():
            probe_start = hit["start_pos"]
            probe_end = probe_start + len(hit["probe_seq"])
            # Probe is within amplicon if it's fully contained
            if probe_start >= amp_start and probe_end <= amp_end:
                mm = hit["mismatches"]
                indel = hit["indels"]
                df.at[i, "mm_p"] = mm
                df.at[i, "indel_p"] = indel
                if mm <= max_mismatches and indel == 0:
                    df.at[i, "probe"] = 1
                break

    df["probe"] = df["probe"].astype(int)
    df = df.set_index("seq_id")
    return df


def add_decision_columns(df, has_probe=False):
    """Add decision and reason columns based on classifier and probe status.

    decision = 1 if classifier=1 (and probe=1 when probe mode is enabled).
    reason = explanation when decision=0.
    """
    df["decision"] = 0
    df["reason"] = ""

    if has_probe:
        mask_pass = (df["classifier"] == 1) & (df["probe"] == 1)
        df.loc[mask_pass, "decision"] = 1

        mask_unmapped = df["regressor"] == "unmapped"
        mask_amp_fail = (df["classifier"] == 0) & ~mask_unmapped
        mask_probe_fail = (df["classifier"] == 1) & (df["probe"] == 0)

        df.loc[mask_unmapped, "reason"] = "unmapped"
        df.loc[mask_amp_fail, "reason"] = "amplification fails"
        # Distinguish probe failure reasons
        if "indel_p" in df.columns:
            mask_indel = mask_probe_fail & (df["indel_p"] > 0)
            mask_mm = mask_probe_fail & (df["indel_p"] == 0) & (df["mm_p"].notna())
            mask_no_map = mask_probe_fail & (df["mm_p"].isna())
            df.loc[mask_indel, "reason"] = "probe indel"
            df.loc[mask_mm, "reason"] = "probe mismatch"
            df.loc[mask_no_map, "reason"] = "probe unmapped"
        else:
            df.loc[mask_probe_fail, "reason"] = "probe fails"
    else:
        mask_pass = df["classifier"] == 1
        df.loc[mask_pass, "decision"] = 1

        mask_unmapped = df["regressor"] == "unmapped"
        mask_amp_fail = (df["classifier"] == 0) & ~mask_unmapped

        df.loc[mask_unmapped, "reason"] = "unmapped"
        df.loc[mask_amp_fail, "reason"] = "amplification fails"

    return df


def build_dimerization_table(df, probe_seq=None):
    """
    Build dimerization table.

    2×2 (forward/reverse) when no probe, 3×3 when probe is provided.

    Returns:
        DataFrame with ΔG values for all primer-primer combinations
    """
    df_on = df[df["eval_type"] == "on"]

    fseq = df_on["pseq_f"].values[0].replace("-", "")
    rseq = df_on["pseq_r"].values[0].replace("-", "")

    if probe_seq:
        seqs = [fseq, rseq, probe_seq]
        labels = ["forward", "reverse", "probe"]
    else:
        seqs = [fseq, rseq]
        labels = ["forward", "reverse"]

    data = [[compute_dimer_dg(s1, s2) for s2 in seqs] for s1 in seqs]

    return pd.DataFrame(data, index=labels, columns=labels).round(1)


def _get_ref_seq_ids(ref_path):
    """Get all sequence IDs from a reference FASTA file."""
    if not ref_path:
        return None
    seq_ids = []
    for rec in SeqIO.parse(ref_path, "fasta"):
        seq_ids.append(rec.id)
    return seq_ids


def _deduplicate_coverage(df_subset, has_probe=False, ref_total=None):
    """Compute coverage by unique seq_id.

    For each unique seq_id, the sequence is "covered" if ANY alignment
    has classifier=1 (and probe=1 if applicable).

    Args:
        ref_total: Total number of sequences in the reference FASTA.
                   If provided, used as denominator instead of counting from df.

    Returns (n_covered, n_total, covered_regressor_values).
    """
    df_reset = df_subset.reset_index()

    if has_probe:
        df_reset["_covered"] = (df_reset["classifier"] == 1) & (df_reset["probe"] == 1)
    else:
        df_reset["_covered"] = df_reset["classifier"] == 1

    # Per seq_id: covered if any alignment is covered
    seq_covered = df_reset.groupby("seq_id")["_covered"].any()
    n_total = ref_total if ref_total is not None else len(seq_covered)
    n_covered = int(seq_covered.sum())

    # Regressor values for covered sequences (best per seq_id)
    covered_seqs = seq_covered[seq_covered].index
    act = df_reset[df_reset["seq_id"].isin(covered_seqs)].groupby("seq_id")["regressor"].max()

    return n_covered, n_total, act


def build_sensitivity_table(df, has_probe=False, ref_total=None):
    """Build sensitivity summary for on-target."""
    df_on = df[df["eval_type"] == "on"]

    if df_on.empty:
        return pd.DataFrame()

    n_covered, n_total, act = _deduplicate_coverage(df_on, has_probe, ref_total)
    coverage_str = f"{n_covered} / {n_total}"

    return pd.DataFrame(
        {
            "Coverage": [coverage_str],
            "Act_mean": [round(act.mean(), 3) if len(act) > 0 else np.nan],
            "Act_median": [round(act.median(), 3) if len(act) > 0 else np.nan],
            "Act_min": [round(act.min(), 3) if len(act) > 0 else np.nan],
            "Act_max": [round(act.max(), 3) if len(act) > 0 else np.nan],
        },
        index=[df["target"].values[0]],
    )


def build_specificity_table(df, has_probe=False):
    """Build specificity summary for off-target."""
    df_off = df[df["eval_type"] == "off"]

    if df_off.empty:
        return pd.DataFrame()

    results = []
    for target, group in df_off.groupby("target"):
        # Check if this is a placeholder (no real amplification predicted)
        is_placeholder = (
            len(group) == 1
            and group.index[0] == "(none)"
        )

        if is_placeholder:
            off_total = group["_off_ref_total"].iloc[0]
            off_total = int(off_total) if pd.notna(off_total) else 0
            results.append({
                "target": target,
                "Coverage": f"0 / {off_total}",
                "Act_mean": np.nan,
                "Act_median": np.nan,
                "Act_min": np.nan,
                "Act_max": np.nan,
            })
        else:
            n_covered, n_total, act = _deduplicate_coverage(group, has_probe)
            results.append({
                "target": target,
                "Coverage": f"{n_covered} / {n_total}",
                "Act_mean": round(act.mean(), 3) if len(act) > 0 else np.nan,
                "Act_median": round(act.median(), 3) if len(act) > 0 else np.nan,
                "Act_min": round(act.min(), 3) if len(act) > 0 else np.nan,
                "Act_max": round(act.max(), 3) if len(act) > 0 else np.nan,
            })

    return pd.DataFrame(results).set_index("target")


def save_eval_excel(df, outdir, filename, probe_seq=None, ref_total=None):
    """
    Save evaluation results to Excel with formatted alignment columns.

    Creates two sheets:
      - summary: dimerization, sensitivity, specificity tables
      - detail: per-target alignment and scoring data
    """
    outdir = Path(outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    xlsx_path = outdir / filename

    has_probe = probe_seq is not None

    # Drop internal columns from detail sheet
    drop_cols = ["pseq_f", "pseq_r", "_off_ref_total"]
    drop_cols = [c for c in drop_cols if c in df.columns]
    # Remove placeholder rows from detail sheet
    df_detail = df[df.index != "(none)"]
    df_detail.drop(columns=drop_cols).to_excel(xlsx_path, sheet_name="detail")

    wb = load_workbook(xlsx_path)

    # Format alignment columns in detail sheet
    ws = wb["detail"]

    courier = Font(name="Courier New")
    wrap = Alignment(wrap_text=True)

    align_cols = [
        idx for idx, cell in enumerate(ws[1], start=1)
        if "align" in str(cell.value)
    ]

    for col_idx in align_cols:
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 50
        for row in ws.iter_rows(min_row=2):
            cell = row[col_idx - 1]
            if cell.value:
                cell.font = courier
                cell.alignment = wrap

    # Create summary sheet
    ws_sum = wb.create_sheet("summary")
    wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(ws_sum)))

    dimer_df = build_dimerization_table(df, probe_seq=probe_seq)
    sens_df = build_sensitivity_table(df, has_probe=has_probe, ref_total=ref_total)
    spec_df = build_specificity_table(df, has_probe=has_probe)

    current_row = 1

    # Write primer/probe sequences at top of summary
    df_on = df[df["eval_type"] == "on"]
    fseq = df_on["pseq_f"].values[0].replace("-", "")
    rseq = df_on["pseq_r"].values[0].replace("-", "")
    if probe_seq:
        seq_headers = ["Forward", "Reverse", "Probe"]
        seq_values = [fseq, rseq, probe_seq]
    else:
        seq_headers = ["Forward", "Reverse"]
        seq_values = [fseq, rseq]
    for col_idx, h in enumerate(seq_headers, start=1):
        ws_sum.cell(row=current_row, column=col_idx, value=h).font = Font(bold=True)
    current_row += 1
    for col_idx, v in enumerate(seq_values, start=1):
        ws_sum.cell(row=current_row, column=col_idx, value=v)
    current_row += 2

    def write_table(title, table):
        nonlocal current_row

        ws_sum.cell(row=current_row, column=1, value=title).font = Font(bold=True)
        current_row += 1

        if table.empty:
            ws_sum.cell(row=current_row, column=1, value="(no data)")
            current_row += 2
            return

        for col_idx, col in enumerate([""] + list(table.columns), start=1):
            ws_sum.cell(row=current_row, column=col_idx, value=col).font = Font(bold=True)
        current_row += 1

        for idx, row in table.iterrows():
            ws_sum.cell(row=current_row, column=1, value=idx)
            for col_idx, val in enumerate(row, start=2):
                ws_sum.cell(row=current_row, column=col_idx, value=val)
            current_row += 1

        current_row += 2

    write_table("Dimerization", dimer_df)
    write_table("Sensitivity", sens_df)
    write_table("Specificity", spec_df)

    wb.save(xlsx_path)


def get_target_name(path):
    """Extract target name from evaluation file path."""
    return Path(path).name.split(".")[1]


def add_unmapped_rows(df, ref_seq_ids):
    """Add rows for sequences in reference that have no alignment.

    Unmapped sequences get classifier=0, regressor='unmapped'.
    """
    if ref_seq_ids is None:
        return df

    mapped_ids = set(df.index)
    unmapped_ids = [sid for sid in ref_seq_ids if sid not in mapped_ids]

    if not unmapped_ids:
        return df

    # Build unmapped rows with same columns
    unmapped_rows = []
    for sid in unmapped_ids:
        row = {col: np.nan for col in df.columns}
        row["classifier"] = 0
        row["regressor"] = "unmapped"
        row["eval_type"] = df["eval_type"].iloc[0] if not df.empty else "on"
        row["target"] = df["target"].iloc[0] if not df.empty else ""
        unmapped_rows.append(row)

    unmapped_df = pd.DataFrame(unmapped_rows, index=pd.Index(unmapped_ids, name="seq_id"))
    return pd.concat([df, unmapped_df], axis=0)


def load_probe_seq(probe_seqs_path, pid):
    """Load probe sequence for a given primer set ID from FASTA."""
    if not probe_seqs_path:
        return None
    probe_name = f"{pid}_pro"
    for rec in SeqIO.parse(probe_seqs_path, "fasta"):
        if rec.id == probe_name:
            return str(rec.seq)
    return None


def run(args):
    """Run the export-report command."""
    outdir = Path(args.out)
    outdir.mkdir(parents=True, exist_ok=True)

    print(f"Output directory: {outdir}")
    print(f"Primer sets to evaluate: {', '.join(args.names)}")

    has_probe = args.probe_mapping_on is not None
    probe_mapping_on = args.probe_mapping_on
    probe_max_mm = args.probe_max_mismatches
    probe_mapping_off = {
        get_target_name(p): p for p in args.probe_mapping_off
    } if args.probe_mapping_off else {}

    # Load reference seq_ids for unmapped row detection
    ref_seq_ids = _get_ref_seq_ids(args.ref) if args.ref else None
    ref_total = len(ref_seq_ids) if ref_seq_ids else None

    # Build off-target reference totals: target_name → seq count
    off_ref_totals = {}
    for off_path, off_ref_path in zip(args.off, args.off_ref):
        target_name = get_target_name(off_path)
        off_ref_ids = _get_ref_seq_ids(off_ref_path)
        if off_ref_ids is not None:
            off_ref_totals[target_name] = len(off_ref_ids)

    cols = [
        "target", "eval_type", "decision", "reason",
        "classifier", "regressor",
        "pname_f", "pname_r", "pseq_f", "pseq_r",
        "align_f", "align_r",
        "prod_len", "prod_Tm",
        "mm_f", "indel_f", "len_f", "Tm_f", "GC_f",
        "mm_r", "indel_r", "len_r", "Tm_r", "GC_r"
    ]

    if has_probe:
        cols += ["probe", "mm_p", "indel_p"]

    for pid in args.names:
        dfs = []
        probe_seq = load_probe_seq(args.probe_seqs, pid) if has_probe else None

        # Process on-target evaluation
        try:
            df_on = eval_to_target_df(
                eval_path=args.on,
                pid=pid,
                eval_type="on"
            )
            df_on["eval_type"] = "on"
            df_on["target"] = get_target_name(args.on)
            if has_probe:
                df_on = annotate_probe(df_on, probe_mapping_on, pid, probe_max_mm)
            # Add unmapped sequences from reference
            df_on = add_unmapped_rows(df_on, ref_seq_ids)
            df_on = add_decision_columns(df_on, has_probe)
            dfs.append(df_on[cols].sort_values(
                "regressor", ascending=False,
                key=lambda s: pd.to_numeric(s, errors="coerce"),
            ))
        except Exception as e:
            print(f"[WARNING] on-target eval failed for {pid}: {e}")

        # Process off-target evaluations
        for off_path in args.off:
            try:
                target_name = get_target_name(off_path)

                try:
                    df_off = eval_to_target_df(
                        eval_path=off_path,
                        pid=pid,
                        eval_type="off"
                    )
                except ValueError:
                    # No alignments found for this primer set against this
                    # off-target — treat as zero amplification.
                    df_off = pd.DataFrame()

                if not df_off.empty:
                    df_off["eval_type"] = "off"
                    df_off["target"] = target_name
                    if has_probe:
                        off_probe_path = probe_mapping_off.get(target_name)
                        df_off = annotate_probe(df_off, off_probe_path, pid, probe_max_mm)
                    df_off = add_decision_columns(df_off, has_probe)
                    dfs.append(df_off[cols].sort_values("regressor", ascending=False))
                else:
                    # No predicted amplification — add a placeholder row so
                    # the specificity table can show "0 / N"
                    placeholder = {col: np.nan for col in cols}
                    placeholder["eval_type"] = "off"
                    placeholder["target"] = target_name
                    placeholder["classifier"] = 0
                    placeholder["regressor"] = "no_amplification"
                    placeholder["decision"] = 0
                    placeholder["reason"] = ""
                    placeholder["_off_ref_total"] = off_ref_totals.get(target_name)
                    if has_probe:
                        placeholder["probe"] = 0
                    dfs.append(pd.DataFrame([placeholder], index=pd.Index(["(none)"], name="seq_id")))

            except Exception as e:
                print(f"[WARNING] off-target eval failed for {pid} ({off_path}): {e}")

        # Combine and save
        if not dfs:
            print(f"[WARNING] No evaluation results for primer set {pid}, skipping")
            continue

        df_all = pd.concat(dfs, axis=0)
        save_eval_excel(df_all, outdir=outdir, filename=f"{pid}.xlsx",
                        probe_seq=probe_seq, ref_total=ref_total)
        print(f"Created {outdir / f'{pid}.xlsx'}")

    print("All primer sets processed")
    (outdir / "done").touch()
