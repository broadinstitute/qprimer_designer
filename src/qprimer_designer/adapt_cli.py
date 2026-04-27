"""adapt - User-facing CLI for running the ADAPT PCR primer design pipeline."""

import argparse
import csv
import io
import os
import re
import shutil
import smtplib
import subprocess
import sys
import urllib.request
from datetime import datetime, timedelta
from email.message import EmailMessage
from pathlib import Path

from qprimer_designer.utils.params import parse_params, parse_list_param


# Resolve template path relative to package location
# Package is at src/qprimer_designer/, template is at workflows/
TEMPLATE_PATH = Path(__file__).resolve().parent.parent.parent / "workflows" / "Snakefile.template"


def _setup_run_dir(run_id: str, params_file: Path, targets: list[str]) -> Path:
    """Create run directory and copy inputs into it.

    Args:
        run_id: Run identifier (used as directory name).
        params_file: Path to the source params.txt.
        targets: List of all target names whose FASTA files should be copied.

    Returns:
        Path to the created run directory.
    """
    run_dir = Path(run_id)
    if run_dir.exists():
        print(f"Error: run directory '{run_dir}' already exists.", file=sys.stderr)
        sys.exit(1)

    run_dir.mkdir(parents=True)

    # Copy params.txt
    shutil.copy2(params_file, run_dir / "params.txt")

    # Copy target FASTA files
    target_seq_dir = run_dir / "target_seqs" / "original"
    target_seq_dir.mkdir(parents=True)

    source_dir = Path("target_seqs") / "original"
    if not source_dir.is_dir():
        print(f"Error: target_seqs/original/ directory not found.", file=sys.stderr)
        sys.exit(1)

    for name in targets:
        # Try common FASTA extensions
        found = False
        for ext in (".fa", ".fasta", ".fna"):
            src = source_dir / f"{name}{ext}"
            if src.exists():
                shutil.copy2(src, target_seq_dir / f"{name}{ext}")
                found = True
                break
        if not found:
            print(f"Warning: FASTA file for '{name}' not found in {source_dir}", file=sys.stderr)

    return run_dir


def _build_snakefile(
    targets: list[str],
    cross: list[str],
    host: list[str],
    panel: list[str],
) -> str:
    """Read Snakefile.template and substitute header variable assignments."""
    template = TEMPLATE_PATH.read_text()

    def _list_literal(items: list[str]) -> str:
        return "[" + ", ".join(f"'{i}'" for i in items) + "]"

    replacements = {
        r"^TARGETS\s*=\s*\[.*?\]": f"TARGETS = {_list_literal(targets)}",
        r"^CROSS\s*=\s*\[.*?\]": f"CROSS   = {_list_literal(cross)}",
        r"^HOST\s*=\s*\[.*?\]": f"HOST    = {_list_literal(host)}",
        r"^PANEL\s*=\s*\[.*?\]": f"PANEL = {_list_literal(panel)}",
        r'^TARGET_DIR\s*=\s*".*?"': 'TARGET_DIR = "target_seqs/original"',
        r'^PARAMS\s*=\s*".*?"': 'PARAMS = "params.txt"',
        r'^RUN_ID\s*=\s*".*?"': 'RUN_ID = ""',
    }

    result = template
    for pattern, replacement in replacements.items():
        result = re.sub(pattern, replacement, result, count=1, flags=re.MULTILINE)

    return result


def _generate_snakefile(
    run_dir: Path,
    targets: list[str],
    cross: list[str],
    host: list[str],
    panel: list[str],
) -> None:
    """Generate Snakefile in the run directory from the template."""
    snakefile_content = _build_snakefile(targets, cross, host, panel)
    (run_dir / "Snakefile").write_text(snakefile_content)


def _run_snakemake(run_dir: Path, config_args: list[str], cores: int, dry_run: bool) -> int:
    """Execute snakemake in the run directory.

    Returns the process exit code.
    """
    cmd = ["snakemake", "-s", "Snakefile", "--cores", str(cores)]

    if config_args:
        cmd += ["--config"] + config_args

    if dry_run:
        cmd.append("--dry-run")

    print(f"Running: {' '.join(cmd)}")
    print(f"Working directory: {run_dir.resolve()}")

    result = subprocess.run(cmd, cwd=str(run_dir))
    return result.returncode


def cmd_design(args):
    """Run the design pipeline."""
    params_file = Path(args.params)
    if not params_file.exists():
        print(f"Error: params file '{params_file}' not found.", file=sys.stderr)
        sys.exit(1)

    params = parse_params(params_file)

    targets = parse_list_param(params, "TARGETS")
    cross = parse_list_param(params, "CROSS")
    host = parse_list_param(params, "HOST")
    panel = parse_list_param(params, "PANEL")

    if args.multiplex:
        if not panel:
            print("Error: --multiplex requires PANEL to be set in params.txt", file=sys.stderr)
            sys.exit(1)
        all_targets = list(set(panel + host))
    else:
        if not targets:
            print("Error: TARGETS must be set in params.txt", file=sys.stderr)
            sys.exit(1)
        all_targets = list(set(targets + cross + host))

    run_id = args.runid or datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = _setup_run_dir(run_id, params_file, all_targets)

    _generate_snakefile(run_dir, targets, cross, host, panel)

    config_args = []
    if args.probe:
        config_args.append("probe=1")
    if args.multiplex:
        config_args.append("multiplex=1")

    rc = _run_snakemake(run_dir, config_args, args.cores, args.dry_run)
    sys.exit(rc)


def cmd_evaluate(args):
    """Run the evaluate pipeline."""
    params_file = Path(args.params)
    if not params_file.exists():
        print(f"Error: params file '{params_file}' not found.", file=sys.stderr)
        sys.exit(1)

    params = parse_params(params_file)

    targets = parse_list_param(params, "TARGETS")
    cross = parse_list_param(params, "CROSS")
    host = parse_list_param(params, "HOST")

    if not targets:
        print("Error: TARGETS must be set in params.txt for evaluate mode", file=sys.stderr)
        sys.exit(1)

    all_targets = list(set(targets + cross + host))

    run_id = args.runid or datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = _setup_run_dir(run_id, params_file, all_targets)

    _generate_snakefile(run_dir, targets, cross, host, panel=[])

    config_args = ["evaluate=1"]
    if args.pset:
        pset_path = Path(args.pset).resolve()
        # Copy pset file into run dir for reproducibility
        shutil.copy2(pset_path, run_dir / pset_path.name)
        config_args.append(f"pset={pset_path.name}")
    else:
        config_args.append(f"for={args.forward}")
        config_args.append(f"rev={args.reverse}")
        if args.probe:
            config_args.append(f"pro={args.probe}")

    rc = _run_snakemake(run_dir, config_args, args.cores, args.dry_run)
    sys.exit(rc)


# ---------------------------------------------------------------------------
# fetch helpers
# ---------------------------------------------------------------------------

# Boolean flags: only add --flag when value is TRUE
_BOOLEAN_FLAGS = {
    "refseq_only",
    "is_sars_cov2",
    "is_alphainfluenza",
    "genbank_metadata",
    "proteins_complete",
}

# Flags that accept explicit true/false values
_TRUE_FALSE_FLAGS = {
    "annotated",
    "lab_passaged",
}

# Columns that are metadata only (not passed to gget)
_METADATA_COLUMNS = {"Pathogen", "Pathogen_abb", "Gene Segment", "Target name",
                      "Primer name", "query_id", "expected_count", "latest_count",
                      "diff", "Forward", "Reverse", "Probe"}

# Max-date columns to strip when monitoring (monitor always fetches up to the latest)
_MAX_DATE_COLUMNS = {"max_collection_date", "max_release_date"}


def _extract_spreadsheet_id(url: str) -> str:
    """Extract the spreadsheet ID from a Google Sheets URL."""
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if not match:
        print(f"Error: cannot extract spreadsheet ID from URL: {url}", file=sys.stderr)
        sys.exit(1)
    return match.group(1)


def _download_spreadsheet_csv(spreadsheet_id: str) -> str:
    """Download a Google Sheets spreadsheet as CSV (must be publicly shared)."""
    export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv"

    # Try urllib first, fall back to curl if SSL fails
    try:
        with urllib.request.urlopen(export_url) as resp:
            return resp.read().decode("utf-8")
    except urllib.error.URLError:
        pass

    # Fallback: use curl
    try:
        result = subprocess.run(
            ["curl", "-s", "-L", export_url],
            capture_output=True, text=True, check=True,
        )
        return result.stdout
    except (subprocess.CalledProcessError, FileNotFoundError) as e:
        print(f"Error downloading spreadsheet: {e}", file=sys.stderr)
        print("Make sure the spreadsheet is shared with 'Anyone with the link'.", file=sys.stderr)
        sys.exit(1)


def _is_empty(value) -> bool:
    """Check if a value is empty/missing."""
    if value is None:
        return True
    val_str = str(value).strip().lower()
    return val_str in ("", "n/a", "na", "nan", "none")


def _parse_boolean(value) -> str | None:
    """Parse a boolean-like value. Returns 'true', 'false', or None."""
    if _is_empty(value):
        return None
    val_str = str(value).strip().upper()
    if val_str == "TRUE":
        return "true"
    if val_str == "FALSE":
        return "false"
    return None


def _load_spreadsheet(csv_text: str) -> tuple[list[str], list[dict]]:
    """Parse the mastersheet CSV and return (headers, data_rows).

    Spreadsheet layout:
      Row 0: category headers (ignored)
      Row 1: column headers
      Row 2: descriptions/examples (ignored)
      Row 3+: data
    """
    reader = csv.reader(io.StringIO(csv_text))
    rows = list(reader)

    if len(rows) < 4:
        print("Error: spreadsheet must have at least 4 rows.", file=sys.stderr)
        sys.exit(1)

    headers = rows[1]

    # Find TaxID column — gget arguments start here
    try:
        taxid_idx = headers.index("TaxID")
    except ValueError:
        print("Error: 'TaxID' column not found in spreadsheet.", file=sys.stderr)
        sys.exit(1)

    # Find metadata column indices
    metadata_indices = {}
    for col in _METADATA_COLUMNS:
        if col in headers:
            metadata_indices[col] = headers.index(col)

    data_rows = []
    for row in rows[3:]:
        if len(row) <= taxid_idx:
            continue

        row_dict = {}
        # Columns from TaxID onwards → gget arguments
        for i in range(taxid_idx, len(headers)):
            if i < len(row):
                row_dict[headers[i]] = row[i]
            else:
                row_dict[headers[i]] = ""

        # Add metadata columns
        for col, idx in metadata_indices.items():
            if idx < len(row):
                row_dict[col] = row[idx]
            else:
                row_dict[col] = ""

        if _is_empty(row_dict.get("TaxID")):
            continue

        data_rows.append(row_dict)

    return headers, data_rows


def _build_gget_command(row: dict, output_dir: str) -> list[str] | None:
    """Build a gget virus command from a spreadsheet row."""
    taxid = row.get("TaxID")
    if _is_empty(taxid):
        return None

    cmd = ["gget", "virus", str(taxid).strip(), "--out", output_dir]

    for col, value in row.items():
        if col in _METADATA_COLUMNS or col == "TaxID":
            continue
        if _is_empty(value):
            continue

        if col in _BOOLEAN_FLAGS:
            if _parse_boolean(value) == "true":
                cmd.append(f"--{col}")
        elif col in _TRUE_FALSE_FLAGS:
            bool_val = _parse_boolean(value)
            if bool_val is not None:
                cmd.extend([f"--{col}", bool_val])
        else:
            cmd.extend([f"--{col}", str(value).strip()])

    return cmd


def _make_target_name(row: dict) -> str:
    """Get target name from the 'Target name' column, falling back to Pathogen_abb."""
    target_name = str(row.get("Target name", "")).strip()
    if target_name:
        return target_name

    abb = str(row.get("Pathogen_abb", "")).strip()
    if not abb:
        abb = str(row.get("Pathogen", "unknown")).strip().replace(" ", "_")

    segment = str(row.get("Gene Segment", "")).strip()
    if segment and segment.lower() != "non-segmented":
        return f"{abb}_{segment}"
    return abb


def _fasta_stats(fasta_path: Path) -> None:
    """Print sequence count and length distribution for a FASTA file."""
    lengths = []
    current_len = 0
    with open(fasta_path) as f:
        for line in f:
            if line.startswith(">"):
                if current_len > 0:
                    lengths.append(current_len)
                current_len = 0
            else:
                current_len += len(line.strip())
        if current_len > 0:
            lengths.append(current_len)

    if not lengths:
        print(f"  {fasta_path.name}: 0 sequences")
        return

    lengths.sort()
    n = len(lengths)
    p5 = lengths[max(0, int(n * 0.05))]
    p95 = lengths[min(n - 1, int(n * 0.95))]
    med = lengths[n // 2]

    print(f"  {fasta_path.name}: {n} sequences, "
          f"median {med:,} bp, range {lengths[0]:,}–{lengths[-1]:,} bp")

    # Histogram between p5 and p95
    if p5 == p95 or n < 3:
        return

    n_bins = 10
    bin_width = (p95 - p5) / n_bins
    bins = [0] * n_bins
    below = 0
    above = 0
    for l in lengths:
        if l < p5:
            below += 1
        elif l >= p95:
            above += 1
        else:
            idx = min(int((l - p5) / bin_width), n_bins - 1)
            bins[idx] += 1

    max_count = max(bins) if max(bins) > 0 else 1
    bar_max = 30

    print(f"  Length distribution (p5={p5:,}, p95={p95:,}):")
    if below:
        print(f"    <{p5:>8,}: {below:>5}")
    for i, count in enumerate(bins):
        lo = int(p5 + i * bin_width)
        hi = int(p5 + (i + 1) * bin_width)
        bar = "█" * int(count / max_count * bar_max) if count else ""
        print(f"    {lo:>8,}–{hi:<8,}: {count:>5} {bar}")
    if above:
        print(f"    >{p95:>8,}: {above:>5}")


def cmd_fetch(args):
    """Fetch virus sequences from NCBI using gget virus."""
    if not args.dry_run and not shutil.which("gget"):
        print("Error: 'gget' is not installed. Run: pip install gget", file=sys.stderr)
        sys.exit(1)

    params_file = Path(args.params)
    if not params_file.exists():
        print(f"Error: params file '{params_file}' not found.", file=sys.stderr)
        sys.exit(1)

    params = parse_params(params_file)

    # Get spreadsheet URL from params or CLI
    spreadsheet_url = args.url or str(params.get("SPREADSHEET_URL", "")).strip()
    if not spreadsheet_url:
        print("Error: no spreadsheet URL. Set SPREADSHEET_URL in params.txt or use --url.", file=sys.stderr)
        sys.exit(1)

    # Get query IDs to filter
    if args.query_ids:
        query_ids = {q.strip() for q in args.query_ids}
    else:
        query_ids_param = parse_list_param(params, "QUERY_IDS")
        query_ids = set(query_ids_param) if query_ids_param else None

    # Download and parse spreadsheet
    spreadsheet_id = _extract_spreadsheet_id(spreadsheet_url)
    print(f"Downloading spreadsheet {spreadsheet_id}...")
    csv_text = _download_spreadsheet_csv(spreadsheet_id)
    _, data_rows = _load_spreadsheet(csv_text)

    # Filter by query IDs if specified
    if query_ids is not None:
        data_rows = [r for r in data_rows if str(r.get("query_id", "")).strip() in query_ids]
        if not data_rows:
            print(f"Error: no rows match query IDs: {query_ids}", file=sys.stderr)
            sys.exit(1)

    print(f"Processing {len(data_rows)} query(ies)...")

    # Ensure output directory exists
    out_dir = Path(args.outdir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # Save spreadsheet CSV with timestamp for reproducibility
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    mastersheet_name = f"mastersheet_{timestamp}.csv"
    csv_path = out_dir / mastersheet_name
    csv_path.write_text(csv_text, encoding="utf-8")
    print(f"Saved spreadsheet to {csv_path}")

    results = []
    for i, row in enumerate(data_rows):
        target_name = _make_target_name(row)
        query_id = str(row.get("query_id", "")).strip()
        pathogen = str(row.get("Pathogen", "")).strip()

        # Temporary output dir for gget
        gget_out = out_dir / f".gget_tmp_{query_id}_{target_name}"
        gget_out.mkdir(parents=True, exist_ok=True)

        cmd = _build_gget_command(row, str(gget_out))
        if not cmd:
            print(f"  [{i+1}/{len(data_rows)}] Skipping {pathogen} (no TaxID)")
            continue

        print(f"\n[{i+1}/{len(data_rows)}] {pathogen} → {target_name}")
        print(f"  Command: {' '.join(cmd)}")

        if args.dry_run:
            results.append({"name": target_name, "status": "dry-run"})
            shutil.rmtree(gget_out, ignore_errors=True)
            continue

        try:
            subprocess.run(cmd, check=True)

            # Find the output FASTA and move to target_seqs/original/
            fasta_files = list(gget_out.glob("*.fa")) + list(gget_out.glob("*.fasta"))
            if fasta_files:
                dest = out_dir / f"{target_name}.fa"
                shutil.move(str(fasta_files[0]), str(dest))
                # Remove duplicate sequences
                n_deduped = _deduplicate_fasta(dest)
                if n_deduped > 0:
                    print(f"  Removed {n_deduped} duplicate sequence(s)")
                # Count sequences
                with open(dest) as f:
                    n_seqs = sum(1 for line in f if line.startswith(">"))
                print(f"  → {dest} ({n_seqs} unique sequences)")
                results.append({"name": target_name, "status": "success", "count": n_seqs, "path": dest})
            else:
                print(f"  Warning: no FASTA output found in {gget_out}")
                results.append({"name": target_name, "status": "no_output"})
        except subprocess.CalledProcessError as e:
            print(f"  Error: gget failed with exit code {e.returncode}")
            results.append({"name": target_name, "status": "error"})
        finally:
            shutil.rmtree(gget_out, ignore_errors=True)

    # Summary
    success = [r for r in results if r["status"] == "success"]
    errors = [r for r in results if r["status"] == "error"]
    print(f"\nDone: {len(success)} succeeded, {len(errors)} failed out of {len(results)} total.")

    # Update fetch log (append new entries, replace if same fasta_file exists)
    if success:
        log_path = out_dir / "fetch_log.csv"

        # Read existing log entries
        existing = {}
        if log_path.exists():
            with open(log_path, "r", newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    existing[row["fasta_file"]] = row

        # Update with new entries (replaces existing same-name files)
        for r in success:
            fname = f"{r['name']}.fa"
            existing[fname] = {
                "fasta_file": fname,
                "sequences": r["count"],
                "mastersheet": mastersheet_name,
            }

        # Write back
        with open(log_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=["fasta_file", "sequences", "mastersheet"])
            writer.writeheader()
            for row in sorted(existing.values(), key=lambda r: r["fasta_file"]):
                writer.writerow(row)

        print(f"Updated {log_path}")

    # Show stats for fetched files
    if success:
        print(f"\n{'='*60}")
        print("Sequence statistics")
        print("="*60)
        for r in success:
            _fasta_stats(r["path"])


# ---------------------------------------------------------------------------
# monitor helpers
# ---------------------------------------------------------------------------

def _extract_accessions(fasta_path: Path) -> set[str]:
    """Extract accession IDs from FASTA headers (first whitespace-delimited token)."""
    accessions = set()
    with open(fasta_path) as f:
        for line in f:
            if line.startswith(">"):
                acc = line[1:].strip().split()[0]
                accessions.add(acc)
    return accessions


def _filter_fasta_by_accessions(
    source: Path, accessions: set[str], dest: Path
) -> int:
    """Write sequences whose accession is in the set to dest. Returns count."""
    count = 0
    writing = False
    with open(source) as fin, open(dest, "w") as fout:
        for line in fin:
            if line.startswith(">"):
                acc = line[1:].strip().split()[0]
                writing = acc in accessions
                if writing:
                    count += 1
            if writing:
                fout.write(line)
    return count


def _deduplicate_fasta(fasta_path: Path) -> int:
    """Remove duplicate sequences (same content, different accessions) in place.

    Keeps the first accession encountered for each unique sequence.
    Returns the number of duplicates removed.
    """
    safe_root = (Path(__file__).resolve().parent.parent.parent / "target_seqs" / "original").resolve()

    candidate = Path(fasta_path)
    candidate_name = candidate.name
    if str(candidate) != candidate_name:
        raise ValueError(f"Refusing non-filename FASTA path: {fasta_path}")

    resolved_fasta_path = (safe_root / candidate_name).resolve()
    try:
        resolved_fasta_path.relative_to(safe_root)
    except ValueError as exc:
        raise ValueError(f"Refusing to access FASTA outside allowed directory: {fasta_path}") from exc

    if resolved_fasta_path.suffix.lower() not in {".fa", ".fasta", ".fna"}:
        raise ValueError(f"Refusing to access non-FASTA file: {fasta_path}")
    if not resolved_fasta_path.exists() or not resolved_fasta_path.is_file():
        raise ValueError(f"FASTA path is not a regular file: {fasta_path}")

    seen_seqs: dict[str, str] = {}  # sequence -> first header
    current_header = None
    current_seq_parts: list[str] = []
    entries: list[tuple[str, str]] = []  # (header_line, sequence)

    with open(resolved_fasta_path) as f:
        for line in f:
            if line.startswith(">"):
                if current_header is not None:
                    seq = "".join(current_seq_parts).upper()
                    entries.append((current_header, seq))
                current_header = line
                current_seq_parts = []
            else:
                current_seq_parts.append(line.strip())
        if current_header is not None:
            seq = "".join(current_seq_parts).upper()
            entries.append((current_header, seq))

    n_before = len(entries)
    unique_entries: list[tuple[str, str]] = []
    for header, seq in entries:
        if seq not in seen_seqs:
            seen_seqs[seq] = header
            unique_entries.append((header, seq))

    n_removed = n_before - len(unique_entries)
    if n_removed > 0:
        with open(resolved_fasta_path, "w") as f:
            for header, seq in unique_entries:
                f.write(header)
                # Write sequence in 80-char lines
                for i in range(0, len(seq), 80):
                    f.write(seq[i:i+80] + "\n")

    return n_removed



def _monitor_fetch(
    data_rows: list[dict],
    work_dir: Path,
    date_dir: Path,
    date_str: str,
    dry_run: bool = False,
) -> dict[str, dict]:
    """Fetch sequences for monitoring. Returns per-target results.

    Fetched FASTA files are stored in date_dir/{target_name}.fa.
    Previous fetches are found by scanning sibling date directories
    under work_dir.

    Result dict per target:
        target_name, fasta_path, metadata_path, new_accessions,
        total_count, new_count, is_first_fetch
    """
    # Group rows by target name
    target_groups: dict[str, list[dict]] = {}
    for row in data_rows:
        name = _make_target_name(row)
        target_groups.setdefault(name, []).append(row)

    results = {}

    _OVERLAP_DAYS = 7  # Fetch overlap to avoid missing late-indexed sequences

    for target_name, rows in target_groups.items():
        # Use the first row to build gget command (same TaxID for same pathogen)
        # Strip max-date columns so monitor always fetches up to the latest
        row = {k: v for k, v in rows[0].items() if k not in _MAX_DATE_COLUMNS}
        pathogen = str(row.get("Pathogen", "")).strip()
        print(f"\n{'='*50}")
        print(f"Fetching: {pathogen} ({target_name})")

        dated_fasta = date_dir / f"{target_name}.fa"
        dated_meta = date_dir / f"{target_name}_metadata.csv"

        # Find previous fetch by scanning sibling date directories for accession lists
        prev_acc_files = sorted(
            [d / f"{target_name}_accessions.txt"
             for d in work_dir.iterdir()
             if d.is_dir() and d.name != date_str
             and (d / f"{target_name}_accessions.txt").exists()],
            key=lambda p: p.parent.name,
            reverse=True,
        )

        is_first = len(prev_acc_files) == 0

        # For subsequent fetches, override min_release_date and
        # min_collection_date to (prev_fetch_date - overlap) so we only
        # download recent sequences instead of the entire history.
        if not is_first:
            prev_date_str = prev_acc_files[0].parent.name  # directory name is YYYYMMDD
            try:
                prev_date = datetime.strptime(prev_date_str, "%Y%m%d")
                cutoff = (prev_date - timedelta(days=_OVERLAP_DAYS)).strftime("%Y-%m-%d")
                for date_col in ("min_release_date", "min_collection_date"):
                    orig = str(row.get(date_col, "")).strip()
                    # Use whichever is later: spreadsheet value or cutoff
                    if not orig or orig < cutoff:
                        row[date_col] = cutoff
                print(f"  Narrowing fetch window: min dates → {cutoff} (prev fetch: {prev_date_str})")
            except ValueError:
                pass  # Could not parse date from directory name; use original dates

        if dry_run:
            cmd = _build_gget_command(row, str(date_dir / ".gget_tmp"))
            print(f"  Command: {' '.join(cmd)}")
            print(f"  Output: {dated_fasta}")
            results[target_name] = {
                "target_name": target_name,
                "status": "dry-run",
            }
            continue

        # Run gget
        gget_tmp = date_dir / ".gget_tmp"
        gget_tmp.mkdir(parents=True, exist_ok=True)

        cmd = _build_gget_command(row, str(gget_tmp))
        if not cmd:
            print(f"  Skipping (no TaxID)")
            continue

        try:
            subprocess.run(cmd, check=True)
        except subprocess.CalledProcessError as e:
            print(f"  Error: gget failed with exit code {e.returncode}")
            results[target_name] = {"target_name": target_name, "status": "error"}
            shutil.rmtree(gget_tmp, ignore_errors=True)
            continue

        # Move FASTA output
        fasta_files = list(gget_tmp.glob("*.fa")) + list(gget_tmp.glob("*.fasta"))
        meta_files = list(gget_tmp.glob("*_metadata.csv"))

        if not fasta_files:
            print(f"  Warning: no FASTA output")
            results[target_name] = {"target_name": target_name, "status": "no_output"}
            shutil.rmtree(gget_tmp, ignore_errors=True)
            continue

        shutil.move(str(fasta_files[0]), str(dated_fasta))
        if meta_files:
            shutil.move(str(meta_files[0]), str(dated_meta))
        shutil.rmtree(gget_tmp, ignore_errors=True)

        # Remove duplicate sequences (same content, different accessions)
        n_deduped = _deduplicate_fasta(dated_fasta)
        if n_deduped > 0:
            print(f"  Removed {n_deduped} duplicate sequence(s)")

        # Count fetched sequences and diff against previous
        current_acc = _extract_accessions(dated_fasta)
        total_count = len(current_acc)
        print(f"  Fetched {total_count} unique sequences → {dated_fasta.name}")

        if not is_first:
            prev_acc = set(prev_acc_files[0].read_text().strip().split("\n"))
            new_acc = current_acc - prev_acc
            print(f"  Previous: {prev_acc_files[0].parent.name} ({len(prev_acc)} accessions)")
            print(f"  New sequences: {len(new_acc)}")
        else:
            new_acc = current_acc
            print(f"  First fetch — all {total_count} sequences are new")

        # Write new-only FASTA
        new_fasta = None
        if new_acc:
            new_fasta = date_dir / f"{target_name}_new.fa"
            n_written = _filter_fasta_by_accessions(dated_fasta, new_acc, new_fasta)
            print(f"  New sequences FASTA: {new_fasta.name} ({n_written} seqs)")

        # Save accession list for future comparison, then delete full FASTA
        acc_file = date_dir / f"{target_name}_accessions.txt"
        acc_file.write_text("\n".join(sorted(current_acc)) + "\n")
        dated_fasta.unlink()
        print(f"  Saved {len(current_acc)} accessions to {acc_file.name} (full FASTA removed)")

        results[target_name] = {
            "target_name": target_name,
            "status": "success",
            "fasta_path": new_fasta,
            "new_fasta_path": new_fasta,
            "metadata_path": dated_meta if dated_meta.exists() else None,
            "new_accessions": new_acc,
            "total_count": total_count,
            "new_count": len(new_acc),
            "is_first_fetch": is_first,
        }

    return results


def _build_pset_fa(rows: list[dict], output_path: Path) -> list[str]:
    """Build pset FASTA from spreadsheet rows with Forward/Reverse/Probe.

    Returns list of primer pair IDs written.
    """
    pair_ids = []
    with open(output_path, "w") as f:
        for row in rows:
            fwd = str(row.get("Forward", "")).strip()
            rev = str(row.get("Reverse", "")).strip()
            pro = str(row.get("Probe", "")).strip()
            qid = str(row.get("query_id", "")).strip()

            if not fwd or not rev:
                continue

            primer_name = str(row.get("Primer name", "")).replace(" ", "").strip()
            pair_id = primer_name or (f"query{qid}" if qid else f"pair_{len(pair_ids)+1}")

            f.write(f">{pair_id}_for\n{fwd}\n")
            f.write(f">{pair_id}_rev\n{rev}\n")
            if pro:
                f.write(f">{pair_id}_pro\n{pro}\n")

            pair_ids.append(pair_id)

    return pair_ids


def _get_new_seq_table(
    new_accessions: set[str], metadata_path: Path | None
) -> str:
    """Build a text table of new sequences with metadata."""
    if not new_accessions:
        return "No new sequences.\n"

    if not metadata_path or not metadata_path.exists():
        return f"{len(new_accessions)} new sequences (metadata unavailable).\n"

    import pandas as pd
    meta = pd.read_csv(metadata_path)

    # Filter to new accessions
    meta = meta[meta["accession"].isin(new_accessions)]

    if meta.empty:
        return f"{len(new_accessions)} new sequences (not found in metadata).\n"

    # Select display columns
    cols = ["accession", "Length", "Geographic Region", "Release date"]
    available = [c for c in cols if c in meta.columns]
    display = meta[available].fillna("—")

    lines = [f"{len(new_accessions)} new sequence(s):\n"]

    # Header
    header = "  ".join(f"{c:<20}" for c in available)
    lines.append(header)
    lines.append("-" * len(header))

    for _, row in display.iterrows():
        line = "  ".join(f"{str(row[c]):<20}" for c in available)
        lines.append(line)

    return "\n".join(lines) + "\n"


def _read_excel_summary(xlsx_path: Path) -> str:
    """Extract summary sheet content from an evaluate Excel file as text."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True)
    if "summary" not in wb.sheetnames:
        return "(no summary sheet)\n"

    ws = wb["summary"]
    lines = []
    for row in ws.iter_rows(values_only=True):
        vals = [str(v) if v is not None else "" for v in row]
        if any(vals):
            lines.append("  ".join(vals))

    wb.close()
    return "\n".join(lines) + "\n" if lines else "(empty summary)\n"


def _monitor_evaluate(
    target_name: str,
    target_fasta: Path,
    pset_fa: Path,
    params_file: Path,
    date_dir: Path,
    cores: int,
    reuse: bool = False,
) -> Path | None:
    """Run evaluate for a target with a pset. Returns output dir or None."""
    eval_dir = date_dir / target_name
    if not reuse and eval_dir.exists():
        shutil.rmtree(eval_dir)
    eval_dir.mkdir(parents=True, exist_ok=True)

    # Copy inputs into eval dir
    shutil.copy2(params_file, eval_dir / "params.txt")
    shutil.copy2(pset_fa, eval_dir / pset_fa.name)

    # Create target_seqs/original with the target FASTA
    target_seq_dir = eval_dir / "target_seqs" / "original"
    target_seq_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(target_fasta, target_seq_dir / f"{target_name}.fa")

    # Write params.txt with TARGETS
    # Append TARGETS to copied params if not already present
    params_path = eval_dir / "params.txt"
    params_text = params_path.read_text()
    if "TARGETS" not in params_text:
        with open(params_path, "a") as f:
            f.write(f"\nTARGETS = {target_name}\n")
    else:
        import re as _re
        params_text = _re.sub(
            r"^TARGETS\s*=.*$",
            f"TARGETS = {target_name}",
            params_text,
            flags=_re.MULTILINE,
        )
        params_path.write_text(params_text)

    # Generate Snakefile
    _generate_snakefile(eval_dir, [target_name], [], [], [])

    # Run snakemake
    config_args = ["evaluate=1", f"pset={pset_fa.name}"]

    print(f"\n  Running evaluate for {target_name}...")
    rc = subprocess.run(
        ["snakemake", "-s", "Snakefile", "--cores", str(cores),
         "--config"] + config_args,
        cwd=str(eval_dir),
    ).returncode

    if rc != 0:
        print(f"  Evaluate failed for {target_name} (exit code {rc})")
        return None

    # Find output Excel files
    xlsx_files = list(eval_dir.rglob("*.xlsx"))
    if xlsx_files:
        print(f"  Evaluate complete: {len(xlsx_files)} report(s) generated")
        return eval_dir
    else:
        print(f"  Warning: no Excel output found")
        return None


def _send_email(
    sender: str,
    password: str,
    recipients: list[str],
    subject: str,
    body: str,
    attachments: list[Path] | None = None,
) -> bool:
    """Send email via Gmail SMTP with optional attachments."""
    msg = EmailMessage()
    msg["From"] = sender
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = subject
    msg.set_content(body)

    if attachments:
        for path in attachments:
            with open(path, "rb") as f:
                data = f.read()
            msg.add_attachment(
                data,
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=path.name,
            )

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender, password)
            server.send_message(msg)
        print(f"Email sent to {', '.join(recipients)}")
        return True
    except Exception as e:
        print(f"Error sending email: {e}", file=sys.stderr)
        return False


_CRON_SCHEDULES = {
    "biweekly": ("0 0 1,15 * *", "1st and 15th of every month"),
    "monthly": ("0 0 1 * *", "1st of every month"),
    "quarterly": ("0 0 1 1,4,7,10 *", "1st of Jan, Apr, Jul, Oct"),
}


def _install_cron(work_dir: Path, params_file: Path, schedule_email: str,
                  frequency: str = "monthly"):
    """Install a cron job for adapt monitor."""
    # Build the command to run
    adapt_path = shutil.which("adapt")
    if not adapt_path:
        print("Error: 'adapt' not found in PATH. Install the package first.", file=sys.stderr)
        sys.exit(1)

    cmd = (
        f"cd {work_dir.resolve()} && "
        f"{adapt_path} monitor --params {params_file.resolve()} "
        f"--workdir {work_dir.resolve()}"
    )
    if schedule_email:
        cmd += f" --email {schedule_email}"

    schedule, description = _CRON_SCHEDULES.get(frequency, _CRON_SCHEDULES["monthly"])
    cron_line = f"{schedule} {cmd}  # adapt-monitor"

    # Read existing crontab
    try:
        result = subprocess.run(
            ["crontab", "-l"], capture_output=True, text=True
        )
        existing = result.stdout if result.returncode == 0 else ""
    except FileNotFoundError:
        existing = ""

    # Remove any existing adapt-monitor line
    lines = [l for l in existing.strip().split("\n") if "adapt-monitor" not in l]
    lines.append(cron_line)

    # Install new crontab
    new_cron = "\n".join(lines) + "\n"
    proc = subprocess.run(
        ["crontab", "-"], input=new_cron, text=True
    )
    if proc.returncode == 0:
        print(f"Cron job installed ({description} at midnight)")
        print(f"  {cron_line}")
    else:
        print("Error installing cron job", file=sys.stderr)
        sys.exit(1)


def _uninstall_cron():
    """Remove the adapt monitor cron job."""
    try:
        result = subprocess.run(
            ["crontab", "-l"], capture_output=True, text=True
        )
        if result.returncode != 0:
            print("No crontab found.")
            return
        existing = result.stdout
    except FileNotFoundError:
        print("crontab not available.")
        return

    lines = [l for l in existing.strip().split("\n") if "adapt-monitor" not in l]
    new_cron = "\n".join(lines) + "\n" if lines else ""

    subprocess.run(["crontab", "-"], input=new_cron, text=True)
    print("Cron job removed.")


def cmd_monitor(args):
    """Run the monitoring pipeline: fetch → evaluate → email alert."""
    if args.unschedule:
        _uninstall_cron()
        return

    if args.schedule:
        _install_cron(
            work_dir=Path(args.workdir),
            params_file=Path(args.params),
            schedule_email=args.email,
        )
        return

    # --- Main monitor flow ---
    params_file = Path(args.params)
    if not params_file.exists():
        print(f"Error: params file '{params_file}' not found.", file=sys.stderr)
        sys.exit(1)

    params = parse_params(params_file)
    base_dir = Path(args.workdir)
    date_str = datetime.now().strftime("%Y%m%d")
    cores = args.cores

    # Email config
    email_sender = str(params.get("EMAIL_SENDER", "")).strip()
    email_password = str(params.get("EMAIL_PASSWORD", "")).strip()
    email_recipients = [
        e.strip() for e in str(params.get("EMAIL_RECIPIENTS", "")).split(",")
        if e.strip()
    ]
    if args.email:
        email_recipients = [args.email]

    if args.skip_fetch:
        # --skip-fetch: reuse existing date directory, skip spreadsheet/fetch entirely
        runid = args.runid
        if not runid:
            print("Error: --runid is required with --skip-fetch.", file=sys.stderr)
            sys.exit(1)
        work_dir = base_dir / runid
        date_dir = work_dir / date_str
        if not date_dir.exists():
            print(f"Error: date directory '{date_dir}' not found.", file=sys.stderr)
            sys.exit(1)

        print(f"Run ID: {runid}")
        print(f"Skipping fetch (--skip-fetch). Using existing files in {date_dir}")

        # Discover targets from existing _new.fa files
        fetch_results = {}
        for new_fa in sorted(date_dir.glob("*_new.fa")):
            target_name = new_fa.name.removesuffix("_new.fa")
            meta = date_dir / f"{target_name}_metadata.csv"
            with open(new_fa) as f:
                new_acc = {l[1:].strip().split()[0] for l in f if l.startswith(">")}
            fetch_results[target_name] = {
                "target_name": target_name,
                "status": "success",
                "new_fasta_path": new_fa,
                "metadata_path": meta if meta.exists() else None,
                "new_accessions": new_acc,
                "total_count": len(new_acc),
                "new_count": len(new_acc),
            }
            print(f"  {target_name}: {new_fa.name} ({len(new_acc)} seqs)")

        if not fetch_results:
            print("No *_new.fa files found in date directory.")
            return

        target_groups = {t: [] for t in fetch_results}
    else:
        # Normal flow: download spreadsheet and fetch
        spreadsheet_url = args.url or str(params.get("SPREADSHEET_URL", "")).strip()
        if not spreadsheet_url:
            print("Error: no spreadsheet URL.", file=sys.stderr)
            sys.exit(1)

        spreadsheet_id = _extract_spreadsheet_id(spreadsheet_url)
        print(f"Downloading spreadsheet...")
        csv_text = _download_spreadsheet_csv(spreadsheet_id)
        _, data_rows = _load_spreadsheet(csv_text)

        # Filter by query IDs if specified
        if args.query_ids:
            query_ids = {q.strip() for q in args.query_ids}
            data_rows = [r for r in data_rows if str(r.get("query_id", "")).strip() in query_ids]

        if not data_rows:
            print("No rows to process.")
            return

        # Resolve run ID: explicit > derived from spreadsheet name
        runid = args.runid or spreadsheet_id[:8]
        work_dir = base_dir / runid
        work_dir.mkdir(parents=True, exist_ok=True)
        print(f"Run ID: {runid}")

        # Create date directory for this run's outputs
        date_dir = work_dir / date_str
        date_dir.mkdir(parents=True, exist_ok=True)

        # Save spreadsheet snapshot
        csv_path = date_dir / "mastersheet.csv"
        csv_path.write_text(csv_text, encoding="utf-8")

        # Group rows by target name
        target_groups: dict[str, list[dict]] = {}
        for row in data_rows:
            name = _make_target_name(row)
            target_groups.setdefault(name, []).append(row)

        print(f"Monitoring {len(target_groups)} target(s): {', '.join(target_groups.keys())}")

        if not args.dry_run and not shutil.which("gget"):
            print("Error: 'gget' not installed.", file=sys.stderr)
            sys.exit(1)

        fetch_results = _monitor_fetch(data_rows, work_dir, date_dir, date_str, args.dry_run)

        if args.dry_run:
            print("\nDry-run complete.")
            return

    # Step 2: Build pset.fa and evaluate for each target with new sequences
    email_body_parts = []
    email_body_parts.append(f"ADAPT Monitor Report — {date_str}\n{'='*50}\n")

    all_xlsx: list[Path] = []

    for target_name, rows in target_groups.items():
        result = fetch_results.get(target_name, {})
        status = result.get("status")

        email_body_parts.append(f"\n== {target_name} ==")

        if status != "success":
            email_body_parts.append(f"Fetch failed (status: {status})\n")
            continue

        new_count = result["new_count"]
        total_count = result["total_count"]
        new_acc = result["new_accessions"]
        metadata_path = result.get("metadata_path")

        email_body_parts.append(
            f"{new_count} new sequence(s) detected ({total_count} total)\n"
        )

        # New sequence table
        seq_table = _get_new_seq_table(new_acc, metadata_path)
        email_body_parts.append(seq_table)

        if new_count == 0:
            continue

        # Determine target FASTA for evaluate
        target_fasta = result.get("new_fasta_path")
        if not target_fasta:
            continue

        # Use existing pset.fa or build from spreadsheet rows
        pset_fa = date_dir / f"{target_name}_pset.fa"

        if not pset_fa.exists():
            primer_rows = [
                r for r in rows
                if not _is_empty(r.get("Forward")) and not _is_empty(r.get("Reverse"))
            ]

            if not primer_rows:
                email_body_parts.append("No primer sequences — skipping evaluate.\n")
                continue

            pair_ids = _build_pset_fa(primer_rows, pset_fa)
            print(f"\n  Built pset.fa with {len(pair_ids)} primer set(s): {', '.join(pair_ids)}")
        else:
            print(f"\n  Using existing {pset_fa.name}")

        # Run evaluate
        eval_dir = _monitor_evaluate(
            target_name=target_name,
            target_fasta=target_fasta,
            pset_fa=pset_fa,
            params_file=params_file,
            date_dir=date_dir,
            cores=cores,
            reuse=args.skip_fetch,
        )

        if eval_dir:
            # Move Excel files to date_dir with {target}_{primer} naming
            xlsx_files = sorted(eval_dir.rglob("*.xlsx"))
            for xlsx in xlsx_files:
                dest = date_dir / f"{target_name}_{xlsx.name}"
                shutil.move(str(xlsx), str(dest))
                all_xlsx.append(dest)

            # TODO: re-enable cleanup after debugging
            # shutil.rmtree(eval_dir, ignore_errors=True)

            email_body_parts.append("Evaluation results:")
            for xlsx in all_xlsx[-len(xlsx_files):]:
                email_body_parts.append(f"\n  --- {xlsx.stem} ---")
                summary_text = _read_excel_summary(xlsx)
                email_body_parts.append(summary_text)
        else:
            email_body_parts.append("Evaluate failed.\n")

    # Step 3: Send email
    email_body = "\n".join(email_body_parts)

    if email_sender and email_password and email_recipients:
        subject = f"[ADAPT Monitor] {date_str} — {len(target_groups)} target(s)"
        _send_email(
            sender=email_sender,
            password=email_password,
            recipients=email_recipients,
            subject=subject,
            body=email_body,
            attachments=all_xlsx if all_xlsx else None,
        )
    else:
        if not email_sender:
            print("\nEmail not configured (set EMAIL_SENDER, EMAIL_PASSWORD, "
                  "EMAIL_RECIPIENTS in params.txt)")

    print("\nMonitor complete.")


def main():
    """Main entry point for the adapt CLI."""
    parser = argparse.ArgumentParser(
        prog="adapt",
        description="ADAPT PCR primer design pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Singleplex design
  adapt design --params params.txt

  # Design with probes
  adapt design --params params.txt --probe

  # Multiplex design with probes
  adapt design --params params.txt --multiplex --probe

  # Evaluate primer sequences
  adapt evaluate --params params.txt --for ATCGATCG --rev GCTAGCTA

  # Evaluate from primer set file
  adapt evaluate --params params.txt --pset my_primers.fa

  # Custom run ID and core count
  adapt design --params params.txt --runid my_experiment --cores 8

  # Fetch virus sequences from NCBI
  adapt fetch --params params.txt

  # Fetch specific queries only
  adapt fetch --params params.txt --query-ids 1 2 3

  # Fetch with explicit spreadsheet URL
  adapt fetch --url 'https://docs.google.com/spreadsheets/d/...' --outdir target_seqs/original

  # Monitor: fetch new sequences, evaluate, send email
  adapt monitor --params params.txt

  # Monitor dry-run
  adapt monitor --params params.txt --dry-run

  # Install monthly cron job
  adapt monitor --params params.txt --schedule --email user@gmail.com

  # Remove cron job
  adapt monitor --unschedule
""",
    )
    parser.add_argument("--version", action="version", version="%(prog)s 0.1.0")

    subparsers = parser.add_subparsers(
        dest="command",
        title="commands",
        metavar="<command>",
    )

    # --- design subcommand ---
    p_design = subparsers.add_parser(
        "design",
        help="Design primers for target sequences",
        description="Run the full primer design pipeline.",
    )
    p_design.add_argument("--params", default="params.txt", help="Parameters file (default: params.txt)")
    p_design.add_argument("--probe", action="store_true", help="Enable probe design")
    p_design.add_argument("--multiplex", action="store_true", help="Enable multiplex mode")
    p_design.add_argument("--runid", help="Run ID for output directory (default: timestamp)")
    p_design.add_argument("--cores", type=int, default=os.cpu_count() or 1, help="Number of cores (default: all)")
    p_design.add_argument("--dry-run", action="store_true", help="Show what would be done without executing")
    p_design.set_defaults(func=cmd_design)

    # --- evaluate subcommand ---
    p_eval = subparsers.add_parser(
        "evaluate",
        help="Evaluate existing primer sequences",
        description="Evaluate user-provided primer sequences against targets.",
    )
    p_eval.add_argument("--params", default="params.txt", help="Parameters file (default: params.txt)")
    p_eval.add_argument("--runid", help="Run ID for output directory (default: timestamp)")
    p_eval.add_argument("--cores", type=int, default=os.cpu_count() or 1, help="Number of cores (default: all)")
    p_eval.add_argument("--dry-run", action="store_true", help="Show what would be done without executing")

    # Primer input: either --pset or --for/--rev/--pro
    eval_pset = p_eval.add_argument_group(
        "option 1: primer set file",
    )
    eval_pset.add_argument("--pset", help="FASTA file with primer set (*_for and *_rev entries)")

    eval_seq = p_eval.add_argument_group(
        "option 2: individual primer sequences",
    )
    eval_seq.add_argument("--for", dest="forward", help="Forward primer sequence")
    eval_seq.add_argument("--rev", dest="reverse", help="Reverse primer sequence")
    eval_seq.add_argument("--pro", dest="probe", help="Probe sequence (optional)")

    p_eval.set_defaults(func=cmd_evaluate)

    # --- fetch subcommand ---
    p_fetch = subparsers.add_parser(
        "fetch",
        help="Fetch virus sequences from NCBI via gget",
        description="Download virus sequences from NCBI using gget virus, "
                    "configured by a Google Sheets spreadsheet.",
    )
    p_fetch.add_argument("--params", default="params.txt", help="Parameters file (default: params.txt)")
    p_fetch.add_argument("--url", help="Google Sheets URL (overrides SPREADSHEET_URL in params.txt)")
    p_fetch.add_argument("--query-ids", nargs="+", dest="query_ids",
                         help="Query IDs to fetch (default: all, or QUERY_IDS from params.txt)")
    p_fetch.add_argument("--outdir", default="target_seqs/original",
                         help="Output directory for FASTA files (default: target_seqs/original)")
    p_fetch.add_argument("--dry-run", action="store_true", help="Show commands without executing")
    p_fetch.set_defaults(func=cmd_fetch)

    # --- monitor subcommand ---
    p_monitor = subparsers.add_parser(
        "monitor",
        help="Monitor primer performance against new sequences",
        description="Periodically fetch new virus sequences, evaluate primers, "
                    "and send email alerts with results.",
    )
    p_monitor.add_argument("--params", default="params.txt", help="Parameters file (default: params.txt)")
    p_monitor.add_argument("--url", help="Google Sheets URL (overrides SPREADSHEET_URL in params.txt)")
    p_monitor.add_argument("--query-ids", nargs="+", dest="query_ids",
                           help="Query IDs to monitor (default: all)")
    p_monitor.add_argument("--workdir", default="monitor",
                           help="Working directory for monitor data (default: monitor/)")
    p_monitor.add_argument("--runid", help="Run ID to group results (default: derived from spreadsheet name)")
    p_monitor.add_argument("--email", help="Recipient email (overrides EMAIL_RECIPIENTS in params.txt)")
    p_monitor.add_argument("--cores", type=int, default=os.cpu_count() or 1,
                           help="Number of cores (default: all)")
    p_monitor.add_argument("--skip-fetch", action="store_true",
                           help="Skip fetch step and use existing files in date directory")
    p_monitor.add_argument("--dry-run", action="store_true", help="Show what would be done")
    p_monitor.add_argument("--schedule", action="store_true",
                           help="Install monthly cron job")
    p_monitor.add_argument("--unschedule", action="store_true",
                           help="Remove monthly cron job")
    p_monitor.set_defaults(func=cmd_monitor)

    args = parser.parse_args()

    if args.command is None:
        parser.print_help()
        sys.exit(1)

    # Validate evaluate primer input: --pset or --for/--rev (not both, not neither)
    if args.command == "evaluate":
        has_pset = args.pset is not None
        has_seq = args.forward is not None
        if has_pset and has_seq:
            p_eval.error("use either --pset or --for/--rev, not both")
        if not has_pset and not has_seq:
            p_eval.error("provide either --pset or --for/--rev")
        if has_seq and not args.reverse:
            p_eval.error("--for requires --rev")

    args.func(args)


if __name__ == "__main__":
    main()
