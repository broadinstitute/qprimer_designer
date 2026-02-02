#!/usr/bin/env python3

import argparse
from pathlib import Path
import sys
from collections import defaultdict
import pandas as pd
import ast
import subprocess
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def just_complement(seq):
    comp = str.maketrans("ACGTacgt", "TGCAtgca")
    return seq.translate(comp)

def get_dg(seq1, seq2, RNAduplex):
    inp = f'{seq1}\n{seq2}\n'
    res = subprocess.run([RNAduplex, '--noconv', '--paramFile=DNA'], input=inp,
                         capture_output=True, text=True, check=True)
    deltaG = re.search(r'\(([-+]?\d*\.?\d+)\)', res.stdout)
    if not deltaG:
        raise ValueError(f'Could not parse ΔG from RNAduplex output:\n{res.stdout}')
    return float(deltaG.group(1))  

def filter_by_primer_id(df, pid, eval_type="on"):
    pid_for = f"{pid}_for"
    pid_rev = f"{pid}_rev"

    if eval_type == "on":
        mask = (
            (df["pname_f"] == pid_for) &
            (df["pname_r"] == pid_rev)
        )

    elif eval_type == "off":
        valid = {pid_for, pid_rev}
        mask = (
            df["pname_f"].isin(valid) &
            df["pname_r"].isin(valid)
        )

    else:
        raise ValueError(f"Unknown eval_type: {eval_type}")

    return df.loc[mask].copy()

def build_alignment_string(df, side):
    """
    side: 'f' or 'r'
    """

    if side == "f":
        return (
            "1    " + df["pseq_f"] + df["len_f"].astype(str).str.rjust(5) + "\n" +
            "     " + df["match_f"] + "     \n" +
            df["starts"].astype(str).str.ljust(5) + df["tseq_f"].apply(just_complement) +
            (df["starts"] + df["tseq_f"].str.replace("-","").str.len() - 1).astype(str).str.rjust(5)
        )
    else:
        return (
            "1    " + df["pseq_r"] + df["len_r"].astype(str).str.rjust(5) + "\n" +
            "     " + df["match_r"] + "     \n" +
            (df["starts"]+df["prod_len"]-df["tseq_r"].str.replace("-",'').str.len()).astype(str).str.ljust(5) + 
            df["tseq_r"].apply(just_complement) + (df["starts"]+df["prod_len"]-1).astype(str).str.rjust(5)
        )

def eval_to_target_df(eval_path, pid, eval_type="on"):
    """
    eval_type:
        - 'on'  : use all rows
        - 'off' : keep only rows with classifier > 0.5
    """
    df = pd.read_csv(eval_path+".full")

    df = filter_by_primer_id(df, pid, eval_type)

    if df.empty:
        raise ValueError(f"No rows found for primer set {pid} in {eval_path}")
    
    df["targets"] = df["targets"].apply(ast.literal_eval)
    df["starts"] = df["starts"].apply(ast.literal_eval)
    df = df.explode(["targets","starts"])

    df = df.set_index("targets")
    df.index.name = "seq_id"

    if eval_type == "off":
        df = df[df["classifier"] > 0.5].copy()

        if df.empty:
            return pd.DataFrame()

    df["align_f"] = build_alignment_string(df, "f")
    df["align_r"] = build_alignment_string(df, "r")

    df["classifier"] = (df["classifier"] > 0.5).astype(int)

    df = df.drop(columns=[
        "match_f", "tseq_f",
        "match_r", "tseq_r",
    ])

    return df

def build_dimerization_table(df, viennaPath):
    """
    Returns a 2x2 DataFrame:
        index   : forward / reverse
        columns : forward / reverse
    """

    df_on = df[df["eval_type"] == "on"]
    rnadup = f'{viennaPath}/src/bin/RNAduplex'

    fseq = df_on["pseq_f"].values[0].replace("-","")
    rseq = df_on["pseq_r"].values[0].replace("-","")
    return pd.DataFrame(
        [[get_dg(fseq,fseq,rnadup), get_dg(fseq,rseq,rnadup)],
         [get_dg(rseq,fseq,rnadup), get_dg(rseq,rseq,rnadup)]],
        index=["forward", "reverse"],
        columns=["forward", "reverse"]
    ).round(1)

def build_sensitivity_table(df):
    df_on = df[df["eval_type"] == "on"]

    if df_on.empty:
        return pd.DataFrame()

    act = df_on["regressor"]

    return pd.DataFrame(
        {
            "Coverage": [df_on["classifier"].sum()],
            "Act_mean": [act.mean()],
            "Act_median": [act.median()],
            "Act_min": [act.min()],
            "Act_max": [act.max()],
        },
        index=[df["target"].values[0]]
    )

def build_specificity_table(df):
    df_off = df[df["eval_type"] == "off"]

    if df_off.empty:
        return pd.DataFrame()

    return (
        df_off
        .groupby("target")
        .agg(
            Coverage=("classifier", "sum"),
            Act_mean=("regressor", "mean"),
            Act_median=("regressor", "median"),
            Act_min=("regressor", "min"),
            Act_max=("regressor", "max"),
        )
    )

def save_eval_excel(df, outdir, filename, viennaPath):
    """
    Save merged on/off target dataframe to Excel,
    applying Courier font to alignment columns.
    """
    outdir = Path(outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    xlsx_path = outdir / filename

    # ------------------
    # Write main sheet
    # ------------------
    df.drop(columns=["pseq_f","pseq_r"]).to_excel(xlsx_path, sheet_name="detail")

    wb = load_workbook(xlsx_path)

    # ------------------
    # Format evaluation sheet (existing code)
    # ------------------
    ws = wb["detail"]

    courier = Font(name="Courier New")
    wrap = Alignment(wrap_text=True)

    align_cols = [
        idx for idx, cell in enumerate(ws[1], start=1)
        if "align" in str(cell.value)
    ]

    for col_idx in align_cols:
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 40
        for row in ws.iter_rows(min_row=2):
            cell = row[col_idx - 1]
            if cell.value:
                cell.font = courier
                cell.alignment = wrap

    # ------------------
    # Build summary tables
    # ------------------
    ws_sum = wb.create_sheet("summary")
    wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(ws_sum)))

    dimer_df = build_dimerization_table(df, viennaPath)
    sens_df  = build_sensitivity_table(df)
    spec_df  = build_specificity_table(df)

    current_row = 1

    def write_table(title, table):
        nonlocal current_row

        ws_sum.cell(row=current_row, column=1, value=title).font = Font(bold=True)
        current_row += 1

        if table.empty:
            ws_sum.cell(row=current_row, column=1, value="(no data)")
            current_row += 2
            return

        for col_idx, col in enumerate([""] + list(table.columns), start=1):
            ws_sum.cell(row=current_row, column=col_idx, value=col).font = Font(bold=True)
        current_row += 1

        for idx, row in table.iterrows():
            ws_sum.cell(row=current_row, column=1, value=idx)
            for col_idx, val in enumerate(row, start=2):
                ws_sum.cell(row=current_row, column=col_idx, value=val)
            current_row += 1

        current_row += 2  # spacer

    write_table("Dimerization", dimer_df)
    write_table("Sensitivity", sens_df)
    write_table("Specificity", spec_df)

    wb.save(xlsx_path)

def parse_args():
    parser = argparse.ArgumentParser(
        description="Evaluate primer sets using on/off-target alignments"
    )

    parser.add_argument(
        "--on", required=True,
        help="On-target eval file"
    )

    parser.add_argument(
        "--off", nargs="+", required=True,
        help="Off-target eval files"
    )

    parser.add_argument(
        "--fa", required=True,
        help="FASTA file containing primer sets (*_for / *_rev)"
    )

    parser.add_argument(
        "--out", required=True,
        help="Output directory"
    )

    parser.add_argument(
        "--names", nargs="+", required=True,
        help="primer set IDs"
    )
 
    parser.add_argument(
        "--params", required=True,
        help="Evaluation parameter string"
    )

    parser.add_argument(
        "--program", required=True,
        help="Path to ViennaRNA (or other backend)"
    )

    return parser.parse_args()

def get_target_name(path):
    return Path(path).name.split(".")[1]

def main():
    args = parse_args()

    outdir = Path(args.out)
    outdir.mkdir(parents=True, exist_ok=True)

    print(f"Output directory: {outdir}")
    print(f"Primer sets to evaluate: {', '.join(args.names)}")
    
    cols = [ "target", "eval_type", "classifier", "regressor", 
             "pname_f", "pname_r","pseq_f","pseq_r",
             "align_f", "align_r",
             "prod_len", "prod_Tm",
             "mm_f", "indel_f", "len_f", "Tm_f", "GC_f",
             "mm_r", "indel_r", "len_r", "Tm_r", "GC_r" ]

    for pid in args.names:
        dfs = []

        # ------------------
        # on-target eval
        # ------------------
        try:
            df_on = eval_to_target_df(
                eval_path=args.on,
                pid=pid,
                eval_type="on"
            )
            df_on["eval_type"] = "on"
            df_on["target"] = get_target_name(args.on)
            dfs.append(df_on[cols].sort_values("regressor", ascending=False))
        except Exception as e:
            print(f"[WARNING] on-target eval failed for {pid}: {e}")

        # ------------------
        # off-target evals
        # ------------------
        for off_path in args.off:
            try:
                df_off = eval_to_target_df(
                    eval_path=off_path,
                    pid=pid,
                    eval_type="off"
                )

                if not df_off.empty:
                    df_off["eval_type"] = "off"
                    df_off["target"] = get_target_name(off_path)
                    dfs.append(df_off[cols].sort_values("regressor", ascending=False))

            except Exception as e:
                print(f"[WARNING] off-target eval failed for {pid} ({off_path}): {e}")

        # ------------------
        # concat & save
        # ------------------
        if not dfs:
            print(f"[WARNING] No evaluation results for primer set {pid}, skipping")
            continue

        df_all = pd.concat(dfs, axis=0)

        xlsx_path = outdir / f"{pid}.xlsx"
        save_eval_excel(df_all, outdir=outdir, filename=f"{pid}.xlsx", viennaPath=args.program)

    print("All primer sets processed")
    (outdir / "done").touch()

if __name__ == "__main__":
    main()

