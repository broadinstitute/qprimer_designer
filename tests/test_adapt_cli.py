"""Tests for adapt_cli helper functions."""

import csv
import io
import textwrap
from pathlib import Path
from unittest.mock import patch

import pytest

from qprimer_designer.adapt_cli import (
    _is_empty,
    _parse_boolean,
    _extract_spreadsheet_id,
    _load_spreadsheet,
    _make_target_name,
    _build_gget_command,
    _build_pset_fa,
    _extract_accessions,
    _filter_fasta_by_accessions,
    _deduplicate_fasta,
    _get_new_seq_table,
    _read_excel_summary,
)


# ---------------------------------------------------------------------------
# _is_empty
# ---------------------------------------------------------------------------
class TestIsEmpty:
    def test_none(self):
        assert _is_empty(None) is True

    def test_empty_string(self):
        assert _is_empty("") is True
        assert _is_empty("   ") is True

    def test_na_variants(self):
        assert _is_empty("N/A") is True
        assert _is_empty("na") is True
        assert _is_empty("NaN") is True
        assert _is_empty("none") is True
        assert _is_empty("None") is True

    def test_nonempty(self):
        assert _is_empty("hello") is False
        assert _is_empty("0") is False
        assert _is_empty(123) is False

    def test_numeric_zero(self):
        assert _is_empty(0) is False


# ---------------------------------------------------------------------------
# _parse_boolean
# ---------------------------------------------------------------------------
class TestParseBoolean:
    def test_true_variants(self):
        assert _parse_boolean("TRUE") == "true"
        assert _parse_boolean("true") == "true"
        assert _parse_boolean("True") == "true"
        assert _parse_boolean("  TRUE  ") == "true"

    def test_false_variants(self):
        assert _parse_boolean("FALSE") == "false"
        assert _parse_boolean("false") == "false"
        assert _parse_boolean("False") == "false"

    def test_empty_returns_none(self):
        assert _parse_boolean("") is None
        assert _parse_boolean(None) is None
        assert _parse_boolean("N/A") is None

    def test_invalid_returns_none(self):
        assert _parse_boolean("yes") is None
        assert _parse_boolean("1") is None
        assert _parse_boolean("maybe") is None


# ---------------------------------------------------------------------------
# _extract_spreadsheet_id
# ---------------------------------------------------------------------------
class TestExtractSpreadsheetId:
    def test_standard_url(self):
        url = "https://docs.google.com/spreadsheets/d/1aBcDeFgHiJkLmNoPqRsTuVwXyZ/edit#gid=0"
        assert _extract_spreadsheet_id(url) == "1aBcDeFgHiJkLmNoPqRsTuVwXyZ"

    def test_url_with_copy(self):
        url = "https://docs.google.com/spreadsheets/d/ABC123_-x/copy"
        assert _extract_spreadsheet_id(url) == "ABC123_-x"

    def test_url_export(self):
        url = "https://docs.google.com/spreadsheets/d/SHEET_ID/export?format=csv"
        assert _extract_spreadsheet_id(url) == "SHEET_ID"

    def test_invalid_url_exits(self):
        with pytest.raises(SystemExit):
            _extract_spreadsheet_id("https://example.com/not-a-sheet")

    def test_empty_url_exits(self):
        with pytest.raises(SystemExit):
            _extract_spreadsheet_id("")


# ---------------------------------------------------------------------------
# _load_spreadsheet
# ---------------------------------------------------------------------------
class TestLoadSpreadsheet:
    @staticmethod
    def _make_csv(*data_rows, headers=None):
        """Build a 4+ row CSV: category row, headers, description row, data rows."""
        if headers is None:
            headers = ["Pathogen", "Target name", "TaxID", "nuc_completeness"]
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow([""] * len(headers))  # row 0: category
        w.writerow(headers)               # row 1: headers
        w.writerow(["desc"] * len(headers))  # row 2: descriptions
        for row in data_rows:
            w.writerow(row)
        return buf.getvalue()

    def test_basic_parsing(self):
        csv_text = self._make_csv(
            ["SARS-CoV-2", "SC2", "2697049", "complete"],
        )
        headers, data = _load_spreadsheet(csv_text)
        assert "TaxID" in headers
        assert len(data) == 1
        assert data[0]["TaxID"] == "2697049"
        assert data[0]["Target name"] == "SC2"

    def test_skips_empty_taxid(self):
        csv_text = self._make_csv(
            ["SARS-CoV-2", "SC2", "", "complete"],
            ["Zika", "ZIKV", "64320", "complete"],
        )
        _, data = _load_spreadsheet(csv_text)
        assert len(data) == 1
        assert data[0]["TaxID"] == "64320"

    def test_skips_na_taxid(self):
        csv_text = self._make_csv(
            ["Flu", "FLU", "N/A", "complete"],
        )
        _, data = _load_spreadsheet(csv_text)
        assert len(data) == 0

    def test_too_few_rows_exits(self):
        csv_text = "a,b\n1,2\n"
        with pytest.raises(SystemExit):
            _load_spreadsheet(csv_text)

    def test_missing_taxid_column_exits(self):
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["", ""])
        w.writerow(["Pathogen", "Segment"])  # no TaxID
        w.writerow(["desc", "desc"])
        w.writerow(["HIV", "gag"])
        with pytest.raises(SystemExit):
            _load_spreadsheet(buf.getvalue())

    def test_multiple_rows(self):
        csv_text = self._make_csv(
            ["P1", "T1", "111", "complete"],
            ["P2", "T2", "222", "partial"],
            ["P3", "T3", "333", ""],
        )
        _, data = _load_spreadsheet(csv_text)
        assert len(data) == 3
        assert [d["TaxID"] for d in data] == ["111", "222", "333"]

    def test_metadata_columns_included(self):
        headers = ["Pathogen", "Target name", "Primer name", "TaxID", "nuc_completeness"]
        csv_text = self._make_csv(
            ["HIV", "HIV_target", "P1", "12721", "complete"],
            headers=headers,
        )
        _, data = _load_spreadsheet(csv_text)
        assert data[0]["Primer name"] == "P1"
        assert data[0]["Pathogen"] == "HIV"


# ---------------------------------------------------------------------------
# _make_target_name
# ---------------------------------------------------------------------------
class TestMakeTargetName:
    def test_uses_target_name_if_present(self):
        row = {"Target name": "SC2_Omicron", "Pathogen_abb": "SC2"}
        assert _make_target_name(row) == "SC2_Omicron"

    def test_falls_back_to_pathogen_abb(self):
        row = {"Target name": "", "Pathogen_abb": "ZIKV"}
        assert _make_target_name(row) == "ZIKV"

    def test_appends_segment(self):
        row = {"Target name": "", "Pathogen_abb": "IAV", "Gene Segment": "HA"}
        assert _make_target_name(row) == "IAV_HA"

    def test_ignores_non_segmented(self):
        row = {"Target name": "", "Pathogen_abb": "SC2", "Gene Segment": "Non-segmented"}
        assert _make_target_name(row) == "SC2"

    def test_falls_back_to_pathogen(self):
        row = {"Pathogen": "Dengue virus"}
        assert _make_target_name(row) == "Dengue_virus"

    def test_empty_everything(self):
        row = {}
        assert _make_target_name(row) == "unknown"

    def test_whitespace_target_name(self):
        row = {"Target name": "   ", "Pathogen_abb": "HIV"}
        assert _make_target_name(row) == "HIV"


# ---------------------------------------------------------------------------
# _build_gget_command
# ---------------------------------------------------------------------------
class TestBuildGgetCommand:
    def test_basic_command(self):
        row = {"TaxID": "2697049"}
        cmd = _build_gget_command(row, "/tmp/out")
        assert cmd[:3] == ["gget", "virus", "2697049"]
        assert "--out" in cmd
        assert "/tmp/out" in cmd

    def test_empty_taxid_returns_none(self):
        row = {"TaxID": ""}
        assert _build_gget_command(row, "/tmp/out") is None

    def test_none_taxid_returns_none(self):
        row = {"TaxID": None}
        assert _build_gget_command(row, "/tmp/out") is None

    def test_na_taxid_returns_none(self):
        row = {"TaxID": "N/A"}
        assert _build_gget_command(row, "/tmp/out") is None

    def test_string_option(self):
        row = {"TaxID": "123", "nuc_completeness": "complete"}
        cmd = _build_gget_command(row, "/tmp/out")
        idx = cmd.index("--nuc_completeness")
        assert cmd[idx + 1] == "complete"

    def test_boolean_flag_true(self):
        row = {"TaxID": "123", "refseq_only": "TRUE"}
        cmd = _build_gget_command(row, "/tmp/out")
        assert "--refseq_only" in cmd

    def test_boolean_flag_false(self):
        row = {"TaxID": "123", "refseq_only": "FALSE"}
        cmd = _build_gget_command(row, "/tmp/out")
        assert "--refseq_only" not in cmd

    def test_true_false_flag(self):
        row = {"TaxID": "123", "annotated": "TRUE"}
        cmd = _build_gget_command(row, "/tmp/out")
        idx = cmd.index("--annotated")
        assert cmd[idx + 1] == "true"

    def test_metadata_columns_excluded(self):
        row = {"TaxID": "123", "Pathogen": "HIV", "Primer name": "P1", "Forward": "ATCG"}
        cmd = _build_gget_command(row, "/tmp/out")
        assert "--Pathogen" not in cmd
        assert "--Forward" not in cmd

    def test_skips_empty_values(self):
        row = {"TaxID": "123", "nuc_completeness": "", "segment": "N/A"}
        cmd = _build_gget_command(row, "/tmp/out")
        assert "--nuc_completeness" not in cmd
        assert "--segment" not in cmd


# ---------------------------------------------------------------------------
# _build_pset_fa
# ---------------------------------------------------------------------------
class TestBuildPsetFa:
    def test_basic_output(self, tmp_path):
        output = tmp_path / "pset.fa"
        rows = [
            {"Primer name": "N1", "Forward": "ATCGATCG", "Reverse": "GCTAGCTA", "Probe": "TTTTAAAA"},
        ]
        pair_ids = _build_pset_fa(rows, output)
        assert pair_ids == ["N1"]
        content = output.read_text()
        assert ">N1_for\nATCGATCG\n" in content
        assert ">N1_rev\nGCTAGCTA\n" in content
        assert ">N1_pro\nTTTTAAAA\n" in content

    def test_no_probe(self, tmp_path):
        output = tmp_path / "pset.fa"
        rows = [
            {"Primer name": "P1", "Forward": "AAAA", "Reverse": "CCCC", "Probe": ""},
        ]
        _build_pset_fa(rows, output)
        content = output.read_text()
        assert "_pro" not in content

    def test_skips_missing_forward_or_reverse(self, tmp_path):
        output = tmp_path / "pset.fa"
        rows = [
            {"Primer name": "P1", "Forward": "AAAA", "Reverse": ""},
            {"Primer name": "P2", "Forward": "", "Reverse": "CCCC"},
        ]
        pair_ids = _build_pset_fa(rows, output)
        assert pair_ids == []
        assert output.read_text() == ""

    def test_multiple_pairs(self, tmp_path):
        output = tmp_path / "pset.fa"
        rows = [
            {"Primer name": "A", "Forward": "AAAA", "Reverse": "TTTT"},
            {"Primer name": "B", "Forward": "CCCC", "Reverse": "GGGG"},
        ]
        pair_ids = _build_pset_fa(rows, output)
        assert pair_ids == ["A", "B"]

    def test_spaces_removed_from_name(self, tmp_path):
        output = tmp_path / "pset.fa"
        rows = [
            {"Primer name": "nCoV N1", "Forward": "AAAA", "Reverse": "TTTT"},
        ]
        pair_ids = _build_pset_fa(rows, output)
        assert pair_ids == ["nCoVN1"]
        assert ">nCoVN1_for" in output.read_text()

    def test_fallback_pair_id_from_query_id(self, tmp_path):
        output = tmp_path / "pset.fa"
        rows = [
            {"Forward": "AAAA", "Reverse": "TTTT", "query_id": "42"},
        ]
        pair_ids = _build_pset_fa(rows, output)
        assert pair_ids == ["query42"]

    def test_fallback_pair_id_auto(self, tmp_path):
        output = tmp_path / "pset.fa"
        rows = [
            {"Forward": "AAAA", "Reverse": "TTTT"},
        ]
        pair_ids = _build_pset_fa(rows, output)
        assert pair_ids == ["pair_1"]

    def test_empty_rows(self, tmp_path):
        output = tmp_path / "pset.fa"
        pair_ids = _build_pset_fa([], output)
        assert pair_ids == []
        assert output.read_text() == ""


# ---------------------------------------------------------------------------
# _extract_accessions
# ---------------------------------------------------------------------------
class TestExtractAccessions:
    def test_basic(self, tmp_path):
        fasta = tmp_path / "seqs.fa"
        fasta.write_text(">ACC001 some description\nATCG\n>ACC002\nGCTA\n")
        result = _extract_accessions(fasta)
        assert result == {"ACC001", "ACC002"}

    def test_empty_file(self, tmp_path):
        fasta = tmp_path / "empty.fa"
        fasta.write_text("")
        assert _extract_accessions(fasta) == set()

    def test_multiline_sequence(self, tmp_path):
        fasta = tmp_path / "multi.fa"
        fasta.write_text(">SEQ1\nATCG\nGCTA\n>SEQ2\nAAAA\n")
        result = _extract_accessions(fasta)
        assert result == {"SEQ1", "SEQ2"}

    def test_accession_with_version(self, tmp_path):
        fasta = tmp_path / "ver.fa"
        fasta.write_text(">NC_045512.2 SARS-CoV-2\nATCG\n")
        result = _extract_accessions(fasta)
        assert result == {"NC_045512.2"}


# ---------------------------------------------------------------------------
# _filter_fasta_by_accessions
# ---------------------------------------------------------------------------
class TestFilterFastaByAccessions:
    def test_filter_keeps_matching(self, tmp_path):
        source = tmp_path / "all.fa"
        dest = tmp_path / "filtered.fa"
        source.write_text(
            ">ACC1 desc\nATCG\n>ACC2 desc\nGCTA\n>ACC3 desc\nTTTT\n"
        )
        count = _filter_fasta_by_accessions(source, {"ACC1", "ACC3"}, dest)
        assert count == 2
        content = dest.read_text()
        assert ">ACC1" in content
        assert ">ACC3" in content
        assert ">ACC2" not in content

    def test_no_match(self, tmp_path):
        source = tmp_path / "all.fa"
        dest = tmp_path / "filtered.fa"
        source.write_text(">ACC1\nATCG\n")
        count = _filter_fasta_by_accessions(source, {"MISSING"}, dest)
        assert count == 0
        assert dest.read_text() == ""

    def test_empty_accession_set(self, tmp_path):
        source = tmp_path / "all.fa"
        dest = tmp_path / "filtered.fa"
        source.write_text(">ACC1\nATCG\n")
        count = _filter_fasta_by_accessions(source, set(), dest)
        assert count == 0

    def test_multiline_sequence(self, tmp_path):
        source = tmp_path / "all.fa"
        dest = tmp_path / "filtered.fa"
        source.write_text(">ACC1\nATCG\nGCTA\nAAAA\n>ACC2\nTTTT\n")
        count = _filter_fasta_by_accessions(source, {"ACC1"}, dest)
        assert count == 1
        content = dest.read_text()
        assert "ATCG\n" in content
        assert "GCTA\n" in content
        assert "TTTT" not in content


# ---------------------------------------------------------------------------
# _deduplicate_fasta
# ---------------------------------------------------------------------------
class TestDeduplicateFasta:
    def test_rejects_absolute_path(self):
        with pytest.raises(ValueError, match="non-filename"):
            _deduplicate_fasta("/etc/passwd")

    def test_rejects_path_traversal(self):
        with pytest.raises(ValueError, match="non-filename"):
            _deduplicate_fasta("../../etc/passwd")

    def test_rejects_invalid_extension(self):
        with pytest.raises(ValueError, match="invalid"):
            _deduplicate_fasta("file.txt")

    def test_rejects_empty_name(self):
        with pytest.raises(ValueError, match="invalid|unsafe"):
            _deduplicate_fasta("")

    def test_rejects_dotdot(self):
        with pytest.raises(ValueError, match="non-filename|unsafe|invalid"):
            _deduplicate_fasta("..")

    def test_rejects_shell_chars(self):
        with pytest.raises(ValueError, match="unsafe|invalid"):
            _deduplicate_fasta("file;rm -rf.fa")

    def test_nonexistent_file(self):
        with pytest.raises(ValueError, match="not a regular file"):
            _deduplicate_fasta("nonexistent_file_xyz.fa")


# ---------------------------------------------------------------------------
# _get_new_seq_table
# ---------------------------------------------------------------------------
class TestGetNewSeqTable:
    def test_empty_accessions(self, tmp_path):
        result = _get_new_seq_table(set(), None)
        assert "No new sequences" in result

    def test_no_metadata_path(self):
        result = _get_new_seq_table({"ACC1", "ACC2"}, None)
        assert "2 new sequences" in result
        assert "unavailable" in result

    def test_missing_metadata_file(self, tmp_path):
        result = _get_new_seq_table({"ACC1"}, tmp_path / "missing.csv")
        assert "1 new sequences" in result
        assert "unavailable" in result

    def test_with_metadata(self, tmp_path):
        import pandas as pd

        meta_path = tmp_path / "metadata.csv"
        df = pd.DataFrame({
            "accession": ["ACC1", "ACC2", "ACC3"],
            "Length": [1000, 2000, 3000],
            "Geographic Region": ["USA", "UK", "Japan"],
            "Release date": ["2025-01-01", "2025-02-01", "2025-03-01"],
        })
        df.to_csv(meta_path, index=False)

        result = _get_new_seq_table({"ACC1", "ACC3"}, meta_path)
        assert "2 new sequence(s)" in result
        assert "ACC1" in result
        assert "ACC3" in result
        assert "ACC2" not in result

    def test_accessions_not_in_metadata(self, tmp_path):
        import pandas as pd

        meta_path = tmp_path / "metadata.csv"
        df = pd.DataFrame({"accession": ["OTHER"], "Length": [100]})
        df.to_csv(meta_path, index=False)

        result = _get_new_seq_table({"ACC1"}, meta_path)
        assert "not found in metadata" in result


# ---------------------------------------------------------------------------
# _read_excel_summary
# ---------------------------------------------------------------------------
class TestReadExcelSummary:
    def test_with_summary_sheet(self, tmp_path):
        from openpyxl import Workbook

        xlsx = tmp_path / "report.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "summary"
        ws.append(["Primer", "Coverage", "Score"])
        ws.append(["N1", "95%", "0.92"])
        wb.save(xlsx)

        result = _read_excel_summary(xlsx)
        assert "Primer" in result
        assert "N1" in result
        assert "95%" in result

    def test_no_summary_sheet(self, tmp_path):
        from openpyxl import Workbook

        xlsx = tmp_path / "report.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "details"
        ws.append(["data"])
        wb.save(xlsx)

        result = _read_excel_summary(xlsx)
        assert "no summary sheet" in result

    def test_empty_summary_sheet(self, tmp_path):
        from openpyxl import Workbook

        xlsx = tmp_path / "report.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "summary"
        wb.save(xlsx)

        result = _read_excel_summary(xlsx)
        assert isinstance(result, str)
